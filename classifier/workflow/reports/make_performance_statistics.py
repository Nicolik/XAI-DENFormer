"""Generate prediction-level performance statistics and supplementary workbooks."""

from __future__ import annotations

try:
    from ._bootstrap import PROJECT_ROOT  # noqa: F401
except ImportError:
    from _bootstrap import PROJECT_ROOT  # noqa: F401

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import paths
from classifier.workflow.reports.performance_statistics import (
    AnalysisConfig,
    DEFAULT_MODELS,
    DEFAULT_PROTOCOLS,
    AnalysisOutputs,
    run_analysis,
)


DEFAULT_OUTPUT_DIR = Path(paths.logs_dir) / "aggregate_metrics" / "performance_statistics"
DEFAULT_CDHIT_SPLIT = Path(paths.split_files["cdhit"])
DEFAULT_CONTINENT_SPLIT = Path(paths.split_files["continent"])
DEFAULT_TIMEBIN_SPLIT = Path(paths.split_files["timebin"])

DEFAULT_PROTOCOL_SPLIT_FILES = {
    "ohe_cdhit_e100": DEFAULT_CDHIT_SPLIT,
    "ohe_continent_e100": DEFAULT_CONTINENT_SPLIT,
    "ohe_timebin_e100": DEFAULT_TIMEBIN_SPLIT,
}


README_ROWS = [
    ("Purpose", "Paired statistical comparison of model performance using held-out out-of-fold predictions."),
    ("Why not fold-level tests?", "The geographical, temporal, and CD-HIT evaluations contain only a small number of structured outer partitions. Kruskal-Wallis, Wilcoxon, t tests, or Friedman tests on those few partition-level scores are not used for inferential claims."),
    ("Primary endpoint", "Difference in accuracy between two models evaluated on exactly the same held-out genomes."),
    ("Primary inference", "Paired sign-flip permutation at the CD-HIT cluster level, with Holm correction within the prespecified benchmark and ablation comparison families."),
    ("Confidence intervals", "Paired nonparametric bootstrap that resamples complete CD-HIT clusters with replacement, stratified by true serotype."),
    ("Sensitivity analysis", "Exact two-sided McNemar test on paired genome-level correctness indicators."),
    ("Secondary endpoints", "Macro-F1 and per-serotype F1 differences with cluster-bootstrap 95% confidence intervals."),
    ("Inference scope", "The analysis quantifies uncertainty in the available held-out predictions conditional on the fixed trained models; it does not estimate variability across repeated training runs."),
    ("Interpretation", "A statistically supported difference is protocol-specific. Do not describe a model as universally superior; also report the absolute effect in percentage points and its confidence interval."),
    ("Workbook contents", "Benchmark_primary and Ablation_primary provide reader-facing primary accuracy summaries. Accuracy_full contains all primary-test fields. Macro_F1 and Per_class_F1 contain secondary endpoints. Model_summary contains pooled out-of-fold metrics. QC_alignment and Methods_parameters document validation and analysis settings."),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare DENFormer and competing models using paired held-out predictions, "
            "CD-HIT cluster bootstrap intervals, cluster-level permutation tests, and exact McNemar tests."
        )
    )
    parser.add_argument("--logs-dir", type=Path, default=Path(paths.logs_dir))
    parser.add_argument("--cdhit-split-file", type=Path, default=DEFAULT_CDHIT_SPLIT)
    parser.add_argument("--continent-split-file", type=Path, default=DEFAULT_CONTINENT_SPLIT)
    parser.add_argument("--timebin-split-file", type=Path, default=DEFAULT_TIMEBIN_SPLIT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--bootstrap-reps", type=int, default=10_000)
    parser.add_argument("--permutation-reps", type=int, default=100_000)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--confidence-level", type=float, default=0.95)
    parser.add_argument("--bootstrap-batch-size", type=int, default=128)
    parser.add_argument("--permutation-batch-size", type=int, default=512)
    parser.add_argument("--protocols", nargs="+", default=list(DEFAULT_PROTOCOLS))
    parser.add_argument("--models", nargs="+", default=list(DEFAULT_MODELS))
    return parser.parse_args()


def _format_p_value(value: object) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if pd.isna(number):
        return ""
    if number < 0.001:
        return "<0.001"
    return f"{number:.3f}"


def build_primary_summary_table(primary: pd.DataFrame) -> pd.DataFrame:
    table = primary.copy()
    table["Accuracy A (%)"] = 100.0 * table["accuracy_a"]
    table["Accuracy B (%)"] = 100.0 * table["accuracy_b"]
    table["Delta accuracy (pp)"] = table["delta_accuracy_pp"]
    table["95% cluster-bootstrap CI (pp)"] = table.apply(
        lambda row: f"{row['delta_accuracy_ci_low_pp']:.3f} to {row['delta_accuracy_ci_high_pp']:.3f}",
        axis=1,
    )
    table["Cluster permutation Holm p"] = table["cluster_permutation_p_holm"].map(_format_p_value)
    table["Exact McNemar Holm p"] = table["mcnemar_exact_p_holm"].map(_format_p_value)
    table["Discordant pairs A+/B-"] = table["a_correct_b_wrong"]
    table["Discordant pairs A-/B+"] = table["a_wrong_b_correct"]
    return table[
        [
            "family",
            "protocol_display",
            "model_a_display",
            "model_b_display",
            "n_genomes",
            "n_cdhit_clusters",
            "Accuracy A (%)",
            "Accuracy B (%)",
            "Delta accuracy (pp)",
            "95% cluster-bootstrap CI (pp)",
            "Cluster permutation Holm p",
            "Exact McNemar Holm p",
            "Discordant pairs A+/B-",
            "Discordant pairs A-/B+",
            "interpretation",
        ]
    ].rename(
        columns={
            "family": "Comparison family",
            "protocol_display": "Validation protocol",
            "model_a_display": "Model A",
            "model_b_display": "Model B",
            "n_genomes": "Held-out genomes",
            "n_cdhit_clusters": "CD-HIT clusters",
            "interpretation": "Interpretation",
        }
    )


def _write_workbook(outputs: AnalysisOutputs, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    readme_df = pd.DataFrame(README_ROWS, columns=["Topic", "Description"])
    primary_summary = build_primary_summary_table(outputs.primary_accuracy)
    benchmark = primary_summary[primary_summary["Comparison family"].eq("benchmark")].copy()
    ablation = primary_summary[primary_summary["Comparison family"].eq("ablation")].copy()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False)
        benchmark.to_excel(writer, sheet_name="Benchmark_primary", index=False)
        ablation.to_excel(writer, sheet_name="Ablation_primary", index=False)
        outputs.primary_accuracy.to_excel(writer, sheet_name="Accuracy_full", index=False)
        outputs.macro_f1.to_excel(writer, sheet_name="Macro_F1", index=False)
        outputs.per_class_f1.to_excel(writer, sheet_name="Per_class_F1", index=False)
        outputs.model_summary.to_excel(writer, sheet_name="Model_summary", index=False)
        outputs.qc_alignment.to_excel(writer, sheet_name="QC_alignment", index=False)
        outputs.parameters.to_excel(writer, sheet_name="Methods_parameters", index=False)

    _style_workbook(out_path)


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E1F2")
    bottom_border = Border(bottom=thin_gray)

    percent_columns = {
        "Accuracy_full": {
            "accuracy_a", "accuracy_b", "delta_accuracy", "delta_accuracy_ci_low",
            "delta_accuracy_ci_high", "accuracy_a_ci_low", "accuracy_a_ci_high",
            "accuracy_b_ci_low", "accuracy_b_ci_high",
        },
        "Macro_F1": {
            "macro_f1_a", "macro_f1_b", "delta_macro_f1",
            "delta_macro_f1_ci_low", "delta_macro_f1_ci_high",
        },
        "Per_class_F1": {
            "f1_a", "f1_b", "delta_f1", "delta_f1_ci_low", "delta_f1_ci_high",
        },
        "Model_summary": {
            "accuracy", "macro_f1", "f1_denv1", "f1_denv2", "f1_denv3", "f1_denv4",
        },
    }

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = {cell.value: cell.column for cell in ws[1]}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = bottom_border

        for name in percent_columns.get(ws.title, set()):
            col = headers.get(name)
            if col is None:
                continue
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col).number_format = "0.0000%"

        for header in ("Accuracy A (%)", "Accuracy B (%)", "Delta accuracy (pp)"):
            col = headers.get(header)
            if col is None:
                continue
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col).number_format = "0.000"

        for col_idx in range(1, ws.max_column + 1):
            values = [ws.cell(row_idx, col_idx).value for row_idx in range(1, min(ws.max_row, 200) + 1)]
            max_len = max((len(str(value)) for value in values if value is not None), default=8)
            width = min(max(max_len + 2, 12), 42)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    if "README" in wb.sheetnames:
        readme = wb["README"]
        readme.column_dimensions["A"].width = 25
        readme.column_dimensions["B"].width = 115
        for row_idx in range(2, readme.max_row + 1):
            readme.cell(row_idx, 1).fill = section_fill
            readme.cell(row_idx, 1).font = Font(bold=True)
            readme.row_dimensions[row_idx].height = 48

    wb.save(path)


def _write_csv_outputs(outputs: AnalysisOutputs, out_dir: Path) -> None:
    csv_dir = out_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    outputs.primary_accuracy.to_csv(csv_dir / "primary_accuracy.csv", index=False)
    outputs.macro_f1.to_csv(csv_dir / "macro_f1.csv", index=False)
    outputs.per_class_f1.to_csv(csv_dir / "per_class_f1.csv", index=False)
    outputs.model_summary.to_csv(csv_dir / "model_summary.csv", index=False)
    outputs.qc_alignment.to_csv(csv_dir / "qc_alignment.csv", index=False)
    outputs.parameters.to_csv(csv_dir / "methods_parameters.csv", index=False)


def main() -> None:
    args = parse_args()
    config = AnalysisConfig(
        bootstrap_reps=args.bootstrap_reps,
        permutation_reps=args.permutation_reps,
        seed=args.seed,
        confidence_level=args.confidence_level,
        bootstrap_batch_size=args.bootstrap_batch_size,
        permutation_batch_size=args.permutation_batch_size,
    )

    print(f"Logs directory: {args.logs_dir}")
    print(f"CD-HIT split file: {args.cdhit_split_file}")
    print(f"Continent split file: {args.continent_split_file}")
    print(f"Time-bin split file: {args.timebin_split_file}")
    print(f"Output directory: {args.output_dir}")
    print(f"Bootstrap replicates: {config.bootstrap_reps}")
    print(f"Permutation replicates: {config.permutation_reps}")

    protocol_split_files = {
        "ohe_cdhit_e100": args.cdhit_split_file,
        "ohe_continent_e100": args.continent_split_file,
        "ohe_timebin_e100": args.timebin_split_file,
    }

    outputs = run_analysis(
        logs_dir=args.logs_dir,
        cdhit_split_file=args.cdhit_split_file,
        config=config,
        protocols=args.protocols,
        models=args.models,
        protocol_split_files=protocol_split_files,
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = args.output_dir / "Supplementary_File_S3_Model_Performance_Statistics.xlsx"
    _write_workbook(outputs, workbook_path)
    _write_csv_outputs(outputs, args.output_dir)

    print(f"Saved: {workbook_path}")
    print(f"Saved CSV tables under: {args.output_dir / 'csv'}")


if __name__ == "__main__":
    main()

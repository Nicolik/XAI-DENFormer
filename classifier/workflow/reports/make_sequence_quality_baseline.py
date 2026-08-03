"""Sequence-length and sequence-quality baselines for DENV serotype prediction.

This report intentionally avoids nucleotide-order information. It extracts simple
metadata-like features from the original FASTA records and evaluates inexpensive
classifiers under the exact geographical, temporal, and CD-HIT split definitions
used by the main study.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from collections import defaultdict, deque
from typing import Iterable

import numpy as np
import pandas as pd
from Bio import SeqIO
from sklearn.compose import ColumnTransformer
from sklearn.dummy import DummyClassifier
from sklearn.ensemble import HistGradientBoostingClassifier
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    confusion_matrix,
    f1_score,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

import paths
from dataset.config import FASTA_FILES, SEROTYPES
from classifier.workflow.utils import load_and_validate_folds


RANDOM_STATE = 42
EXCEL_FONT_NAME = "Helvetica"
EXCEL_FONT_SIZE = 8
ACGT = frozenset("ACGT")
IUPAC_AMBIGUOUS = frozenset("RYWSKMBDHV")
GAP_CHARS = frozenset("-.")
LABEL_TO_INDEX = {label: idx for idx, label in enumerate(SEROTYPES)}
INDEX_TO_LABEL = {idx: label for label, idx in LABEL_TO_INDEX.items()}

SUPPLEMENTARY_MODEL_ORDER = (
    "majority",
    "length_logreg",
    "n_logreg",
    "gc_logreg",
    "length_gc_logreg",
    "length_n_gc_logreg",
    "quality_logreg",
    "quality_histgb",
)
SUPPLEMENTARY_MODEL_LABELS = {
    "majority": "Majority class",
    "length_logreg": "Length (LR)",
    "n_logreg": "N fraction (LR)",
    "gc_logreg": "GC fraction (LR)",
    "length_gc_logreg": "Length + GC (LR)",
    "length_n_gc_logreg": "Length + N + GC (LR)",
    "quality_logreg": "All quality features (LR)",
    "quality_histgb": "All quality features (HGB)",
}
SUPPLEMENTARY_PROTOCOL_ORDER = ("continent", "timebin", "cdhit")
SUPPLEMENTARY_PROTOCOL_LABELS = {
    "continent": "Geographical",
    "timebin": "Temporal",
    "cdhit": "CD-HIT",
}


@dataclass(frozen=True)
class ModelSpec:
    name: str
    feature_columns: tuple[str, ...]
    kind: str


MODEL_SPECS = (
    ModelSpec("majority", tuple(), "dummy"),
    ModelSpec("length_logreg", ("sequence_length",), "logreg"),
    ModelSpec("n_logreg", ("n_fraction",), "logreg"),
    ModelSpec("gc_logreg", ("gc_fraction",), "logreg"),
    ModelSpec(
        "length_gc_logreg",
        ("sequence_length", "gc_fraction"),
        "logreg",
    ),
    ModelSpec(
        "length_n_gc_logreg",
        ("sequence_length", "n_fraction", "gc_fraction"),
        "logreg",
    ),
    ModelSpec(
        "quality_logreg",
        (
            "sequence_length",
            "n_fraction",
            "iupac_fraction",
            "gap_fraction",
            "other_fraction",
            "gc_fraction",
        ),
        "logreg",
    ),
    ModelSpec(
        "quality_histgb",
        (
            "sequence_length",
            "n_fraction",
            "iupac_fraction",
            "gap_fraction",
            "other_fraction",
            "gc_fraction",
        ),
        "histgb",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Evaluate length- and sequence-quality-only serotype baselines using "
            "the existing geographical, temporal, and CD-HIT splits."
        )
    )
    parser.add_argument(
        "--genomes-dir",
        type=Path,
        default=paths.genomes_dir,
        help="Directory containing DENV1-DENV4 merged FASTA files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=paths.logs_dir / "aggregate_metrics" / "sequence_quality_baseline",
        help="Directory for CSV, XLSX, and JSON outputs.",
    )
    parser.add_argument(
        "--sequence-id-order-file",
        type=Path,
        default=paths.embeddings_dir / "seq_ids_ohe.txt",
        help=(
            "Canonical sequence-ID order used by dataset_ohe.h5 and the split files. "
            "When present, FASTA-derived features are reindexed to this order."
        ),
    )
    parser.add_argument(
        "--label-order-file",
        type=Path,
        default=paths.embeddings_dir / "label_matrix.txt",
        help=(
            "Optional canonical numeric labels aligned with --sequence-id-order-file. "
            "Used to disambiguate duplicate sequence IDs."
        ),
    )
    parser.add_argument(
        "--split",
        dest="splits",
        action="append",
        choices=tuple(paths.split_files),
        help="Run only one named split protocol. Repeat to select multiple.",
    )
    parser.add_argument(
        "--no-histgb",
        action="store_true",
        help="Skip the non-linear quality-only HistGradientBoosting baseline.",
    )
    return parser.parse_args()


def _safe_fraction(numerator: int, denominator: int) -> float:
    return float(numerator / denominator) if denominator else 0.0


def sequence_features(sequence: str) -> dict[str, float | int]:
    seq = str(sequence).upper()
    length = len(seq)
    counts = {char: seq.count(char) for char in ACGT}
    acgt_count = sum(counts.values())
    n_count = seq.count("N")
    iupac_count = sum(seq.count(char) for char in IUPAC_AMBIGUOUS)
    gap_count = sum(seq.count(char) for char in GAP_CHARS)
    recognized = acgt_count + n_count + iupac_count + gap_count
    other_count = max(0, length - recognized)
    non_acgt_count = length - acgt_count
    gc_count = counts["G"] + counts["C"]

    return {
        "sequence_length": length,
        "a_count": counts["A"],
        "c_count": counts["C"],
        "g_count": counts["G"],
        "t_count": counts["T"],
        "acgt_count": acgt_count,
        "n_count": n_count,
        "iupac_count": iupac_count,
        "gap_count": gap_count,
        "other_count": other_count,
        "non_acgt_count": non_acgt_count,
        "n_fraction": _safe_fraction(n_count, length),
        "iupac_fraction": _safe_fraction(iupac_count, length),
        "gap_fraction": _safe_fraction(gap_count, length),
        "other_fraction": _safe_fraction(other_count, length),
        "non_acgt_fraction": _safe_fraction(non_acgt_count, length),
        "gc_fraction": _safe_fraction(gc_count, acgt_count),
    }


def extract_feature_table(genomes_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    source_index = 0

    for filename, serotype in FASTA_FILES.items():
        fasta_path = Path(genomes_dir) / filename
        if not fasta_path.exists():
            raise FileNotFoundError(f"Missing FASTA file: {fasta_path}")

        for record in SeqIO.parse(str(fasta_path), "fasta"):
            row: dict[str, object] = {
                "source_index": source_index,
                "sequence_id": record.id,
                # Source-file assignment is retained only as provenance. The final
                # target label is replaced from label_matrix.txt after canonical
                # OHE-order alignment, exactly as in training and inference.
                "fasta_serotype": serotype,
                "serotype": serotype,
                "label": LABEL_TO_INDEX[serotype],
            }
            row.update(sequence_features(str(record.seq)))
            rows.append(row)
            source_index += 1

    if not rows:
        raise ValueError(f"No FASTA records found under {genomes_dir}")

    features = pd.DataFrame(rows)
    max_length = int(features["sequence_length"].max())
    features["padding_length"] = max_length - features["sequence_length"]
    features["padding_fraction"] = features["padding_length"] / max_length
    return features


def _read_nonempty_lines(path: Path) -> list[str]:
    return [line.strip() for line in Path(path).read_text(encoding="utf-8").splitlines() if line.strip()]


def align_features_to_canonical_order(
    features: pd.DataFrame,
    sequence_id_order_file: Path,
    label_order_file: Path | None = None,
) -> pd.DataFrame:
    """Align FASTA-derived features to the exact OHE sample and target order.

    ``seq_ids_ohe.txt`` defines the row order of ``dataset_ohe.h5`` and
    ``label_matrix.txt`` defines the target used by the original training and
    inference workflows. Both files are required: this report deliberately does
    not fall back to current FASTA order or to labels parsed from FASTA headers.
    Duplicate sequence IDs are disambiguated with the canonical numeric label.
    """

    sequence_id_order_file = Path(sequence_id_order_file)
    if not sequence_id_order_file.exists():
        raise FileNotFoundError(
            "Canonical OHE sequence-ID order is required for this analysis: "
            f"{sequence_id_order_file}"
        )
    if label_order_file is None or not Path(label_order_file).exists():
        raise FileNotFoundError(
            "Canonical training/inference labels are required for this analysis: "
            f"{label_order_file}"
        )

    label_order_file = Path(label_order_file)
    ordered_ids = _read_nonempty_lines(sequence_id_order_file)
    label_lines = _read_nonempty_lines(label_order_file)

    if len(ordered_ids) != len(features):
        raise ValueError(
            f"Canonical sequence-ID file has {len(ordered_ids):,} rows, but FASTA "
            f"feature extraction produced {len(features):,} rows: {sequence_id_order_file}"
        )
    if len(label_lines) != len(ordered_ids):
        raise ValueError(
            f"Canonical label file has {len(label_lines):,} rows, but sequence-ID "
            f"order has {len(ordered_ids):,} rows: {label_order_file}"
        )

    ordered_labels = [int(value) for value in label_lines]
    invalid = sorted(set(ordered_labels).difference(INDEX_TO_LABEL))
    if invalid:
        raise ValueError(f"Unexpected numeric labels in {label_order_file}: {invalid}")

    # Match on (ID, canonical label), not ID alone, because a small number of
    # accessions/headers can be repeated or contain a conflicting serotype token.
    buckets: dict[tuple[str, int], deque[int]] = defaultdict(deque)
    source = features.reset_index(drop=True)
    for row_position, row in source.iterrows():
        key = (str(row["sequence_id"]), int(row["label"]))
        buckets[key].append(int(row_position))

    selected_positions: list[int] = []
    missing: list[tuple[str, int]] = []
    for sequence_id, label in zip(ordered_ids, ordered_labels):
        key = (sequence_id, label)
        if not buckets[key]:
            missing.append(key)
            if len(missing) >= 5:
                break
        else:
            selected_positions.append(buckets[key].popleft())

    if missing:
        raise ValueError(
            "Could not align canonical OHE rows to the current FASTA collection. "
            f"First missing (sequence_id, label) keys: {missing}"
        )

    leftovers = sum(len(queue) for queue in buckets.values())
    if leftovers:
        raise ValueError(
            f"Canonical alignment left {leftovers:,} unmatched FASTA records; "
            "the FASTA collection and OHE index files are not the same dataset."
        )

    aligned = source.iloc[selected_positions].reset_index(drop=True).copy()
    aligned.insert(0, "index", np.arange(len(aligned), dtype=int))

    # Make label_matrix.txt explicitly authoritative, matching get_dataset().
    expected_labels = np.asarray(ordered_labels, dtype=int)
    aligned["label"] = expected_labels
    aligned["serotype"] = [INDEX_TO_LABEL[int(label)] for label in expected_labels]
    aligned["target_source"] = "label_matrix.txt"

    if aligned["sequence_id"].astype(str).tolist() != ordered_ids:
        raise ValueError("Canonical sequence-ID alignment failed an exact order check")
    if not np.array_equal(aligned["label"].to_numpy(dtype=int), expected_labels):
        raise ValueError("Canonical target assignment failed to reproduce label_matrix.txt")

    return aligned


def folds_from_main_loader(path: Path, n_samples: int) -> list[dict[str, object]]:
    """Load folds exactly as the original training and inference workflows do."""

    return load_and_validate_folds(path, n_samples)


def build_estimator(spec: ModelSpec) -> object:
    if spec.kind == "dummy":
        return DummyClassifier(strategy="most_frequent", random_state=RANDOM_STATE)

    if spec.kind == "logreg":
        preprocessor = ColumnTransformer(
            transformers=[
                (
                    "numeric",
                    Pipeline(
                        steps=[
                            ("imputer", SimpleImputer(strategy="median")),
                            ("scaler", StandardScaler()),
                        ]
                    ),
                    list(spec.feature_columns),
                )
            ],
            remainder="drop",
            verbose_feature_names_out=False,
        )
        classifier = LogisticRegression(
            max_iter=5000,
            solver="lbfgs",
            random_state=RANDOM_STATE,
        )
        return Pipeline([("features", preprocessor), ("classifier", classifier)])

    if spec.kind == "histgb":
        preprocessor = ColumnTransformer(
            transformers=[
                (
                    "numeric",
                    SimpleImputer(strategy="median"),
                    list(spec.feature_columns),
                )
            ],
            remainder="drop",
            verbose_feature_names_out=False,
        )
        classifier = HistGradientBoostingClassifier(
            learning_rate=0.08,
            max_iter=200,
            max_leaf_nodes=15,
            l2_regularization=1.0,
            random_state=RANDOM_STATE,
        )
        return Pipeline([("features", preprocessor), ("classifier", classifier)])

    raise ValueError(f"Unknown model kind: {spec.kind}")


def aligned_probabilities(estimator: object, frame: pd.DataFrame) -> np.ndarray:
    probabilities = np.asarray(estimator.predict_proba(frame), dtype=float)
    classes = np.asarray(estimator.classes_, dtype=int)
    aligned = np.zeros((len(frame), len(SEROTYPES)), dtype=float)
    for source_col, class_index in enumerate(classes):
        aligned[:, int(class_index)] = probabilities[:, source_col]
    return aligned


def metric_row(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float | int]:
    present_labels = sorted(np.unique(y_true).tolist())
    return {
        "n_test": int(len(y_true)),
        "accuracy": float(accuracy_score(y_true, y_pred)),
        "balanced_accuracy": float(balanced_accuracy_score(y_true, y_pred)),
        "macro_f1_present": float(
            f1_score(y_true, y_pred, labels=present_labels, average="macro", zero_division=0)
        ),
        "macro_f1_all4": float(
            f1_score(
                y_true,
                y_pred,
                labels=list(range(len(SEROTYPES))),
                average="macro",
                zero_division=0,
            )
        ),
    }


def coefficient_rows(
    estimator: object,
    spec: ModelSpec,
    protocol: str,
    fold: str,
) -> list[dict[str, object]]:
    if spec.kind != "logreg":
        return []

    classifier = estimator.named_steps["classifier"]
    rows: list[dict[str, object]] = []

    classes = np.asarray(classifier.classes_, dtype=int)
    coefficients = np.asarray(classifier.coef_, dtype=float)
    intercepts = np.asarray(classifier.intercept_, dtype=float)

    if len(classes) == 2 and coefficients.shape[0] == 1:
        class_parameters = [
            (classes[0], -coefficients[0], -intercepts[0]),
            (classes[1], coefficients[0], intercepts[0]),
        ]
    else:
        class_parameters = [
            (class_index, coefficients[position], intercepts[position])
            for position, class_index in enumerate(classes)
        ]

    for class_index, class_coefficients, class_intercept in class_parameters:
        for feature_position, feature_name in enumerate(spec.feature_columns):
            rows.append(
                {
                    "protocol": protocol,
                    "fold": fold,
                    "model": spec.name,
                    "class_index": int(class_index),
                    "serotype": INDEX_TO_LABEL[int(class_index)],
                    "feature": feature_name,
                    "standardized_coefficient": float(class_coefficients[feature_position]),
                    "intercept": float(class_intercept),
                }
            )
    return rows


def evaluate_protocol(
    protocol: str,
    folds: list[dict[str, object]],
    features: pd.DataFrame,
    model_specs: Iterable[ModelSpec],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    fold_metrics: list[dict[str, object]] = []
    predictions: list[pd.DataFrame] = []
    coefficients: list[dict[str, object]] = []

    for fold_spec in folds:
        fold = str(fold_spec["fold_id"])
        train_indices = np.asarray(fold_spec["train_idx"], dtype=np.int64)
        test_indices = np.asarray(fold_spec["test_idx"], dtype=np.int64)
        if len(test_indices) == 0:
            continue

        train = features.iloc[train_indices].copy()
        test = features.iloc[test_indices].copy()
        y_train = train["label"].to_numpy(dtype=int)
        y_test = test["label"].to_numpy(dtype=int)

        for spec in model_specs:
            estimator = build_estimator(spec)
            fit_frame = train if spec.feature_columns else np.zeros((len(train), 1))
            test_frame = test if spec.feature_columns else np.zeros((len(test), 1))
            estimator.fit(fit_frame, y_train)
            y_pred = np.asarray(estimator.predict(test_frame), dtype=int)
            probabilities = aligned_probabilities(estimator, test_frame)

            row: dict[str, object] = {
                "protocol": protocol,
                "fold": str(fold),
                "model": spec.name,
                "features": ",".join(spec.feature_columns) if spec.feature_columns else "none",
                "n_train": int(len(train)),
            }
            row.update(metric_row(y_test, y_pred))
            fold_metrics.append(row)
            coefficients.extend(coefficient_rows(estimator, spec, protocol, str(fold)))

            pred = test[
                [
                    "index",
                    "sequence_id",
                    "serotype",
                    "sequence_length",
                    "padding_fraction",
                    "n_fraction",
                    "non_acgt_fraction",
                    "gc_fraction",
                ]
            ].copy()
            pred.insert(0, "protocol", protocol)
            pred.insert(1, "fold", str(fold))
            pred.insert(2, "model", spec.name)
            pred["true_label"] = y_test
            pred["predicted_label"] = y_pred
            pred["predicted_serotype"] = [INDEX_TO_LABEL[int(x)] for x in y_pred]
            pred["correct"] = y_pred == y_test
            for class_index, serotype in INDEX_TO_LABEL.items():
                pred[f"prob_{serotype}"] = probabilities[:, class_index]
            predictions.append(pred)

    return (
        pd.DataFrame(fold_metrics),
        pd.concat(predictions, ignore_index=True),
        pd.DataFrame(coefficients),
    )


def aggregate_predictions(predictions: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    aggregate_rows: list[dict[str, object]] = []
    per_class_rows: list[dict[str, object]] = []
    confusion_rows: list[dict[str, object]] = []

    for (protocol, model), group in predictions.groupby(["protocol", "model"], sort=False):
        y_true = group["true_label"].to_numpy(dtype=int)
        y_pred = group["predicted_label"].to_numpy(dtype=int)
        row: dict[str, object] = {"protocol": protocol, "model": model}
        row.update(metric_row(y_true, y_pred))
        row["n_unique_test_indices"] = int(group["index"].nunique())
        row["n_prediction_rows"] = int(len(group))
        aggregate_rows.append(row)

        scores = f1_score(
            y_true,
            y_pred,
            labels=list(range(len(SEROTYPES))),
            average=None,
            zero_division=0,
        )
        supports = np.bincount(y_true, minlength=len(SEROTYPES))
        for class_index, serotype in INDEX_TO_LABEL.items():
            per_class_rows.append(
                {
                    "protocol": protocol,
                    "model": model,
                    "serotype": serotype,
                    "support": int(supports[class_index]),
                    "f1": float(scores[class_index]),
                }
            )

        matrix = confusion_matrix(y_true, y_pred, labels=list(range(len(SEROTYPES))))
        for true_index, true_serotype in INDEX_TO_LABEL.items():
            for pred_index, pred_serotype in INDEX_TO_LABEL.items():
                confusion_rows.append(
                    {
                        "protocol": protocol,
                        "model": model,
                        "true_serotype": true_serotype,
                        "predicted_serotype": pred_serotype,
                        "n": int(matrix[true_index, pred_index]),
                    }
                )

    return (
        pd.DataFrame(aggregate_rows),
        pd.DataFrame(per_class_rows),
        pd.DataFrame(confusion_rows),
    )


def split_qc(
    protocol: str,
    folds: list[dict[str, object]],
    n_features: int,
) -> dict[str, object]:
    test_indices = np.concatenate([
        np.asarray(fold["test_idx"], dtype=np.int64)
        for fold in folds
        if len(fold["test_idx"])
    ])
    return {
        "protocol": protocol,
        "target_label_source": "label_matrix.txt",
        "split_loader": "classifier.workflow.utils.load_and_validate_folds",
        "n_folds": int(len(folds)),
        "n_feature_rows": int(n_features),
        "n_unique_test_indices": int(len(np.unique(test_indices))),
        "n_test_assignments": int(len(test_indices)),
    }


def feature_summary(features: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "sequence_length",
        "padding_fraction",
        "n_fraction",
        "iupac_fraction",
        "gap_fraction",
        "other_fraction",
        "non_acgt_fraction",
        "gc_fraction",
    ]
    rows: list[dict[str, object]] = []
    for serotype, group in features.groupby("serotype", sort=False):
        for column in columns:
            values = group[column].astype(float)
            rows.append(
                {
                    "serotype": serotype,
                    "feature": column,
                    "n": int(values.notna().sum()),
                    "mean": float(values.mean()),
                    "std": float(values.std(ddof=1)),
                    "min": float(values.min()),
                    "q25": float(values.quantile(0.25)),
                    "median": float(values.median()),
                    "q75": float(values.quantile(0.75)),
                    "max": float(values.max()),
                }
            )
    return pd.DataFrame(rows)


def build_supplementary_table(aggregate_metrics: pd.DataFrame) -> pd.DataFrame:
    """Build a Word-ready table with one atomic metric value per cell."""

    required_columns = {
        "protocol",
        "model",
        "accuracy",
        "balanced_accuracy",
        "macro_f1_all4",
    }
    missing_columns = required_columns.difference(aggregate_metrics.columns)
    if missing_columns:
        raise ValueError(
            "Aggregate metrics are missing columns required for the supplementary table: "
            f"{sorted(missing_columns)}"
        )

    indexed = aggregate_metrics.set_index(["model", "protocol"])
    metric_rows = (
        ("Accuracy", "accuracy"),
        ("Balanced accuracy", "balanced_accuracy"),
        ("Macro-F1", "macro_f1_all4"),
    )
    rows: list[dict[str, object]] = []

    for model in SUPPLEMENTARY_MODEL_ORDER:
        if model not in SUPPLEMENTARY_MODEL_LABELS:
            continue
        for metric_label, metric_column in metric_rows:
            row: dict[str, object] = {
                "Baseline": SUPPLEMENTARY_MODEL_LABELS[model],
                "Metric": metric_label,
            }
            for protocol in SUPPLEMENTARY_PROTOCOL_ORDER:
                key = (model, protocol)
                if key not in indexed.index:
                    raise ValueError(
                        "Missing aggregate result required for the supplementary table: "
                        f"model={model}, protocol={protocol}"
                    )
                result = indexed.loc[key]
                if isinstance(result, pd.DataFrame):
                    raise ValueError(
                        "Duplicate aggregate result for supplementary-table cell: "
                        f"model={model}, protocol={protocol}"
                    )
                row[SUPPLEMENTARY_PROTOCOL_LABELS[protocol]] = round(
                    100.0 * float(result[metric_column]), 1
                )
            rows.append(row)

    return pd.DataFrame(
        rows,
        columns=[
            "Baseline",
            "Metric",
            *[SUPPLEMENTARY_PROTOCOL_LABELS[p] for p in SUPPLEMENTARY_PROTOCOL_ORDER],
        ],
    )


def apply_excel_font(workbook) -> None:
    """Apply the shared Helvetica 8 typography to every generated workbook."""

    from copy import copy
    from openpyxl.styles import Font

    # Set the workbook-level Normal style for any cells added later.
    normal_style = workbook._named_styles["Normal"]
    normal_style.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE)

    # Pandas/openpyxl may assign explicit fonts to populated cells. Preserve
    # emphasis and colour while normalizing only family and size.
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                font = copy(cell.font)
                font.name = EXCEL_FONT_NAME
                font.sz = EXCEL_FONT_SIZE
                cell.font = font


def write_supplementary_table_workbook(
    output_path: Path,
    aggregate_metrics: pd.DataFrame,
) -> None:
    """Write only the atomic table values, styled like the other appendix tables."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    table = build_supplementary_table(aggregate_metrics)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table_S1"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    group_fill = PatternFill("solid", fgColor="EAF0F8")
    border_side = Side(style="thin", color="B4C7E7")
    table_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    for column_index, column_name in enumerate(table.columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = table_border
    sheet.row_dimensions[1].height = 26

    for row_index, record in enumerate(table.itertuples(index=False), start=2):
        for column_index, value in enumerate(record, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, color="000000")
            cell.border = table_border
            if column_index == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif column_index == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.0"
        sheet.row_dimensions[row_index].height = 15

    # Three metric rows form one baseline block, matching the compact grouped
    # treatment used by the other supplementary tables.
    first_data_row = 2
    rows_per_model = 3
    for model_offset in range(len(SUPPLEMENTARY_MODEL_ORDER)):
        start_row = first_data_row + model_offset * rows_per_model
        end_row = start_row + rows_per_model - 1
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        baseline_cell = sheet.cell(row=start_row, column=1)
        baseline_cell.fill = group_fill
        baseline_cell.font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True, color="000000")
        baseline_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        baseline_cell.border = table_border
        for row_index in range(start_row, end_row + 1):
            sheet.cell(row=row_index, column=1).fill = group_fill
            sheet.cell(row=row_index, column=1).border = table_border

    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15
    sheet.column_dimensions["E"].width = 15
    sheet.freeze_panes = "C2"
    sheet.print_area = f"A1:E{1 + len(table)}"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    apply_excel_font(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_outputs(
    output_dir: Path,
    features: pd.DataFrame,
    fold_metrics: pd.DataFrame,
    predictions: pd.DataFrame,
    coefficients: pd.DataFrame,
    aggregate_metrics: pd.DataFrame,
    per_class_f1: pd.DataFrame,
    confusion: pd.DataFrame,
    qc: pd.DataFrame,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary = feature_summary(features)

    # Remove the obsolete v10 diagnostic, which reflected a legacy descriptive
    # split column rather than the targets used by training/inference.
    obsolete_mismatch_file = output_dir / "split_serotype_mismatches.csv"
    if obsolete_mismatch_file.exists():
        obsolete_mismatch_file.unlink()

    features.to_csv(output_dir / "sequence_quality_features.csv", index=False)
    fold_metrics.to_csv(output_dir / "fold_metrics.csv", index=False)
    aggregate_metrics.to_csv(output_dir / "aggregate_metrics.csv", index=False)
    per_class_f1.to_csv(output_dir / "per_class_f1.csv", index=False)
    confusion.to_csv(output_dir / "confusion_matrices_long.csv", index=False)
    qc.to_csv(output_dir / "split_qc.csv", index=False)
    summary.to_csv(output_dir / "feature_summary_by_serotype.csv", index=False)
    coefficients.to_csv(output_dir / "logistic_coefficients.csv", index=False)
    predictions.to_csv(output_dir / "oof_predictions.csv.gz", index=False, compression="gzip")

    readme_rows = [
        {
            "field": "analysis_scope",
            "value": "Metadata-like sequence features only; no nucleotide order or embeddings.",
        },
        {
            "field": "primary_question",
            "value": "Can serotype be predicted from length and simple sequence-quality/composition features?",
        },
        {"field": "random_state", "value": str(RANDOM_STATE)},
        {
            "field": "target_label_source",
            "value": (
                "label_matrix.txt, exactly matching the numeric targets loaded by the original "
                "training and inference workflows."
            ),
        },
        {
            "field": "sample_alignment_source",
            "value": "seq_ids_ohe.txt, matching the row order of dataset_ohe.h5.",
        },
        {
            "field": "split_loader",
            "value": (
                "classifier.workflow.utils.load_and_validate_folds, the same loader used by "
                "the original training and inference workflows."
            ),
        },
        {
            "field": "models",
            "value": ", ".join(aggregate_metrics["model"].drop_duplicates().tolist()),
        },
        {
            "field": "important_caveat",
            "value": (
                "Low baseline performance shows that these features are insufficient in isolation; "
                "it does not by itself prove that a sequence model never uses padding."
            ),
        },
    ]
    readme = pd.DataFrame(readme_rows)

    workbook = output_dir / "Sequence_Quality_Baseline.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        aggregate_metrics.to_excel(writer, sheet_name="Aggregate_metrics", index=False)
        fold_metrics.to_excel(writer, sheet_name="Fold_metrics", index=False)
        per_class_f1.to_excel(writer, sheet_name="Per_class_F1", index=False)
        qc.to_excel(writer, sheet_name="Split_QC", index=False)
        summary.to_excel(writer, sheet_name="Feature_summary", index=False)
        coefficients.to_excel(writer, sheet_name="LogReg_coefficients", index=False)
        apply_excel_font(writer.book)

    supplementary_workbook = output_dir / "Supplementary_Table_Sequence_Quality_Baselines.xlsx"
    write_supplementary_table_workbook(supplementary_workbook, aggregate_metrics)

    manifest = {
        "output_dir": str(output_dir),
        "n_sequences": int(len(features)),
        "target_label_source": "label_matrix.txt",
        "sample_alignment_source": "seq_ids_ohe.txt",
        "protocols": aggregate_metrics["protocol"].drop_duplicates().tolist(),
        "models": aggregate_metrics["model"].drop_duplicates().tolist(),
        "files": sorted(path.name for path in output_dir.iterdir() if path.is_file()),
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main() -> None:
    args = parse_args()
    selected_protocols = args.splits or list(paths.split_files)
    model_specs = [spec for spec in MODEL_SPECS if not (args.no_histgb and spec.kind == "histgb")]

    print(f"Reading FASTA records from: {args.genomes_dir}")
    features = extract_feature_table(args.genomes_dir)
    print(f"Extracted features for {len(features):,} sequences")

    print(f"Aligning samples to canonical OHE order: {args.sequence_id_order_file}")
    print(f"Loading canonical targets used by training/inference: {args.label_order_file}")
    features = align_features_to_canonical_order(
        features,
        args.sequence_id_order_file,
        args.label_order_file,
    )

    all_fold_metrics: list[pd.DataFrame] = []
    all_predictions: list[pd.DataFrame] = []
    all_coefficients: list[pd.DataFrame] = []
    qc_rows: list[dict[str, object]] = []

    for protocol in selected_protocols:
        split_path = Path(paths.split_files[protocol])
        print(f"Running protocol={protocol} using {split_path}")
        folds = folds_from_main_loader(split_path, len(features))
        qc_rows.append(split_qc(protocol, folds, len(features)))
        fold_metrics, predictions, coefficients = evaluate_protocol(
            protocol,
            folds,
            features,
            model_specs,
        )
        all_fold_metrics.append(fold_metrics)
        all_predictions.append(predictions)
        all_coefficients.append(coefficients)

    fold_metrics = pd.concat(all_fold_metrics, ignore_index=True)
    predictions = pd.concat(all_predictions, ignore_index=True)
    coefficient_frames = [frame for frame in all_coefficients if not frame.empty]
    coefficients = (
        pd.concat(coefficient_frames, ignore_index=True)
        if coefficient_frames
        else pd.DataFrame()
    )
    aggregate_metrics, per_class_f1, confusion = aggregate_predictions(predictions)
    qc = pd.DataFrame(qc_rows)

    write_outputs(
        args.output_dir,
        features,
        fold_metrics,
        predictions,
        coefficients,
        aggregate_metrics,
        per_class_f1,
        confusion,
        qc,
    )

    print("\nAggregate out-of-fold results")
    print(
        aggregate_metrics[
            ["protocol", "model", "accuracy", "balanced_accuracy", "macro_f1_all4"]
        ].to_string(index=False)
    )
    print(f"\nOutputs written to: {args.output_dir}")


if __name__ == "__main__":
    main()

try:
    from ._bootstrap import PROJECT_ROOT  # noqa: F401
except ImportError:
    from _bootstrap import PROJECT_ROOT  # noqa: F401

from pathlib import Path
import argparse
import shutil
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from scipy.stats import kruskal, wilcoxon
from statsmodels.stats.multitest import multipletests
import paths


MODELS = [
    "denformer_first",
    "denformer_max",
    "denformer_mean",
    "longformer",
    "performer",
    "ffnn",
    "logreg",
]

DENFORMER_VARIANTS = [
    "denformer_first",
    "denformer_max",
    "denformer_mean",
]

SELECTED_DENFORMER = "denformer_mean"

OTHER_MODELS = [
    model for model in MODELS
    if model not in DENFORMER_VARIANTS
]

COMPARISON_GROUPS = {
    "all_models": MODELS,
    "denformer_variants": DENFORMER_VARIANTS,
    "selected_denformer_vs_others": [SELECTED_DENFORMER] + OTHER_MODELS,
}

STATISTICAL_GROUPS = {
    "denformer_variants": DENFORMER_VARIANTS,
    "selected_denformer_vs_others": [SELECTED_DENFORMER] + OTHER_MODELS,
}


DISPLAY_NAMES = {
    "ffnn": "FFNN",
    "denformer_first": "DENFormer-first",
    "denformer_mean": "DENFormer-mean",
    "denformer_max": "DENFormer-max",
    "longformer": "Longformer",
    "performer": "Performer",
    "logreg": "LogisticRegression",
}

MODEL_COLORS = {
    # DENFormer variants: lighter shades of the same blue family.
    # Kept visually related while avoiding a near-black dark blue in figures.
    "denformer_first": "#c6dbef",
    "denformer_max": "#9ecae1",
    "denformer_mean": "#6baed6",
    # Baselines: stable, model-specific colors used consistently in every figure.
    "longformer": "#ff7f0e",
    "performer": "#2ca02c",
    "ffnn": "#d62728",
    "logreg": "#9467bd",
}


def model_color(model):
    return MODEL_COLORS.get(model)

EXPERIMENTS = [
    "ohe_cdhit_e100",
    "ohe_continent_e100",
    "ohe_timebin_e100",
]

# Keep the wording aligned with the stats 2x2 panel in dataset/scripts.
EXPERIMENT_TITLES = {
    "all_experiments": "Overall",
    "ohe_cdhit_e100": "CD-HIT cluster-aware folds",
    "ohe_continent_e100": "Geographical distribution by continent",
    "ohe_timebin_e100": "Temporal distribution",
}

EXPERIMENT_PANEL_ORDER = [
    ("A", "all_experiments"),
    ("B", "ohe_cdhit_e100"),
    ("C", "ohe_continent_e100"),
    ("D", "ohe_timebin_e100"),
]

EXPERIMENT_ROW_PANEL_ORDER = [
    ("A", "ohe_cdhit_e100"),
    ("B", "ohe_continent_e100"),
    ("C", "ohe_timebin_e100"),
]

FOLD_GRID_EXPERIMENT_ORDER = [
    "ohe_continent_e100",
    "ohe_timebin_e100",
    "ohe_cdhit_e100",
]

FOLD_IDS_KEY = "_fold_ids"



RESULT_TABLE_EXPERIMENT_ORDER = [
    "ohe_continent_e100",
    "ohe_timebin_e100",
    "ohe_cdhit_e100",
]

RESULT_TABLE_SPLIT_LABELS = {
    "ohe_continent_e100": "Geographical",
    "ohe_timebin_e100": "Temporal",
    "ohe_cdhit_e100": "CD-HIT",
}

RESULT_TABLE_COLUMNS = [
    "Model",
    "Split Strategy",
    "Fold",
    "Accuracy",
    "DENV1",
    "DENV2",
    "DENV3",
    "DENV4",
]

RESULT_TABLE_METRIC_COLUMNS = {
    "Accuracy": "Accuracy",
    "F1-DENV1": "DENV1",
    "F1-DENV2": "DENV2",
    "F1-DENV3": "DENV3",
    "F1-DENV4": "DENV4",
}

GROUP_TITLES = {
    "all_models": "All models",
    "denformer_variants": "DENFormer variants",
    "selected_denformer_vs_others": "DENFormer-mean vs other models",
}

RECAP_1X3_DIR_NAME = "recap_1x3"


def copy_figure_for_paper(source_path, destination_dir, destination_name=None):
    """Copy manuscript-ready recap figures into a single summary folder."""
    source_path = Path(source_path)
    if not source_path.exists():
        return None
    destination_dir = Path(destination_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination_path = destination_dir / (destination_name or source_path.name)
    shutil.copy2(source_path, destination_path)
    return destination_path

BAR_TITLE_SIZE = 17
BAR_AXIS_LABEL_SIZE = 16
BAR_TICK_SIZE = 15
BAR_LEGEND_SIZE = 15
DOT_SIZE = 38
DOT_ALPHA = 0.75
DOT_SUMMARY_MARKER_SIZE = 90
HEATMAP_TITLE_SIZE = 17
HEATMAP_AXIS_LABEL_SIZE = 14
HEATMAP_TICK_SIZE = 13
HEATMAP_TEXT_SIZE = 12
HEATMAP_CBAR_SIZE = 13
FOLD_HEATMAP_TEXT_SIZE = 8
FOLD_HEATMAP_ANNOTATE_MAX_CELLS = 240
PANEL_TITLE_SIZE = 17
PANEL_SUPTITLE_SIZE = 23

METRICS = [
    "Accuracy",
    "F1-DENV1",
    "F1-DENV2",
    "F1-DENV3",
    "F1-DENV4",
    # "macro_precision",
    # "macro_recall",
    # "macro_f1",
    # "weighted_precision",
    # "weighted_recall",
    # "weighted_f1",
]

SEROTYPE_LABELS = {
    "F1-DENV1": "0",
    "F1-DENV2": "1",
    "F1-DENV3": "2",
    "F1-DENV4": "3",
}

METRIC_DISPLAY_NAMES = {
    "Accuracy": "Accuracy",
    "F1-DENV1": "DENV1",
    "F1-DENV2": "DENV2",
    "F1-DENV3": "DENV3",
    "F1-DENV4": "DENV4",
    "macro_precision": "Macro precision",
    "macro_recall": "Macro recall",
    "macro_f1": "Macro F1",
    "weighted_precision": "Weighted precision",
    "weighted_recall": "Weighted recall",
    "weighted_f1": "Weighted F1",
}


def display_metric_name(metric_name):
    return METRIC_DISPLAY_NAMES.get(metric_name, metric_name)


def display_metric_names(metrics):
    return [display_metric_name(metric) for metric in metrics]


def apply_score_axis_format(ax, as_percent=False):
    if as_percent:
        ax.yaxis.set_major_formatter(FuncFormatter(lambda y, _: f"{y * 100:.0f}%"))

ERROR_BAR_CHOICES = ("minmax", "std")
DEFAULT_ERROR_BAR_MODE = "std"

SUMMARY_CHOICES = ("mean", "median")
DEFAULT_SUMMARY_MODE = "mean"
RESULT_TABLE_VALUES_AS_PERCENT = False
BARPLOT_AUTO_YMIN = True
BARPLOT_Y_PADDING_FRACTION = 0.08


def output_subdir_name(summary_mode, error_bar_mode):
    return f"{summary_mode}_{error_bar_mode}"

def parse_args():
    parser = argparse.ArgumentParser(
        description="Aggregate fold metrics and plot model bar charts."
    )
    parser.add_argument(
        "--error-bar",
        choices=ERROR_BAR_CHOICES,
        default=DEFAULT_ERROR_BAR_MODE,
        help=(
            "Error-bar definition for plots. "
            "'minmax' draws bars from fold minimum to fold maximum around the selected summary; "
            "'std' draws selected summary +/- standard deviation. Default: std."
        ),
    )
    parser.add_argument(
        "--summary",
        choices=SUMMARY_CHOICES,
        default=DEFAULT_SUMMARY_MODE,
        help=(
            "Central summary shown by bars. "
            "Use 'mean' with --error-bar std, or 'median' with --error-bar minmax. "
            "Default: mean."
        ),
    )
    parser.add_argument(
        "--result-table-percent",
        action="store_true",
        default=RESULT_TABLE_VALUES_AS_PERCENT,
        help=(
            "Format manuscript/template result tables as percentages (XX.xx %%). "
            "By default, values are formatted as decimals (0.XXXX)."
        ),
    )
    parser.add_argument(
        "--barplot-y-from-zero",
        action="store_true",
        help=(
            "Force metric barplots/dotplots to start from 0. "
            "By default, the y-axis is zoomed dynamically so error bars remain inside the panel."
        ),
    )
    return parser.parse_args()


def summarize_values(values):
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        return np.nan, np.nan, np.nan, np.nan, np.nan
    return (
        float(np.mean(values)),
        float(np.median(values)),
        float(np.std(values)),
        float(np.min(values)),
        float(np.max(values)),
    )


def summary_std_str(summary_value, std):
    return f"{summary_value:.4f} ({std:.4f})"


def summary_minmax_str(summary_value, min_value, max_value):
    return f"{summary_value:.4f} [{min_value:.4f}-{max_value:.4f}]"


def metric_center(row, metric, summary_mode):
    return float(row[f"{metric}_{summary_mode}"])


def metric_error_bounds(row, metric, error_bar_mode, summary_mode):
    center = metric_center(row, metric, summary_mode)

    if error_bar_mode == "std":
        std = float(row[f"{metric}_std"])
        return center - std, center + std

    if error_bar_mode == "minmax":
        return float(row[f"{metric}_min"]), float(row[f"{metric}_max"])

    raise ValueError(f"Unsupported error_bar_mode: {error_bar_mode}")


def normalize_fold_id(fold_id):
    if fold_id is None:
        raise ValueError("Missing fold identifier in inference_summary.json")
    value = str(fold_id).strip()
    if not value:
        raise ValueError("Empty fold identifier in inference_summary.json")
    return value


def fold_sort_key(experiment_name, fold_id):
    """Sort real fold identifiers without inferring their identity by position."""
    value = normalize_fold_id(fold_id)
    compact = value.replace(" ", "")

    if experiment_name == "ohe_timebin_e100":
        if compact.startswith("<"):
            try:
                return (0, int(compact[1:]), value.casefold())
            except ValueError:
                return (0, -1, value.casefold())
        try:
            start_year = int(compact.split("-", 1)[0])
            return (1, start_year, value.casefold())
        except ValueError:
            return (2, value.casefold())

    if experiment_name == "ohe_cdhit_e100":
        numeric = compact
        for prefix in ("fold_", "Fold_", "fold", "Fold"):
            if numeric.startswith(prefix):
                numeric = numeric[len(prefix):]
                break
        numeric = numeric.lstrip("_- ")
        if numeric.isdigit():
            return (0, int(numeric))

    return (0, value.casefold())


def display_fold_label(experiment_name, fold_id):
    """Format the real fold identifier for manuscript-ready outputs."""
    value = normalize_fold_id(fold_id)
    compact = value.replace(" ", "")

    if experiment_name == "ohe_timebin_e100" and compact.startswith("<"):
        return f"< {compact[1:]}"

    if experiment_name == "ohe_cdhit_e100":
        numeric = compact
        for prefix in ("fold_", "Fold_", "fold", "Fold"):
            if numeric.startswith(prefix):
                numeric = numeric[len(prefix):]
                break
        numeric = numeric.lstrip("_- ")
        if numeric.isdigit():
            return f"Fold {int(numeric)}"

    return value


def validate_metric_fold_lengths(metric_values, context):
    fold_ids = metric_values.get(FOLD_IDS_KEY, [])
    expected = len(fold_ids)
    if len(set(fold_ids)) != expected:
        raise ValueError(f"Duplicate fold identifiers in {context}: {fold_ids}")

    for metric in METRICS:
        actual = len(metric_values.get(metric, []))
        if actual != expected:
            raise ValueError(
                f"Fold/metric length mismatch in {context}: "
                f"{expected} fold identifiers but {actual} values for {metric}."
            )


def validate_model_fold_alignment(metric_values_by_model, models, context):
    reference_model = None
    reference_ids = None
    for model in models:
        metric_values = metric_values_by_model.get(model)
        if metric_values is None:
            continue
        validate_metric_fold_lengths(metric_values, f"{context} / {model}")
        fold_ids = tuple(metric_values[FOLD_IDS_KEY])
        if reference_ids is None:
            reference_model = model
            reference_ids = fold_ids
        elif fold_ids != reference_ids:
            raise ValueError(
                f"Fold alignment mismatch in {context}: {model} has {fold_ids}, "
                f"whereas {reference_model} has {reference_ids}."
            )


def extract_metric_values(summary_json, experiment_name):
    values = {metric: [] for metric in METRICS}
    values[FOLD_IDS_KEY] = []

    ordered_folds = sorted(
        summary_json,
        key=lambda fold: fold_sort_key(experiment_name, fold.get("fold")),
    )

    for fold in ordered_folds:
        fold_id = normalize_fold_id(fold.get("fold"))
        values[FOLD_IDS_KEY].append(fold_id)

        test = fold["test"]
        report = test["classification_report"]

        if "Accuracy" in values:
            values["Accuracy"].append(test["accuracy"])

        for metric_name, report_label in SEROTYPE_LABELS.items():
            if metric_name not in values:
                continue

            if report_label not in report:
                # Some evaluation splits may not contain all serotypes in a fold/test set.
                # Keep the explicit numeric mapping fixed, but store NaN for missing
                # serotypes so summaries and statistical tests can ignore them safely.
                values[metric_name].append(np.nan)
                continue

            values[metric_name].append(report[report_label]["f1-score"])

        # To re-enable aggregate metrics, uncomment the corresponding entries
        # in METRICS and append their values here as well.

    validate_metric_fold_lengths(values, f"{experiment_name} inference summary")
    return values


def build_rows_from_metric_values(metric_values_by_model, models, summary_mode):
    rows_csv = []
    rows_numeric = []

    for model in models:
        metric_values = metric_values_by_model.get(model)

        if metric_values is None:
            continue

        row_csv = {"model": model}
        row_numeric = {"model": model}

        for metric in METRICS:
            mean, median, std, min_value, max_value = summarize_values(metric_values[metric])
            selected_summary = mean if summary_mode == "mean" else median

            row_csv[metric] = summary_std_str(selected_summary, std)
            row_csv[f"{metric}_{summary_mode}_std"] = summary_std_str(selected_summary, std)
            row_csv[f"{metric}_{summary_mode}_minmax"] = summary_minmax_str(selected_summary, min_value, max_value)
            row_numeric[f"{metric}_mean"] = mean
            row_numeric[f"{metric}_median"] = median
            row_numeric[f"{metric}_std"] = std
            row_numeric[f"{metric}_min"] = min_value
            row_numeric[f"{metric}_max"] = max_value

        rows_csv.append(row_csv)
        rows_numeric.append(row_numeric)

    return rows_csv, rows_numeric




def build_fold_values_df(metric_values_by_model, models, experiment_name):
    rows = []
    validate_model_fold_alignment(
        metric_values_by_model,
        models,
        f"{experiment_name} fold-level report",
    )

    for model in models:
        metric_values = metric_values_by_model.get(model)
        if metric_values is None:
            continue

        fold_ids = metric_values[FOLD_IDS_KEY]
        for fold_idx, fold_id in enumerate(fold_ids, start=1):
            for metric in METRICS:
                value = metric_values[metric][fold_idx - 1]
                rows.append({
                    "model": model,
                    "model_display": display_model_name(model),
                    "fold_index": fold_idx,
                    "fold_id": fold_id,
                    "fold_label": display_fold_label(experiment_name, fold_id),
                    "metric": metric,
                    "metric_display": display_metric_name(metric),
                    "value": float(value) if np.isfinite(value) else np.nan,
                })

    return pd.DataFrame(rows)


def add_fold_labels(fold_df, experiment_name):
    if fold_df.empty:
        return fold_df.copy()
    if "fold_id" not in fold_df.columns:
        raise ValueError("fold_id is missing from fold-level reporting data")
    out = fold_df.copy()
    out["fold_label"] = out["fold_id"].apply(
        lambda fold_id: display_fold_label(experiment_name, fold_id)
    )
    return out


def build_fold_metric_wide_df(fold_df, experiment_name):
    """Paper-friendly wide table: one row per model/fold and one column per metric."""
    if fold_df.empty:
        return pd.DataFrame()

    labeled = add_fold_labels(fold_df, experiment_name)
    wide = labeled.pivot_table(
        index=["model", "model_display", "fold_index", "fold_id", "fold_label"],
        columns="metric_display",
        values="value",
        aggfunc="first",
    ).reset_index()
    wide.columns.name = None
    wide = wide.sort_values(["model", "fold_index"])
    return wide
def metric_summary_from_values(values, summary_mode):
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return np.nan
    if summary_mode == "median":
        return float(np.median(arr))
    if summary_mode == "mean":
        return float(np.mean(arr))
    raise ValueError(f"Unsupported summary_mode: {summary_mode}")


def _nice_floor(value, step):
    if not np.isfinite(value):
        return 0.0
    return step * np.floor(value / step)


def _nice_ceil(value, step):
    if not np.isfinite(value):
        return 1.0
    return step * np.ceil(value / step)


def _choose_tick_step(y_min, y_max):
    span = max(float(y_max - y_min), 1e-6)
    if span <= 0.16:
        return 0.02
    if span <= 0.35:
        return 0.05
    return 0.10


def compute_global_ylim_and_ticks(all_numeric_dfs, error_bar_mode, summary_mode, force_zero=False):
    """Return y-axis limits that include the complete error bars.

    Accuracy and F1 are bounded scores, but mean +/- std can extend above 1.0.
    For publication figures it is usually more informative not to start from zero
    when all models are close to saturation; this also prevents error-bar caps from
    being clipped by the top border. Use --barplot-y-from-zero to recover the old
    behavior.
    """
    lowers = []
    uppers = []

    for df_numeric in all_numeric_dfs:
        if df_numeric is None or df_numeric.empty:
            continue
        for _, row in df_numeric.iterrows():
            for metric in METRICS:
                try:
                    center = metric_center(row, metric, summary_mode)
                    lower, upper = metric_error_bounds(row, metric, error_bar_mode, summary_mode)
                except Exception:
                    continue
                if np.isfinite(center):
                    lowers.append(center)
                    uppers.append(center)
                if np.isfinite(lower):
                    lowers.append(lower)
                if np.isfinite(upper):
                    uppers.append(upper)

    if not lowers or not uppers:
        if summary_mode == "median" and error_bar_mode == "minmax" and not force_zero:
            return 0.10, 1.02, np.arange(0.20, 1.0 + 1e-9, 0.20)
        y_min = 0.0
        y_max = 1.05
        y_ticks = np.arange(y_min, y_max + 1e-9, 0.1)
        return y_min, y_max, y_ticks

    data_min = float(np.nanmin(lowers))
    data_max = float(np.nanmax(uppers))

    # Median/min-max panels show bounded raw fold ranges rather than std. For
    # these recap figures, keep the axis compact and paper-like: start at 10%
    # and show ticks every 20 percentage points, while leaving a small headroom
    # above 100% so bars/error-bar caps at 1.0 remain visible.
    if summary_mode == "median" and error_bar_mode == "minmax" and not force_zero:
        y_min = 0.10 if data_min >= 0.10 else max(0.0, _nice_floor(data_min, 0.05))
        y_max = 1.02 if data_max <= 1.0 + 1e-9 else min(1.20, data_max + 0.02)
        y_ticks = np.arange(0.20, 1.0 + 1e-9, 0.20)
        y_ticks = y_ticks[(y_ticks >= y_min - 1e-9) & (y_ticks <= 1.0 + 1e-9)]
        if y_ticks.size == 0 or abs(y_ticks[-1] - 1.0) > 1e-9:
            y_ticks = np.append(y_ticks, 1.0)
        return y_min, y_max, y_ticks

    raw_span = max(data_max - data_min, 0.03)
    padding = max(BARPLOT_Y_PADDING_FRACTION * raw_span, 0.015)

    if force_zero:
        y_min = 0.0
    else:
        y_min = max(0.0, data_min - padding)

    # Accuracy/F1 are bounded at 1.0. Summary error bars (especially std) can
    # extend above 1.0; let the plot area reach up to 1.2 if needed, while the
    # visible tick labels remain capped at 1.0 to avoid misleading score ticks
    # above the metric maximum.
    if data_max <= 1.0 + 1e-9:
        y_max = 1.015
    else:
        y_max = min(1.20, max(data_max + padding, 1.08))

    tick_step = _choose_tick_step(y_min, y_max)
    y_min = 0.0 if force_zero else _nice_floor(y_min, tick_step)
    y_max = _nice_ceil(y_max, tick_step)
    y_max = min(y_max, 1.20)
    if y_max <= y_min:
        y_max = min(1.20, y_min + tick_step)

    tick_upper = min(1.0, y_max)
    y_ticks = np.arange(y_min, tick_upper + tick_step * 0.5, tick_step)
    y_ticks = y_ticks[y_ticks <= tick_upper + 1e-9]
    if y_ticks.size == 0 or abs(y_ticks[-1] - tick_upper) > 1e-9:
        y_ticks = np.append(y_ticks, tick_upper)
    return y_min, y_max, y_ticks




def compute_fold_barplot_grid_ylim_and_ticks(output_tables, group_name, force_zero=False):
    """Return compact y-axis limits for raw fold-level metric barplots.

    These grids do not display summary error bars, so their scale should not be
    inherited from mean +/- std panels. For bounded metrics such as Accuracy and
    F1, keep the top just above 1.0 unless raw values exceed 1.0 because of an
    upstream numeric convention.
    """
    values = []
    for experiment_name in FOLD_GRID_EXPERIMENT_ORDER:
        tables = output_tables.get((experiment_name, group_name))
        if tables is None:
            continue
        fold_df = tables.get("fold_values")
        if fold_df is None or fold_df.empty:
            continue
        vals = pd.to_numeric(fold_df["value"], errors="coerce").to_numpy(dtype=float)
        values.extend(vals[np.isfinite(vals)].tolist())

    if not values:
        y_min, y_max = (0.0, 1.015)
        y_ticks = np.arange(0.0, 1.0 + 1e-9, 0.1)
        return y_min, y_max, y_ticks

    data_min = float(np.nanmin(values))
    data_max = float(np.nanmax(values))
    raw_span = max(data_max - data_min, 0.03)
    padding = max(BARPLOT_Y_PADDING_FRACTION * raw_span, 0.012)

    if force_zero:
        y_min = 0.0
    else:
        y_min = max(0.0, data_min - padding)

    tick_step = _choose_tick_step(y_min, 1.0 if data_max <= 1.0 + 1e-9 else min(1.10, data_max + padding))
    y_min = 0.0 if force_zero else _nice_floor(y_min, tick_step)

    if data_max <= 1.0 + 1e-9:
        # Raw fold-level plots have no error bars. Keep the visible axis just
        # above the biological/metric maximum, but do not add a 1.1 y-tick.
        y_max = 1.015
        y_ticks = np.arange(y_min, 1.0 + tick_step * 0.5, tick_step)
        y_ticks = y_ticks[y_ticks <= 1.0 + 1e-9]
        if y_ticks.size == 0 or abs(y_ticks[-1] - 1.0) > 1e-9:
            y_ticks = np.append(y_ticks, 1.0)
        return y_min, y_max, y_ticks

    y_max = min(1.10, data_max + padding)
    y_max = _nice_ceil(y_max, tick_step)
    y_max = min(y_max, 1.10)
    if y_max <= y_min:
        y_max = min(1.10, y_min + tick_step)

    y_ticks = np.arange(y_min, y_max + tick_step * 0.5, tick_step)
    return y_min, y_max, y_ticks


def draw_barplot(ax, df_numeric, title, y_min, y_max, y_ticks, error_bar_mode, summary_mode, show_xlabel=True, show_ylabel=True, axis_as_percent=False):
    x = np.arange(len(METRICS))

    n_models = len(df_numeric)
    width = min(0.13, 0.82 / max(1, n_models))
    offsets = (np.arange(n_models) - (n_models - 1) / 2) * width

    handles = []
    labels = []

    for i, (_, row) in enumerate(df_numeric.iterrows()):
        centers = np.array([metric_center(row, metric, summary_mode) for metric in METRICS], dtype=float)

        if error_bar_mode == "std":
            yerr = np.array([row[f"{metric}_std"] for metric in METRICS], dtype=float)
        elif error_bar_mode == "minmax":
            mins = np.array([row[f"{metric}_min"] for metric in METRICS], dtype=float)
            maxs = np.array([row[f"{metric}_max"] for metric in METRICS], dtype=float)
            yerr = np.vstack([centers - mins, maxs - centers])
        else:
            raise ValueError(f"Unsupported error_bar_mode: {error_bar_mode}")

        yerr = np.nan_to_num(yerr, nan=0.0)
        label = DISPLAY_NAMES.get(row["model"], row["model"])

        bars = ax.bar(
            x + offsets[i],
            centers,
            width,
            yerr=yerr,
            capsize=3,
            label=label,
            color=model_color(row["model"]),
        )
        handles.append(bars[0])
        labels.append(label)

    ax.set_ylim(y_min, y_max)
    ax.set_yticks(y_ticks)
    apply_score_axis_format(ax, axis_as_percent)
    ax.tick_params(axis="both", labelsize=BAR_TICK_SIZE)
    ax.set_xticks(x)
    ax.set_xticklabels(display_metric_names(METRICS), rotation=30, ha="right", fontsize=BAR_TICK_SIZE)

    if show_ylabel:
        ax.set_ylabel("Score", fontsize=BAR_AXIS_LABEL_SIZE)
    else:
        ax.set_ylabel("")

    if show_xlabel:
        ax.set_xlabel("", fontsize=BAR_AXIS_LABEL_SIZE)
    else:
        ax.set_xlabel("")

    ax.set_title(title, fontsize=BAR_TITLE_SIZE, fontweight="bold", pad=10)
    return handles, labels


def save_barplot(df_numeric, title, out_path, y_min, y_max, y_ticks, error_bar_mode, summary_mode, axis_as_percent=False):
    fig_width = max(8.5, 1.25 * len(METRICS))
    fig, ax = plt.subplots(figsize=(fig_width, 7))
    handles, labels = draw_barplot(
        ax,
        df_numeric,
        title,
        y_min,
        y_max,
        y_ticks,
        error_bar_mode,
        summary_mode,
        show_xlabel=False,
        show_ylabel=True,
        axis_as_percent=axis_as_percent,
    )

    ax.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.20),
        ncol=len(labels),
        frameon=False,
        fontsize=BAR_LEGEND_SIZE,
    )

    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_barplot_panel(output_tables, group_name, out_path, y_min, y_max, y_ticks, error_bar_mode, summary_mode, axis_as_percent=False):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_PANEL_ORDER
        if (experiment_name, group_name) in output_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(2, 2, figsize=(14.5, 10.5), sharey=True)
    axes = axes.ravel()
    legend_handles = None
    legend_labels = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        df_numeric = output_tables[(experiment_name, group_name)]["numeric"]
        handles, labels = draw_barplot(
            ax,
            df_numeric,
            f"{panel_label}. {get_experiment_title(experiment_name)}",
            y_min,
            y_max,
            y_ticks,
            error_bar_mode,
            summary_mode,
            show_xlabel=False,
            show_ylabel=False,
            axis_as_percent=axis_as_percent,
        )
        if legend_handles is None:
            legend_handles = handles
            legend_labels = labels

    for ax in axes[len(available):]:
        ax.axis("off")

    if legend_handles is not None:
        fig.legend(
            legend_handles,
            legend_labels,
            loc="lower center",
            ncol=len(legend_labels),
            frameon=False,
            fontsize=BAR_LEGEND_SIZE,
            bbox_to_anchor=(0.5, 0.015),
        )

    fig.subplots_adjust(left=0.08, right=0.98, top=0.94, bottom=0.14, wspace=0.12, hspace=0.40)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_barplot_row_panel(output_tables, group_name, out_path, y_min, y_max, y_ticks, error_bar_mode, summary_mode, axis_as_percent=False):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_ROW_PANEL_ORDER
        if (experiment_name, group_name) in output_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(1, len(available), figsize=(21.0, 5.4), sharey=True)
    axes = np.atleast_1d(axes).ravel()
    legend_handles = None
    legend_labels = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        df_numeric = output_tables[(experiment_name, group_name)]["numeric"]
        handles, labels = draw_barplot(
            ax,
            df_numeric,
            f"{panel_label}. {get_experiment_title(experiment_name)}",
            y_min,
            y_max,
            y_ticks,
            error_bar_mode,
            summary_mode,
            show_xlabel=False,
            show_ylabel=False,
            axis_as_percent=axis_as_percent,
        )
        if legend_handles is None:
            legend_handles = handles
            legend_labels = labels

    if legend_handles is not None:
        fig.legend(
            legend_handles,
            legend_labels,
            loc="lower center",
            ncol=len(legend_labels),
            frameon=False,
            fontsize=BAR_LEGEND_SIZE,
            bbox_to_anchor=(0.5, 0.015),
        )

    fig.subplots_adjust(left=0.055, right=0.995, top=0.86, bottom=0.27, wspace=0.14)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)





def _available_fold_indices_for_experiment(output_tables, group_name, experiment_name):
    tables = output_tables.get((experiment_name, group_name))
    if tables is None:
        return []
    fold_df = tables["fold_values"]
    if fold_df.empty:
        return []
    return sorted(int(idx) for idx in fold_df["fold_index"].dropna().unique())


def _fold_grid_column_order(output_tables, group_name):
    fold_indices = []
    for experiment_name in FOLD_GRID_EXPERIMENT_ORDER:
        fold_indices.extend(_available_fold_indices_for_experiment(output_tables, group_name, experiment_name))
    if not fold_indices:
        return []
    return list(range(1, max(fold_indices) + 1))


def draw_fold_barplot(ax, fold_df, models, title, y_min, y_max, y_ticks, show_ylabel=False, show_xlabel=False, axis_as_percent=False):
    x = np.arange(len(METRICS))
    present_models = [model for model in models if model in set(fold_df["model"].dropna())]
    n_models = len(present_models)
    width = min(0.13, 0.82 / max(1, n_models))
    offsets = (np.arange(n_models) - (n_models - 1) / 2) * width

    handles = []
    labels = []

    for i, model in enumerate(present_models):
        model_df = fold_df[fold_df["model"].eq(model)]
        centers = []
        for metric in METRICS:
            row = model_df[model_df["metric"].eq(metric)]
            centers.append(float(row.iloc[0]["value"]) if not row.empty else np.nan)
        label = display_model_name(model)
        bars = ax.bar(x + offsets[i], centers, width, label=label, color=model_color(model))
        handles.append(bars[0])
        labels.append(label)

    ax.set_ylim(y_min, y_max)
    ax.set_yticks(y_ticks)
    apply_score_axis_format(ax, axis_as_percent)
    ax.tick_params(axis="both", labelsize=max(9, BAR_TICK_SIZE - 2))
    ax.set_xticks(x)
    ax.set_xticklabels(
        display_metric_names(METRICS),
        rotation=32,
        ha="right",
        fontsize=max(9, BAR_TICK_SIZE - 2),
    )
    ax.grid(axis="y", alpha=0.20, linewidth=0.7)
    ax.set_title(title, fontsize=PANEL_TITLE_SIZE, fontweight="bold", pad=5)
    ax.set_ylabel("")
    ax.set_xlabel("", fontsize=BAR_AXIS_LABEL_SIZE)
    return handles, labels


def save_fold_barplot_grid(output_tables, group_name, out_path, y_min, y_max, y_ticks, axis_as_percent=False):
    """Save a embedding-style grid: rows are split strategies, columns are folds."""
    experiments = [
        experiment_name for experiment_name in FOLD_GRID_EXPERIMENT_ORDER
        if (experiment_name, group_name) in output_tables
    ]
    fold_indices = _fold_grid_column_order(output_tables, group_name)
    if not experiments or not fold_indices:
        return

    n_rows = len(experiments)
    n_cols = len(fold_indices)
    fig_width = max(16.5, 3.95 * n_cols)
    fig_height = max(8.4, 4.05 * n_rows + 1.80)
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(fig_width, fig_height), sharey=True)
    axes = np.asarray(axes).reshape(n_rows, n_cols)

    legend_handles = None
    legend_labels = None

    for row_idx, experiment_name in enumerate(experiments):
        tables = output_tables[(experiment_name, group_name)]
        fold_df = tables["fold_values"]
        models = COMPARISON_GROUPS[group_name]
        available_fold_set = set(int(idx) for idx in fold_df["fold_index"].dropna().unique())

        for col_idx, fold_index in enumerate(fold_indices):
            ax = axes[row_idx, col_idx]
            if fold_index not in available_fold_set:
                ax.axis("off")
                continue

            one_fold = fold_df[fold_df["fold_index"].eq(fold_index)]
            title = str(one_fold.iloc[0]["fold_label"])
            handles, labels = draw_fold_barplot(
                ax,
                one_fold,
                models,
                title,
                y_min,
                y_max,
                y_ticks,
                show_ylabel=(col_idx == 0),
                show_xlabel=(row_idx == n_rows - 1),
            )
            if legend_handles is None and handles:
                legend_handles = handles
                legend_labels = labels

            if col_idx == 0:
                ax.annotate(
                    get_fold_grid_row_label(experiment_name),
                    xy=(-0.38, 0.5),
                    xycoords="axes fraction",
                    rotation=90,
                    ha="center",
                    va="center",
                    fontsize=PANEL_TITLE_SIZE,
                    fontweight="bold",
                )

    if legend_handles is not None:
        fig.legend(
            legend_handles,
            legend_labels,
            loc="lower center",
            ncol=len(legend_labels),
            frameon=False,
            fontsize=BAR_LEGEND_SIZE,
            bbox_to_anchor=(0.5, 0.012),
        )

    fig.suptitle(
        f"Fold-level metric barplots ({get_group_title(group_name)})",
        fontsize=PANEL_SUPTITLE_SIZE,
        fontweight="bold",
        y=0.992,
    )
    fig.subplots_adjust(left=0.105, right=0.995, top=0.905, bottom=0.155, wspace=0.22, hspace=0.68)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

def draw_dotplot(ax, fold_df, title, y_min, y_max, y_ticks, summary_mode, show_xlabel=True, show_ylabel=True, axis_as_percent=False):
    x = np.arange(len(METRICS))
    present_models = [model for model in MODELS if model in set(fold_df["model"].dropna())]
    n_models = len(present_models)
    width = min(0.13, 0.82 / max(1, n_models))
    offsets = (np.arange(n_models) - (n_models - 1) / 2) * width

    handles = []
    labels = []

    for i, model in enumerate(present_models):
        label = DISPLAY_NAMES.get(model, model)
        model_df = fold_df[fold_df["model"].eq(model)]
        first_handle = None

        for j, metric in enumerate(METRICS):
            values = model_df[model_df["metric"].eq(metric)]["value"].dropna().to_numpy(dtype=float)
            if values.size == 0:
                continue

            if values.size == 1:
                jitter = np.array([0.0])
            else:
                jitter = np.linspace(-width * 0.22, width * 0.22, values.size)
            xs = np.full(values.size, x[j] + offsets[i]) + jitter

            scatter = ax.scatter(
                xs,
                values,
                s=DOT_SIZE,
                alpha=DOT_ALPHA,
                edgecolors="black",
                linewidths=0.35,
                label=label if first_handle is None else None,
                color=model_color(model),
                zorder=3,
            )
            if first_handle is None:
                first_handle = scatter

            summary_value = metric_summary_from_values(values, summary_mode)
            if np.isfinite(summary_value):
                ax.scatter(
                    [x[j] + offsets[i]],
                    [summary_value],
                    s=DOT_SUMMARY_MARKER_SIZE,
                    marker="_",
                    linewidths=2.0,
                    color=model_color(model),
                    zorder=4,
                )

        if first_handle is not None:
            handles.append(first_handle)
            labels.append(label)

    ax.set_ylim(y_min, y_max)
    ax.set_yticks(y_ticks)
    apply_score_axis_format(ax, axis_as_percent)
    ax.tick_params(axis="both", labelsize=BAR_TICK_SIZE)
    ax.set_xticks(x)
    ax.set_xticklabels(display_metric_names(METRICS), rotation=30, ha="right", fontsize=BAR_TICK_SIZE)
    ax.grid(axis="y", alpha=0.25, linewidth=0.8)

    if show_ylabel:
        ax.set_ylabel("Score", fontsize=BAR_AXIS_LABEL_SIZE)
    else:
        ax.set_ylabel("")

    if show_xlabel:
        ax.set_xlabel("", fontsize=BAR_AXIS_LABEL_SIZE)
    else:
        ax.set_xlabel("")

    ax.set_title(title, fontsize=BAR_TITLE_SIZE, fontweight="bold", pad=10)
    return handles, labels


def save_dotplot(fold_df, title, out_path, y_min, y_max, y_ticks, summary_mode, axis_as_percent=False):
    if fold_df.empty:
        return

    fig_width = max(8.5, 1.25 * len(METRICS))
    fig, ax = plt.subplots(figsize=(fig_width, 7))
    handles, labels = draw_dotplot(
        ax,
        fold_df,
        title,
        y_min,
        y_max,
        y_ticks,
        summary_mode,
        show_xlabel=False,
        show_ylabel=True,
        axis_as_percent=axis_as_percent,
    )

    ax.legend(
        handles,
        labels,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.20),
        ncol=len(labels),
        frameon=False,
        fontsize=BAR_LEGEND_SIZE,
    )

    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_dotplot_panel(output_tables, group_name, out_path, y_min, y_max, y_ticks, summary_mode, axis_as_percent=False):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_PANEL_ORDER
        if (experiment_name, group_name) in output_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(2, 2, figsize=(14.5, 10.5), sharey=True)
    axes = axes.ravel()
    legend_handles = None
    legend_labels = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        fold_df = output_tables[(experiment_name, group_name)]["fold_values"]
        handles, labels = draw_dotplot(
            ax,
            fold_df,
            f"{panel_label}. {get_experiment_title(experiment_name)}",
            y_min,
            y_max,
            y_ticks,
            summary_mode,
            show_xlabel=False,
            show_ylabel=False,
            axis_as_percent=axis_as_percent,
        )
        if legend_handles is None:
            legend_handles = handles
            legend_labels = labels

    for ax in axes[len(available):]:
        ax.axis("off")

    if legend_handles is not None:
        fig.legend(
            legend_handles,
            legend_labels,
            loc="lower center",
            ncol=len(legend_labels),
            frameon=False,
            fontsize=BAR_LEGEND_SIZE,
            bbox_to_anchor=(0.5, 0.015),
        )

    fig.subplots_adjust(left=0.08, right=0.98, top=0.94, bottom=0.14, wspace=0.12, hspace=0.40)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_dotplot_row_panel(output_tables, group_name, out_path, y_min, y_max, y_ticks, summary_mode, axis_as_percent=False):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_ROW_PANEL_ORDER
        if (experiment_name, group_name) in output_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(1, len(available), figsize=(21.0, 5.4), sharey=True)
    axes = np.atleast_1d(axes).ravel()
    legend_handles = None
    legend_labels = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        fold_df = output_tables[(experiment_name, group_name)]["fold_values"]
        handles, labels = draw_dotplot(
            ax,
            fold_df,
            f"{panel_label}. {get_experiment_title(experiment_name)}",
            y_min,
            y_max,
            y_ticks,
            summary_mode,
            show_xlabel=False,
            show_ylabel=False,
            axis_as_percent=axis_as_percent,
        )
        if legend_handles is None:
            legend_handles = handles
            legend_labels = labels

    if legend_handles is not None:
        fig.legend(
            legend_handles,
            legend_labels,
            loc="lower center",
            ncol=len(legend_labels),
            frameon=False,
            fontsize=BAR_LEGEND_SIZE,
            bbox_to_anchor=(0.5, 0.015),
        )

    fig.subplots_adjust(left=0.055, right=0.995, top=0.86, bottom=0.27, wspace=0.14)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def build_fold_metric_heatmap_matrix(fold_df):
    """Build a model x metric by split/fold matrix for visualization/export."""
    if fold_df.empty:
        return [], [], np.empty((0, 0), dtype=float)

    present_models = [
        model for model in MODELS
        if model in set(fold_df["model"].dropna())
    ]
    fold_metadata = (
        fold_df[["fold_index", "fold_label"]]
        .drop_duplicates()
        .sort_values("fold_index")
    )
    fold_indices = fold_metadata["fold_index"].astype(int).tolist()
    column_labels = fold_metadata["fold_label"].astype(str).tolist()

    row_labels = []
    matrix_rows = []

    for model in present_models:
        model_df = fold_df[fold_df["model"].eq(model)]
        model_label = display_model_name(model)

        for metric in METRICS:
            row_labels.append(f"{model_label} | {display_metric_name(metric)}")
            values = []
            for fold_idx in fold_indices:
                value_row = model_df[
                    model_df["fold_index"].eq(fold_idx)
                    & model_df["metric"].eq(metric)
                ]
                if value_row.empty:
                    values.append(np.nan)
                else:
                    values.append(float(value_row.iloc[0]["value"]))
            matrix_rows.append(values)

    matrix = np.asarray(matrix_rows, dtype=float) if matrix_rows else np.empty((0, len(column_labels)), dtype=float)
    return row_labels, column_labels, matrix


def save_fold_metric_heatmap(fold_df, title, out_path, group_name):
    """Save a heatmap of fold-level metrics for one evaluation protocol."""
    row_labels, column_labels, matrix = build_fold_metric_heatmap_matrix(fold_df)
    if matrix.size == 0 or not row_labels or not column_labels:
        return

    fig_width = max(7.2, 1.05 * len(column_labels) + 4.6)
    fig_height = max(5.4, 0.36 * len(row_labels) + 2.4)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    im = ax.imshow(matrix, aspect="auto", interpolation="nearest", vmin=0.0, vmax=1.0)
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("Score", fontsize=HEATMAP_CBAR_SIZE)
    cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    ax.set_xticks(np.arange(len(column_labels)))
    ax.set_xticklabels(column_labels, rotation=0, ha="center", fontsize=HEATMAP_TICK_SIZE)
    ax.set_yticks(np.arange(len(row_labels)))
    ax.set_yticklabels(row_labels, fontsize=max(8, HEATMAP_TICK_SIZE - 2))
    ax.set_xlabel("Split / fold", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    ax.set_ylabel("Model | metric", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    ax.set_title(
        f"{title}: fold-level metrics ({get_group_title(group_name)})",
        fontsize=HEATMAP_TITLE_SIZE,
        fontweight="bold",
        pad=10,
    )

    # Draw separators between models to make the repeated metric block easier to read.
    for boundary in range(len(METRICS), len(row_labels), len(METRICS)):
        ax.axhline(boundary - 0.5, color="white", linewidth=1.4)

    if matrix.size <= FOLD_HEATMAP_ANNOTATE_MAX_CELLS:
        for i in range(matrix.shape[0]):
            for j in range(matrix.shape[1]):
                value = matrix[i, j]
                if np.isfinite(value):
                    ax.text(
                        j,
                        i,
                        f"{value:.2f}",
                        ha="center",
                        va="center",
                        fontsize=FOLD_HEATMAP_TEXT_SIZE,
                        color="black",
                    )

    plt.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_fold_metric_excel(fold_df, df_csv, df_numeric, out_path, experiment_name):
    """Export fold-level values, wide tables, heatmap data and summaries to Excel."""
    row_labels, column_labels, matrix = build_fold_metric_heatmap_matrix(fold_df)

    if matrix.size:
        heatmap_df = pd.DataFrame(matrix, index=row_labels, columns=column_labels)
        heatmap_df.index.name = "model_metric"
        heatmap_df = heatmap_df.reset_index()
    else:
        heatmap_df = pd.DataFrame()

    export_fold_df = add_fold_labels(fold_df, experiment_name)
    if not export_fold_df.empty:
        export_fold_df = export_fold_df.sort_values(["model", "metric", "fold_index"])

    fold_wide_df = build_fold_metric_wide_df(fold_df, experiment_name)

    with pd.ExcelWriter(out_path) as writer:
        fold_wide_df.to_excel(writer, sheet_name="fold_values_wide", index=False)
        export_fold_df.to_excel(writer, sheet_name="fold_values_long", index=False)
        heatmap_df.to_excel(writer, sheet_name="fold_metric_matrix", index=False)
        df_numeric.to_excel(writer, sheet_name="metric_summary_numeric", index=False)
        df_csv.to_excel(writer, sheet_name="metric_summary_text", index=False)


def save_paper_tables_excel(output_tables, stats_tables, group_name, out_path):
    """Create one compact workbook per comparison group for manuscript tables."""
    summary_rows = []
    fold_long_rows = []
    fold_wide_rows = []
    kruskal_rows = []
    pairwise_rows = []

    for experiment_name in ["all_experiments"] + EXPERIMENTS:
        tables = output_tables.get((experiment_name, group_name))
        if tables is not None:
            summary = tables["numeric"].copy()
            if not summary.empty:
                summary.insert(0, "experiment_title", get_experiment_title(experiment_name))
                summary.insert(0, "experiment", experiment_name)
                summary_rows.append(summary)

            if experiment_name in EXPERIMENTS:
                fold_long = add_fold_labels(tables["fold_values"], experiment_name)
                if not fold_long.empty:
                    fold_long.insert(0, "experiment_title", get_experiment_title(experiment_name))
                    fold_long.insert(0, "experiment", experiment_name)
                    fold_long_rows.append(fold_long)

                fold_wide = build_fold_metric_wide_df(tables["fold_values"], experiment_name)
                if not fold_wide.empty:
                    fold_wide.insert(0, "experiment_title", get_experiment_title(experiment_name))
                    fold_wide.insert(0, "experiment", experiment_name)
                    fold_wide_rows.append(fold_wide)

        stats = stats_tables.get((experiment_name, group_name))
        if stats is not None:
            kruskal = stats["kruskal"].copy()
            if not kruskal.empty:
                kruskal["experiment_title"] = kruskal["experiment"].map(get_experiment_title)
                kruskal_rows.append(kruskal)

            pairwise = stats["pairwise"].copy()
            if not pairwise.empty:
                pairwise["experiment_title"] = pairwise["experiment"].map(get_experiment_title)
                pairwise_rows.append(pairwise)

    summary_df = pd.concat(summary_rows, ignore_index=True) if summary_rows else pd.DataFrame()
    fold_long_df = pd.concat(fold_long_rows, ignore_index=True) if fold_long_rows else pd.DataFrame()
    fold_wide_df = pd.concat(fold_wide_rows, ignore_index=True) if fold_wide_rows else pd.DataFrame()
    kruskal_df = pd.concat(kruskal_rows, ignore_index=True) if kruskal_rows else pd.DataFrame()
    pairwise_df = pd.concat(pairwise_rows, ignore_index=True) if pairwise_rows else pd.DataFrame()

    with pd.ExcelWriter(out_path) as writer:
        summary_df.to_excel(writer, sheet_name="metric_summary", index=False)
        fold_wide_df.to_excel(writer, sheet_name="fold_values_wide", index=False)
        fold_long_df.to_excel(writer, sheet_name="fold_values_long", index=False)
        kruskal_df.to_excel(writer, sheet_name="kruskal_tests", index=False)
        pairwise_df.to_excel(writer, sheet_name="wilcoxon_pairwise", index=False)
        if not kruskal_df.empty:
            kruskal_df[[
                "experiment", "experiment_title", "comparison_group", "metric",
                "n_models", "models", "effect_size", "effect_size_name",
            ]].to_excel(writer, sheet_name="kruskal_effect_sizes", index=False)
        if not pairwise_df.empty:
            pairwise_df[[
                "experiment", "experiment_title", "comparison_group", "metric",
                "comparison", "model_a", "model_b", "n_pairs", "effect_size",
                "effect_size_name", "mean_diff_a_minus_b",
            ]].to_excel(writer, sheet_name="pairwise_effect_sizes", index=False)


def get_result_table_split_label(experiment_name):
    return RESULT_TABLE_SPLIT_LABELS.get(experiment_name, get_experiment_title(experiment_name))


def get_fold_grid_row_label(experiment_name):
    """Compact row labels, aligned with the embedding-grid style."""
    return RESULT_TABLE_SPLIT_LABELS.get(experiment_name, get_experiment_title(experiment_name))


def format_result_metric_value(value, as_percent=False):
    if not np.isfinite(value):
        return ""
    if as_percent:
        return f"{float(value) * 100:.2f} %"
    return f"{float(value):.4f}"


def format_result_summary_value(values, as_percent=False):
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return ""
    mean_value = float(np.mean(arr))
    std_value = float(np.std(arr))
    if as_percent:
        return f"{mean_value * 100:.2f} % +/- {std_value * 100:.2f} %"
    return f"{mean_value:.4f} +/- {std_value:.4f}"


def build_results_template_rows(output_tables, group_name, summary_mode, as_percent=False):
    """Build manuscript-style result rows following the uploaded template layout.

    The template contains one row per model/split/fold with columns:
    Model, Split Strategy, Fold, Accuracy, DENV1-DENV4 F1.
    Rows are grouped by model and split strategy; repeated labels are left blank
    for a cleaner paper-ready layout, while the numeric companion table keeps all
    identifiers explicit for downstream analysis.
    """
    display_rows = []
    numeric_rows = []
    summary_rows = []

    models = COMPARISON_GROUPS[group_name]

    for model in models:
        model_display = display_model_name(model)
        model_started = False

        for experiment_name in RESULT_TABLE_EXPERIMENT_ORDER:
            tables = output_tables.get((experiment_name, group_name))
            if tables is None:
                continue

            fold_df = tables["fold_values"]
            if fold_df.empty or model not in set(fold_df["model"]):
                continue

            labeled = add_fold_labels(fold_df[fold_df["model"] == model], experiment_name)
            if labeled.empty:
                continue

            wide = labeled.pivot_table(
                index=["fold_index", "fold_label"],
                columns="metric",
                values="value",
                aggfunc="first",
            ).reset_index()
            wide.columns.name = None
            wide = wide.sort_values("fold_index")

            split_label = get_result_table_split_label(experiment_name)
            split_started = False

            for _, row in wide.iterrows():
                display_row = {
                    "Model": model_display if not model_started else "",
                    "Split Strategy": split_label if not split_started else "",
                    "Fold": row["fold_label"],
                }
                numeric_row = {
                    "model": model,
                    "model_display": model_display,
                    "experiment": experiment_name,
                    "experiment_title": get_experiment_title(experiment_name),
                    "split_strategy": split_label,
                    "fold_index": int(row["fold_index"]),
                    "fold": row["fold_label"],
                }

                for metric, out_col in RESULT_TABLE_METRIC_COLUMNS.items():
                    value = row.get(metric, np.nan)
                    value = float(value) if np.isfinite(value) else np.nan
                    display_row[out_col] = format_result_metric_value(value, as_percent=as_percent)
                    numeric_row[out_col] = value if np.isfinite(value) else np.nan

                display_rows.append(display_row)
                numeric_rows.append(numeric_row)
                model_started = True
                split_started = True

            summary_row = {
                "Model": model_display,
                "Split Strategy": split_label,
                "Fold": "Mean +/- SD",
            }
            for metric, out_col in RESULT_TABLE_METRIC_COLUMNS.items():
                vals = labeled.loc[labeled["metric"] == metric, "value"].to_numpy(dtype=float)
                summary_row[out_col] = format_result_summary_value(vals, as_percent=as_percent)
            summary_rows.append(summary_row)

    display_df = pd.DataFrame(display_rows, columns=RESULT_TABLE_COLUMNS)
    numeric_columns = [
        "model", "model_display", "experiment", "experiment_title",
        "split_strategy", "fold_index", "fold",
        "Accuracy", "DENV1", "DENV2", "DENV3", "DENV4",
    ]
    numeric_df = pd.DataFrame(numeric_rows, columns=numeric_columns)
    summary_df = pd.DataFrame(summary_rows, columns=RESULT_TABLE_COLUMNS)
    return display_df, numeric_df, summary_df


def _apply_results_template_style(workbook, sheet_name):
    """Style a worksheet like the uploaded results table template."""
    try:
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]
    max_row = max(ws.max_row, 2)

    # Template-like two-row header: F1 Score spans the four serotype columns.
    ws.insert_rows(2)
    ws["E1"] = "F1 Score"
    for cell in ("F1", "G1", "H1"):
        ws[cell] = None
    ws.merge_cells("E1:H1")
    for cell in ("A1", "B1", "C1", "D1"):
        ws.merge_cells(f"{cell}:{cell[0]}2")
    ws["E2"] = "DENV1"
    ws["F2"] = "DENV2"
    ws["G2"] = "DENV3"
    ws["H2"] = "DENV4"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    subheader_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill if cell.row == 1 else subheader_fill
            cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.column >= 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    for col_idx, width in enumerate([20, 18, 18, 12, 12, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{ws.max_row}"


def save_results_template_excel(output_tables, group_name, out_path, summary_mode, as_percent=False):
    """Export paper-ready result tables using the uploaded template structure."""
    display_df, numeric_df, summary_df = build_results_template_rows(
        output_tables,
        group_name,
        summary_mode,
        as_percent=as_percent,
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        display_df.to_excel(writer, sheet_name="fold_results_template", index=False)
        summary_df.to_excel(writer, sheet_name="summary_template", index=False)
        numeric_df.to_excel(writer, sheet_name="fold_results_numeric", index=False)

        workbook = writer.book
        _apply_results_template_style(workbook, "fold_results_template")
        _apply_results_template_style(workbook, "summary_template")

        # Keep the numeric sheet analysis-friendly, not presentation-styled.
        numeric_ws = workbook["fold_results_numeric"]
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            for cell in numeric_ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF3F8")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx in range(1, numeric_ws.max_column + 1):
                numeric_ws.column_dimensions[get_column_letter(col_idx)].width = 18
            for row in numeric_ws.iter_rows(min_row=2, min_col=8, max_col=12):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0000"
            numeric_ws.freeze_panes = "A2"
            numeric_ws.auto_filter.ref = numeric_ws.dimensions
        except ImportError:
            pass

def display_model_name(model_name):
    return DISPLAY_NAMES.get(model_name, model_name)


def display_comparison_name(comparison):
    if " vs " not in comparison:
        return comparison
    model_a, model_b = comparison.split(" vs ", 1)
    return f"{display_model_name(model_a)} vs {display_model_name(model_b)}"


def significance_stars(p_value):
    if not np.isfinite(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return "ns"




def kruskal_epsilon_squared(statistic, groups):
    """Epsilon-squared effect size for Kruskal-Wallis, clipped to [0, 1]."""
    if not np.isfinite(statistic):
        return np.nan
    valid_groups = []
    for values in groups:
        arr = np.asarray(values, dtype=float)
        arr = arr[np.isfinite(arr)]
        if arr.size > 0:
            valid_groups.append(arr)
    n_total = int(sum(arr.size for arr in valid_groups))
    k = len(valid_groups)
    if k < 2 or n_total <= k:
        return np.nan
    eps = (float(statistic) - k + 1.0) / (n_total - k)
    return float(np.clip(eps, 0.0, 1.0))


def paired_rank_biserial(values_a, values_b):
    """Matched-pairs rank-biserial correlation for paired model differences."""
    a = np.asarray(values_a, dtype=float)
    b = np.asarray(values_b, dtype=float)
    n = min(a.size, b.size)
    a = a[:n]
    b = b[:n]
    valid = np.isfinite(a) & np.isfinite(b)
    diff = a[valid] - b[valid]
    diff = diff[~np.isclose(diff, 0.0)]
    if diff.size == 0:
        return 0.0

    abs_diff = np.abs(diff)
    order = np.argsort(abs_diff, kind="mergesort")
    sorted_abs = abs_diff[order]
    ranks_sorted = np.empty_like(sorted_abs, dtype=float)

    start = 0
    while start < sorted_abs.size:
        end = start + 1
        while end < sorted_abs.size and np.isclose(sorted_abs[end], sorted_abs[start]):
            end += 1
        # Ranks are 1-based; ties receive the average rank.
        ranks_sorted[start:end] = (start + 1 + end) / 2.0
        start = end

    ranks = np.empty_like(ranks_sorted)
    ranks[order] = ranks_sorted
    positive_rank_sum = float(ranks[diff > 0].sum())
    negative_rank_sum = float(ranks[diff < 0].sum())
    total_rank_sum = positive_rank_sum + negative_rank_sum
    if total_rank_sum == 0:
        return np.nan
    return float((positive_rank_sum - negative_rank_sum) / total_rank_sum)
def safe_kruskal(groups):
    valid_groups = []
    for values in groups:
        arr = np.asarray(values, dtype=float)
        arr = arr[np.isfinite(arr)]
        if arr.size > 0:
            valid_groups.append(arr)

    if len(valid_groups) < 2:
        return np.nan, np.nan, 0, "not_enough_models"

    concatenated = np.concatenate(valid_groups)
    if concatenated.size == 0:
        return np.nan, np.nan, 0, "no_valid_values"

    if np.allclose(concatenated, concatenated[0]):
        return 0.0, 1.0, int(min(len(g) for g in valid_groups)), "all_values_identical"

    stat, p_value = kruskal(*valid_groups)
    return float(stat), float(p_value), int(min(len(g) for g in valid_groups)), "kruskal_wallis"


def safe_wilcoxon(values_a, values_b):
    a = np.asarray(values_a, dtype=float)
    b = np.asarray(values_b, dtype=float)

    n = min(a.size, b.size)
    a = a[:n]
    b = b[:n]

    valid = np.isfinite(a) & np.isfinite(b)
    a = a[valid]
    b = b[valid]

    if a.size < 2:
        return np.nan, np.nan, int(a.size), "not_enough_pairs"

    diff = a - b
    nonzero_diff = diff[~np.isclose(diff, 0.0)]

    if nonzero_diff.size == 0:
        return 0.0, 1.0, int(a.size), "all_differences_zero"

    if nonzero_diff.size < 2:
        return np.nan, np.nan, int(a.size), "too_few_nonzero_differences"

    # Use the exact Wilcoxon distribution for the small fold counts typical of
    # these experiments. For larger paired samples, SciPy can safely choose the
    # appropriate implementation with method="auto".
    method = "exact" if nonzero_diff.size <= 25 else "auto"
    stat, p_value = wilcoxon(
        nonzero_diff,
        zero_method="wilcox",
        alternative="two-sided",
        method=method,
    )
    return float(stat), float(p_value), int(a.size), f"wilcoxon_{method}"


def compute_statistical_tests(metric_values_by_model, experiment_name, group_name, models):
    validate_model_fold_alignment(
        metric_values_by_model,
        models,
        f"{experiment_name} / {group_name} statistical comparison",
    )
    present_models = [
        model for model in models
        if model in metric_values_by_model
    ]

    kruskal_rows = []
    pairwise_rows = []

    selected_vs_others = group_name == "selected_denformer_vs_others"

    for metric in METRICS:
        groups = [metric_values_by_model[model][metric] for model in present_models]
        stat, p_value, n_min, global_method = safe_kruskal(groups)
        effect_size = kruskal_epsilon_squared(stat, groups)
        kruskal_rows.append({
            "experiment": experiment_name,
            "comparison_group": group_name,
            "metric": metric,
            "n_models": len(present_models),
            "models": ",".join(present_models),
            "test": "Kruskal-Wallis",
            "n_min_per_model": n_min,
            "statistic": stat,
            "p_value": p_value,
            "effect_size": effect_size,
            "effect_size_name": "epsilon_squared",
            "method": global_method,
        })

        if selected_vs_others:
            comparison_pairs = [
                (SELECTED_DENFORMER, model)
                for model in present_models
                if model != SELECTED_DENFORMER
            ]
        else:
            comparison_pairs = [
                (model_a, model_b)
                for i, model_a in enumerate(present_models)
                for model_b in present_models[i + 1:]
            ]

        for model_a, model_b in comparison_pairs:
            values_a = metric_values_by_model[model_a][metric]
            values_b = metric_values_by_model[model_b][metric]
            pair_stat, pair_p, n_pairs, pairwise_method = safe_wilcoxon(values_a, values_b)
            pair_effect_size = paired_rank_biserial(values_a, values_b)
            pairwise_rows.append({
                "experiment": experiment_name,
                "comparison_group": group_name,
                "metric": metric,
                "model_a": model_a,
                "model_b": model_b,
                "comparison": f"{model_a} vs {model_b}",
                "test": "Wilcoxon signed-rank",
                "n_pairs": n_pairs,
                "statistic": pair_stat,
                "p_value": pair_p,
                "effect_size": pair_effect_size,
                "effect_size_name": "matched_pairs_rank_biserial",
                "method": pairwise_method,
                "mean_a": float(np.nanmean(values_a)) if len(values_a) else np.nan,
                "mean_b": float(np.nanmean(values_b)) if len(values_b) else np.nan,
                "mean_diff_a_minus_b": (
                    float(np.nanmean(values_a) - np.nanmean(values_b))
                    if len(values_a) and len(values_b)
                    else np.nan
                ),
            })

    kruskal_df = pd.DataFrame(kruskal_rows)
    pairwise_df = pd.DataFrame(pairwise_rows)

    if not kruskal_df.empty:
        valid = kruskal_df["p_value"].notna() & np.isfinite(kruskal_df["p_value"])
        kruskal_df["p_fdr_bh"] = np.nan
        if valid.any():
            _, q_values, _, _ = multipletests(
                kruskal_df.loc[valid, "p_value"].values,
                alpha=0.05,
                method="fdr_bh",
            )
            kruskal_df.loc[valid, "p_fdr_bh"] = q_values
        kruskal_df["significance"] = kruskal_df["p_fdr_bh"].apply(significance_stars)

    if not pairwise_df.empty:
        pairwise_df["p_fdr_bh"] = np.nan
        for metric in METRICS:
            mask = pairwise_df["metric"].eq(metric)
            valid = mask & pairwise_df["p_value"].notna() & np.isfinite(pairwise_df["p_value"])
            if valid.any():
                _, q_values, _, _ = multipletests(
                    pairwise_df.loc[valid, "p_value"].values,
                    alpha=0.05,
                    method="fdr_bh",
                )
                pairwise_df.loc[valid, "p_fdr_bh"] = q_values
        pairwise_df["significance"] = pairwise_df["p_fdr_bh"].apply(significance_stars)

    return kruskal_df, pairwise_df



def get_experiment_title(experiment_name):
    return EXPERIMENT_TITLES.get(experiment_name, experiment_name)


def get_group_title(group_name):
    return GROUP_TITLES.get(group_name, group_name)


def build_pairwise_heatmap_data(pairwise_df):
    comparisons = list(pairwise_df["comparison"].drop_duplicates())
    if pairwise_df.empty or not comparisons:
        return [], np.empty((0, len(METRICS))), np.empty((0, len(METRICS)), dtype=object)

    matrix = np.full((len(comparisons), len(METRICS)), np.nan, dtype=float)
    labels = np.full((len(comparisons), len(METRICS)), "", dtype=object)

    for i, comparison in enumerate(comparisons):
        for j, metric in enumerate(METRICS):
            row = pairwise_df[
                pairwise_df["comparison"].eq(comparison)
                & pairwise_df["metric"].eq(metric)
            ]
            if row.empty:
                continue
            q_value = float(row.iloc[0]["p_fdr_bh"])
            if np.isfinite(q_value):
                matrix[i, j] = -np.log10(max(q_value, 1e-300))
                labels[i, j] = significance_stars(q_value)

    return comparisons, matrix, labels


def draw_pairwise_pvalue_heatmap(ax, pairwise_df, title, show_xlabel=True, show_ylabel=True):
    comparisons, matrix, labels = build_pairwise_heatmap_data(pairwise_df)

    if not comparisons:
        ax.axis("off")
        ax.set_title(title, fontsize=HEATMAP_TITLE_SIZE, fontweight="bold")
        ax.text(0.5, 0.5, "No pairwise data", ha="center", va="center", transform=ax.transAxes)
        return None

    im = ax.imshow(matrix, aspect="auto", interpolation="nearest")

    ax.set_xticks(np.arange(len(METRICS)))
    ax.set_xticklabels(display_metric_names(METRICS), rotation=35, ha="right", fontsize=HEATMAP_TICK_SIZE)
    ax.set_yticks(np.arange(len(comparisons)))
    if show_ylabel:
        ax.set_yticklabels(
            [display_comparison_name(c) for c in comparisons],
            fontsize=HEATMAP_TICK_SIZE,
        )
        ax.set_ylabel("Pairwise comparison", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    else:
        ax.set_yticklabels([])
        ax.tick_params(axis="y", length=0)

    if show_xlabel:
        ax.set_xlabel("", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    else:
        ax.set_xlabel("")

    ax.set_title(title, fontsize=HEATMAP_TITLE_SIZE, fontweight="bold", pad=10)

    for i in range(len(comparisons)):
        for j in range(len(METRICS)):
            if labels[i, j]:
                ax.text(
                    j,
                    i,
                    labels[i, j],
                    ha="center",
                    va="center",
                    fontsize=HEATMAP_TEXT_SIZE,
                    color="black",
                )

    return im


def save_pairwise_pvalue_heatmap(pairwise_df, title, out_path):
    if pairwise_df.empty:
        return

    fig_height = max(5.2, 0.62 * max(1, pairwise_df["comparison"].nunique()) + 2.8)
    fig_width = max(10.5, 1.55 * len(METRICS) + 4.0)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    im = draw_pairwise_pvalue_heatmap(ax, pairwise_df, title)
    if im is not None:
        cbar = fig.colorbar(im, ax=ax)
        cbar.set_label("-log10(FDR-adjusted Wilcoxon p)", fontsize=HEATMAP_CBAR_SIZE)
        cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    plt.tight_layout()
    plt.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close()


def save_pairwise_pvalue_heatmap_panel(stats_tables, group_name, out_path):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_PANEL_ORDER
        if (experiment_name, group_name) in stats_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(2, 2, figsize=(18.0, 12.0))
    axes = axes.ravel()
    last_im = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        pairwise_df = stats_tables[(experiment_name, group_name)]["pairwise"]
        panel_title = f"{panel_label}. {get_experiment_title(experiment_name)}"
        im = draw_pairwise_pvalue_heatmap(
            ax,
            pairwise_df,
            panel_title,
            show_xlabel=panel_label in {"C", "D"},
            show_ylabel=panel_label in {"A", "C"},
        )
        if im is not None:
            last_im = im

    for ax in axes[len(available):]:
        ax.axis("off")

    if last_im is not None:
        cbar_ax = fig.add_axes([0.925, 0.18, 0.018, 0.64])
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("-log10(FDR-adjusted Wilcoxon p)", fontsize=HEATMAP_CBAR_SIZE)
        cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    fig.subplots_adjust(left=0.17, right=0.89, top=0.94, bottom=0.10, wspace=0.34, hspace=0.42)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_pairwise_pvalue_heatmap_row_panel(stats_tables, group_name, out_path):
    available = [
        (panel_label, experiment_name)
        for panel_label, experiment_name in EXPERIMENT_ROW_PANEL_ORDER
        if (experiment_name, group_name) in stats_tables
    ]
    if not available:
        return

    fig, axes = plt.subplots(1, len(available), figsize=(21.0, 6.2))
    axes = np.atleast_1d(axes).ravel()
    last_im = None

    for ax, (panel_label, experiment_name) in zip(axes, available):
        pairwise_df = stats_tables[(experiment_name, group_name)]["pairwise"]
        panel_title = f"{panel_label}. {get_experiment_title(experiment_name)}"
        im = draw_pairwise_pvalue_heatmap(
            ax,
            pairwise_df,
            panel_title,
            show_xlabel=False,
            show_ylabel=(panel_label == "A"),
        )
        if im is not None:
            last_im = im

    if last_im is not None:
        cbar_ax = fig.add_axes([0.925, 0.24, 0.014, 0.52])
        cbar = fig.colorbar(last_im, cax=cbar_ax)
        cbar.set_label("-log10(FDR-adjusted Wilcoxon p)", fontsize=HEATMAP_CBAR_SIZE)
        cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    fig.subplots_adjust(left=0.16, right=0.90, top=0.83, bottom=0.18, wspace=0.32)
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_kruskal_pvalue_heatmap_panel(stats_tables, group_name, out_path):
    rows = []
    for panel_label, experiment_name in EXPERIMENT_PANEL_ORDER:
        stats = stats_tables.get((experiment_name, group_name))
        if stats is None or stats["kruskal"].empty:
            continue
        for _, row in stats["kruskal"].iterrows():
            rows.append({
                "panel_label": panel_label,
                "experiment": experiment_name,
                "experiment_title": get_experiment_title(experiment_name),
                "metric": row["metric"],
                "q_value": row.get("p_fdr_bh", np.nan),
                "significance": row.get("significance", ""),
            })

    if not rows:
        return

    df = pd.DataFrame(rows)
    experiments = [experiment_name for _, experiment_name in EXPERIMENT_PANEL_ORDER]
    y_labels = [f"{panel}. {get_experiment_title(experiment)}" for panel, experiment in EXPERIMENT_PANEL_ORDER]
    matrix = np.full((len(experiments), len(METRICS)), np.nan, dtype=float)
    labels = np.full((len(experiments), len(METRICS)), "", dtype=object)

    for i, experiment in enumerate(experiments):
        for j, metric in enumerate(METRICS):
            row = df[df["experiment"].eq(experiment) & df["metric"].eq(metric)]
            if row.empty:
                continue
            q_value = float(row.iloc[0]["q_value"])
            if np.isfinite(q_value):
                matrix[i, j] = -np.log10(max(q_value, 1e-300))
                labels[i, j] = row.iloc[0]["significance"]

    fig, ax = plt.subplots(figsize=(12.5, 5.8))
    im = ax.imshow(matrix, aspect="auto", interpolation="nearest")
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("-log10(FDR-adjusted Kruskal-Wallis p)", fontsize=HEATMAP_CBAR_SIZE)
    cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    ax.set_xticks(np.arange(len(METRICS)))
    ax.set_xticklabels(display_metric_names(METRICS), rotation=35, ha="right", fontsize=HEATMAP_TICK_SIZE)
    ax.set_yticks(np.arange(len(experiments)))
    ax.set_yticklabels(y_labels, fontsize=HEATMAP_TICK_SIZE)
    ax.set_xlabel("", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    ax.set_ylabel("Evaluation protocol", fontsize=HEATMAP_AXIS_LABEL_SIZE)
    ax.set_title(
        f"{get_group_title(group_name)}: Kruskal-Wallis tests",
        fontsize=HEATMAP_TITLE_SIZE,
        fontweight="bold",
        pad=10,
    )

    for i in range(len(experiments)):
        for j in range(len(METRICS)):
            if labels[i, j]:
                ax.text(j, i, labels[i, j], ha="center", va="center", fontsize=HEATMAP_TEXT_SIZE)

    plt.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def save_kruskal_pvalue_heatmap_row_panel(stats_tables, group_name, out_path):
    rows = []
    for panel_label, experiment_name in EXPERIMENT_ROW_PANEL_ORDER:
        stats = stats_tables.get((experiment_name, group_name))
        if stats is None or stats["kruskal"].empty:
            continue
        for _, row in stats["kruskal"].iterrows():
            rows.append({
                "panel_label": panel_label,
                "experiment": experiment_name,
                "metric": row["metric"],
                "q_value": row.get("p_fdr_bh", np.nan),
                "significance": row.get("significance", ""),
            })

    if not rows:
        return

    df = pd.DataFrame(rows)
    experiments = [experiment_name for _, experiment_name in EXPERIMENT_ROW_PANEL_ORDER]
    y_labels = [f"{panel}. {get_experiment_title(experiment)}" for panel, experiment in EXPERIMENT_ROW_PANEL_ORDER]
    matrix = np.full((len(experiments), len(METRICS)), np.nan, dtype=float)
    labels = np.full((len(experiments), len(METRICS)), "", dtype=object)

    for i, experiment in enumerate(experiments):
        for j, metric in enumerate(METRICS):
            row = df[df["experiment"].eq(experiment) & df["metric"].eq(metric)]
            if row.empty:
                continue
            q_value = float(row.iloc[0]["q_value"])
            if np.isfinite(q_value):
                matrix[i, j] = -np.log10(max(q_value, 1e-300))
                labels[i, j] = row.iloc[0]["significance"]

    fig, ax = plt.subplots(figsize=(12.5, 4.6))
    im = ax.imshow(matrix, aspect="auto", interpolation="nearest")
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("-log10(FDR-adjusted Kruskal-Wallis p)", fontsize=HEATMAP_CBAR_SIZE)
    cbar.ax.tick_params(labelsize=HEATMAP_TICK_SIZE)

    ax.set_xticks(np.arange(len(METRICS)))
    ax.set_xticklabels(display_metric_names(METRICS), rotation=35, ha="right", fontsize=HEATMAP_TICK_SIZE)
    ax.set_yticks(np.arange(len(experiments)))
    ax.set_yticklabels(y_labels, fontsize=HEATMAP_TICK_SIZE)
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.set_title("Kruskal-Wallis tests", fontsize=HEATMAP_TITLE_SIZE, fontweight="bold", pad=10)

    for i in range(len(experiments)):
        for j in range(len(METRICS)):
            if labels[i, j]:
                ax.text(j, i, labels[i, j], ha="center", va="center", fontsize=HEATMAP_TEXT_SIZE)

    plt.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

def main():
    args = parse_args()
    error_bar_mode = args.error_bar
    summary_mode = args.summary
    result_table_as_percent = args.result_table_percent

    base_dir = Path(paths.logs_dir)
    out_dir = base_dir / "aggregate_metrics" / output_subdir_name(summary_mode, error_bar_mode)
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Using output directory: {out_dir}")
    result_tables_root_dir = base_dir / "aggregate_metrics" / "stats"
    result_tables_root_dir.mkdir(parents=True, exist_ok=True)
    print(f"Using manuscript table directory: {result_tables_root_dir}")

    result_tables_recap_dir = result_tables_root_dir / "recap"
    result_tables_recap_dir.mkdir(parents=True, exist_ok=True)
    print(f"Using manuscript table recap directory: {result_tables_recap_dir}")

    recap_1x3_dir = out_dir / RECAP_1X3_DIR_NAME
    recap_1x3_dir.mkdir(parents=True, exist_ok=True)
    print(f"Using recap 1x3 figure directory: {recap_1x3_dir}")

    experiment_metric_values_by_name = {}

    global_metric_values = {
        model: {
            FOLD_IDS_KEY: [],
            **{metric: [] for metric in METRICS},
        }
        for model in MODELS
    }

    for experiment in EXPERIMENTS:
        experiment_metric_values = {}

        for model in MODELS:
            summary_path = (
                base_dir
                / model
                / experiment
                / "metrics"
                / "inference_summary.json"
            )

            if not summary_path.exists():
                print(f"Missing: {summary_path}")
                continue

            with open(summary_path, "r") as f:
                summary_json = json.load(f)

            metric_values = extract_metric_values(summary_json, experiment)
            experiment_metric_values[model] = metric_values

            global_metric_values[model][FOLD_IDS_KEY].extend(
                f"{experiment}::{fold_id}"
                for fold_id in metric_values[FOLD_IDS_KEY]
            )
            for metric in METRICS:
                global_metric_values[model][metric].extend(metric_values[metric])

        if not experiment_metric_values:
            print(f"No data for {experiment}")
            continue

        experiment_metric_values_by_name[experiment] = experiment_metric_values

    if any(any(values[metric] for metric in METRICS) for values in global_metric_values.values()):
        experiment_metric_values_by_name["all_experiments"] = global_metric_values

    output_tables = {}
    stats_tables = {}

    for experiment, metric_values in experiment_metric_values_by_name.items():
        for group_name, models in COMPARISON_GROUPS.items():
            rows_csv, rows_numeric = build_rows_from_metric_values(metric_values, models, summary_mode)
            if not rows_csv:
                print(f"No data for {experiment} - {group_name}")
                continue

            output_tables[(experiment, group_name)] = {
                "csv": pd.DataFrame(rows_csv),
                "numeric": pd.DataFrame(rows_numeric),
                "fold_values": build_fold_values_df(metric_values, models, experiment),
            }

        for group_name, models in STATISTICAL_GROUPS.items():
            kruskal_df, pairwise_df = compute_statistical_tests(
                metric_values,
                experiment,
                group_name,
                models,
            )
            stats_tables[(experiment, group_name)] = {
                "kruskal": kruskal_df,
                "pairwise": pairwise_df,
            }

    all_numeric_dfs = [
        tables["numeric"]
        for tables in output_tables.values()
        if not tables["numeric"].empty
    ]

    if not all_numeric_dfs:
        print("No data found.")
        return

    y_min, y_max, y_ticks = compute_global_ylim_and_ticks(
        all_numeric_dfs,
        error_bar_mode,
        summary_mode,
        force_zero=args.barplot_y_from_zero,
    )

    for (experiment, group_name), tables in output_tables.items():
        group_out_dir = out_dir / group_name
        barplot_dir = group_out_dir / "barplot"
        dotplot_dir = group_out_dir / "dotplot"
        heatmap_dir = group_out_dir / "heatmap"
        stats_dir = group_out_dir / "stats"
        barplot_dir.mkdir(parents=True, exist_ok=True)
        dotplot_dir.mkdir(parents=True, exist_ok=True)
        heatmap_dir.mkdir(parents=True, exist_ok=True)
        stats_dir.mkdir(parents=True, exist_ok=True)

        df_csv = tables["csv"]
        df_numeric = tables["numeric"]
        fold_df = tables["fold_values"]

        csv_path = stats_dir / f"{experiment}_{group_name}_metrics.csv"
        df_csv.to_csv(csv_path, index=False, sep=";")
        print(f"Saved: {csv_path}")

        numeric_csv_path = stats_dir / f"{experiment}_{group_name}_metrics_numeric.csv"
        df_numeric.to_csv(numeric_csv_path, index=False, sep=";")
        print(f"Saved: {numeric_csv_path}")

        fold_values_path = stats_dir / f"{experiment}_{group_name}_fold_values_long.csv"
        fold_df.to_csv(fold_values_path, index=False, sep=";")
        print(f"Saved: {fold_values_path}")

        plot_title = get_experiment_title(experiment)

        barplot_path = barplot_dir / f"{experiment}_{group_name}_metrics_barplot.png"
        save_barplot(df_numeric, plot_title, barplot_path, y_min, y_max, y_ticks, error_bar_mode, summary_mode, axis_as_percent=result_table_as_percent)
        print(f"Saved: {barplot_path}")

        dotplot_path = dotplot_dir / f"{experiment}_{group_name}_metrics_dotplot.png"
        save_dotplot(fold_df, plot_title, dotplot_path, y_min, y_max, y_ticks, summary_mode, axis_as_percent=result_table_as_percent)
        print(f"Saved: {dotplot_path}")

        if experiment in EXPERIMENTS:
            fold_heatmap_path = heatmap_dir / f"{experiment}_{group_name}_fold_metric_heatmap.png"
            save_fold_metric_heatmap(fold_df, plot_title, fold_heatmap_path, group_name)
            if fold_heatmap_path.exists():
                print(f"Saved: {fold_heatmap_path}")

            fold_metric_excel_path = stats_dir / f"{experiment}_{group_name}_fold_metric_heatmap_data.xlsx"
            save_fold_metric_excel(fold_df, df_csv, df_numeric, fold_metric_excel_path, experiment)
            print(f"Saved: {fold_metric_excel_path}")

        stats = stats_tables.get((experiment, group_name))
        if stats is None:
            continue

        kruskal_path = stats_dir / f"{experiment}_{group_name}_stats_kruskal.csv"
        stats["kruskal"].to_csv(kruskal_path, index=False, sep=";")
        print(f"Saved: {kruskal_path}")

        pairwise_path = stats_dir / f"{experiment}_{group_name}_stats_pairwise_wilcoxon.csv"
        stats["pairwise"].to_csv(pairwise_path, index=False, sep=";")
        print(f"Saved: {pairwise_path}")

        heatmap_path = heatmap_dir / f"{experiment}_{group_name}_stats_pairwise_wilcoxon_heatmap.png"
        save_pairwise_pvalue_heatmap(stats["pairwise"], plot_title, heatmap_path)
        if heatmap_path.exists():
            print(f"Saved: {heatmap_path}")

    for group_name in COMPARISON_GROUPS:
        group_out_dir = out_dir / group_name
        barplot_dir = group_out_dir / "barplot"
        dotplot_dir = group_out_dir / "dotplot"
        heatmap_dir = group_out_dir / "heatmap"
        stats_dir = group_out_dir / "stats"
        barplot_dir.mkdir(parents=True, exist_ok=True)
        dotplot_dir.mkdir(parents=True, exist_ok=True)
        heatmap_dir.mkdir(parents=True, exist_ok=True)
        stats_dir.mkdir(parents=True, exist_ok=True)

        barplot_panel_path = barplot_dir / f"all_protocols_{group_name}_metrics_barplot_panel.png"
        save_barplot_panel(
            output_tables,
            group_name,
            barplot_panel_path,
            y_min,
            y_max,
            y_ticks,
            error_bar_mode,
            summary_mode,
            axis_as_percent=result_table_as_percent,
        )
        if barplot_panel_path.exists():
            print(f"Saved: {barplot_panel_path}")

        barplot_row_panel_path = barplot_dir / f"split_protocols_{group_name}_metrics_barplot_panel_1x3.png"
        save_barplot_row_panel(
            output_tables,
            group_name,
            barplot_row_panel_path,
            y_min,
            y_max,
            y_ticks,
            error_bar_mode,
            summary_mode,
            axis_as_percent=result_table_as_percent,
        )
        if barplot_row_panel_path.exists():
            print(f"Saved: {barplot_row_panel_path}")
            recap_path = copy_figure_for_paper(
                barplot_row_panel_path,
                recap_1x3_dir,
                f"{group_name}_metrics_barplot_panel_1x3.png",
            )
            if recap_path is not None:
                print(f"Saved: {recap_path}")

        manuscript_stats_dir = result_tables_root_dir / group_name
        manuscript_stats_dir.mkdir(parents=True, exist_ok=True)

        manuscript_barplot_dir = manuscript_stats_dir / "barplot"
        manuscript_barplot_dir.mkdir(parents=True, exist_ok=True)

        fold_y_min, fold_y_max, fold_y_ticks = compute_fold_barplot_grid_ylim_and_ticks(
            output_tables,
            group_name,
            force_zero=args.barplot_y_from_zero,
        )

        fold_barplot_grid_path = manuscript_barplot_dir / f"split_protocols_{group_name}_metrics_barplot_fold_grid.png"
        save_fold_barplot_grid(
            output_tables,
            group_name,
            fold_barplot_grid_path,
            fold_y_min,
            fold_y_max,
            fold_y_ticks,
            axis_as_percent=result_table_as_percent,
        )
        if fold_barplot_grid_path.exists():
            print(f"Saved: {fold_barplot_grid_path}")

        paper_excel_path = manuscript_stats_dir / f"paper_tables_{group_name}.xlsx"
        save_paper_tables_excel(output_tables, stats_tables, group_name, paper_excel_path)
        if paper_excel_path.exists():
            print(f"Saved: {paper_excel_path}")

        results_template_excel_path = manuscript_stats_dir / f"results_template_{group_name}.xlsx"
        save_results_template_excel(
            output_tables,
            group_name,
            results_template_excel_path,
            summary_mode,
            as_percent=result_table_as_percent,
        )
        if results_template_excel_path.exists():
            print(f"Saved: {results_template_excel_path}")
            recap_results_template_path = result_tables_recap_dir / results_template_excel_path.name
            shutil.copy2(results_template_excel_path, recap_results_template_path)
            print(f"Saved: {recap_results_template_path}")

        dotplot_panel_path = dotplot_dir / f"all_protocols_{group_name}_metrics_dotplot_panel.png"
        save_dotplot_panel(
            output_tables,
            group_name,
            dotplot_panel_path,
            y_min,
            y_max,
            y_ticks,
            summary_mode,
            axis_as_percent=result_table_as_percent,
        )
        if dotplot_panel_path.exists():
            print(f"Saved: {dotplot_panel_path}")

        dotplot_row_panel_path = dotplot_dir / f"split_protocols_{group_name}_metrics_dotplot_panel_1x3.png"
        save_dotplot_row_panel(
            output_tables,
            group_name,
            dotplot_row_panel_path,
            y_min,
            y_max,
            y_ticks,
            summary_mode,
            axis_as_percent=result_table_as_percent,
        )
        if dotplot_row_panel_path.exists():
            print(f"Saved: {dotplot_row_panel_path}")

    for group_name in STATISTICAL_GROUPS:
        stats_dir = out_dir / group_name / "stats"
        heatmap_dir = out_dir / group_name / "heatmap"
        stats_dir.mkdir(parents=True, exist_ok=True)
        heatmap_dir.mkdir(parents=True, exist_ok=True)

        panel_path = heatmap_dir / f"all_protocols_{group_name}_stats_pairwise_wilcoxon_heatmap_panel.png"
        save_pairwise_pvalue_heatmap_panel(stats_tables, group_name, panel_path)
        if panel_path.exists():
            print(f"Saved: {panel_path}")

        row_panel_path = heatmap_dir / f"split_protocols_{group_name}_stats_pairwise_wilcoxon_heatmap_panel_1x3.png"
        save_pairwise_pvalue_heatmap_row_panel(stats_tables, group_name, row_panel_path)
        if row_panel_path.exists():
            print(f"Saved: {row_panel_path}")

        kruskal_panel_path = heatmap_dir / f"all_protocols_{group_name}_stats_kruskal_heatmap_panel.png"
        save_kruskal_pvalue_heatmap_panel(stats_tables, group_name, kruskal_panel_path)
        if kruskal_panel_path.exists():
            print(f"Saved: {kruskal_panel_path}")

        kruskal_row_panel_path = heatmap_dir / f"split_protocols_{group_name}_stats_kruskal_heatmap_panel_1x3.png"
        save_kruskal_pvalue_heatmap_row_panel(stats_tables, group_name, kruskal_row_panel_path)
        if kruskal_row_panel_path.exists():
            print(f"Saved: {kruskal_row_panel_path}")


if __name__ == "__main__":
    main()

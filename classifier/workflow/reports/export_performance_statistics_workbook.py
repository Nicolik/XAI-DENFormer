"""Rebuild the Supplementary File S3 workbook from cached CSV outputs."""

from __future__ import annotations

try:
    from ._bootstrap import PROJECT_ROOT  # noqa: F401
except ImportError:
    from _bootstrap import PROJECT_ROOT  # noqa: F401

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from classifier.workflow.reports.make_performance_statistics import (
    DEFAULT_OUTPUT_DIR,
    _write_workbook,
)
from classifier.workflow.reports.performance_statistics import AnalysisOutputs


DEFAULT_WORKBOOK_NAME = "Supplementary_File_S3_Model_Performance_Statistics.xlsx"

CSV_FILES = {
    "primary_accuracy": "primary_accuracy.csv",
    "macro_f1": "macro_f1.csv",
    "per_class_f1": "per_class_f1.csv",
    "model_summary": "model_summary.csv",
    "qc_alignment": "qc_alignment.csv",
}
PARAMETERS_CSV_NAME = "methods_parameters.csv"
PARAMETERS_SHEET_NAME = "Methods_parameters"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rebuild Supplementary File S3 from cached CSV tables without "
            "rerunning bootstrap or permutation analyses."
        )
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Performance-statistics output directory containing the csv/ folder.",
    )
    parser.add_argument(
        "--csv-dir",
        type=Path,
        default=None,
        help="Optional CSV directory. Defaults to <output-dir>/csv.",
    )
    parser.add_argument(
        "--workbook-path",
        type=Path,
        default=None,
        help=(
            "Destination workbook. Defaults to "
            "<output-dir>/Supplementary_File_S3_Model_Performance_Statistics.xlsx."
        ),
    )
    parser.add_argument(
        "--source-workbook",
        type=Path,
        default=None,
        help=(
            "Existing workbook used only as a fallback source for the "
            "Methods_parameters sheet when methods_parameters.csv is absent. "
            "Defaults to the destination workbook."
        ),
    )
    return parser.parse_args()


def _read_required_csv(csv_dir: Path, filename: str) -> pd.DataFrame:
    path = csv_dir / filename
    if not path.is_file():
        raise FileNotFoundError(
            f"Required cached statistics file not found: {path}. "
            "Run the full performance-statistics analysis once before using export-only mode."
        )
    dataframe = pd.read_csv(path)
    if dataframe.empty:
        raise ValueError(f"Cached statistics file is empty: {path}")
    return dataframe


def _load_parameters(csv_dir: Path, source_workbook: Path) -> pd.DataFrame:
    parameters_csv = csv_dir / PARAMETERS_CSV_NAME
    if parameters_csv.is_file():
        parameters = pd.read_csv(parameters_csv)
        if parameters.empty:
            raise ValueError(f"Cached methods-parameter file is empty: {parameters_csv}")
        return parameters

    if source_workbook.is_file():
        # Read the sheet directly with openpyxl instead of pandas.read_excel().
        # Pandas 2.2 requires openpyxl >= 3.1 for Excel reads, while the project
        # environment may still use openpyxl 3.0.x. The underlying openpyxl API
        # used here is compatible with both versions.
        workbook = load_workbook(source_workbook, read_only=True, data_only=True)
        try:
            if PARAMETERS_SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"Workbook {source_workbook} does not contain the "
                    f"{PARAMETERS_SHEET_NAME!r} sheet."
                )
            worksheet = workbook[PARAMETERS_SHEET_NAME]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows or not rows[0]:
            raise ValueError(
                f"Workbook sheet {PARAMETERS_SHEET_NAME!r} is empty: {source_workbook}"
            )
        headers = [str(value) if value is not None else "" for value in rows[0]]
        parameters = pd.DataFrame(rows[1:], columns=headers).dropna(how="all")
        if parameters.empty:
            raise ValueError(
                f"Workbook sheet {PARAMETERS_SHEET_NAME!r} is empty: {source_workbook}"
            )
        parameters_csv.parent.mkdir(parents=True, exist_ok=True)
        parameters.to_csv(parameters_csv, index=False)
        return parameters

    raise FileNotFoundError(
        f"Neither {parameters_csv} nor an existing source workbook was found. "
        "Keep the current Supplementary File S3 workbook for the first export-only run, "
        "or rerun the full analysis once to create methods_parameters.csv."
    )


def load_cached_outputs(csv_dir: Path, source_workbook: Path) -> AnalysisOutputs:
    """Load all workbook inputs from cached CSVs and existing method metadata."""
    csv_dir = Path(csv_dir)
    source_workbook = Path(source_workbook)
    loaded = {
        field: _read_required_csv(csv_dir, filename)
        for field, filename in CSV_FILES.items()
    }
    parameters = _load_parameters(csv_dir, source_workbook)
    return AnalysisOutputs(parameters=parameters, **loaded)


def export_workbook(
    output_dir: Path,
    csv_dir: Path | None = None,
    workbook_path: Path | None = None,
    source_workbook: Path | None = None,
) -> Path:
    """Write Supplementary File S3 without recomputing any statistics."""
    output_dir = Path(output_dir)
    csv_dir = Path(csv_dir) if csv_dir is not None else output_dir / "csv"
    workbook_path = (
        Path(workbook_path)
        if workbook_path is not None
        else output_dir / DEFAULT_WORKBOOK_NAME
    )
    source_workbook = (
        Path(source_workbook) if source_workbook is not None else workbook_path
    )

    outputs = load_cached_outputs(csv_dir=csv_dir, source_workbook=source_workbook)
    _write_workbook(outputs, workbook_path)
    return workbook_path


def main() -> None:
    args = parse_args()
    workbook_path = export_workbook(
        output_dir=args.output_dir,
        csv_dir=args.csv_dir,
        workbook_path=args.workbook_path,
        source_workbook=args.source_workbook,
    )
    print("Loaded cached statistical tables; bootstrap and permutation analyses were not rerun.")
    print(f"Saved: {workbook_path}")


if __name__ == "__main__":
    main()

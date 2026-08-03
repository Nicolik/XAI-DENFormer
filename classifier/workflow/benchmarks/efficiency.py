"""Controlled training and inference efficiency benchmark.

The benchmark runs every classifier configuration on the same precomputed fold,
software environment, batch size, and CUDA allocation. Training uses one full
warm-up epoch followed by measured epochs. Inference loads the original model
checkpoints, performs batch warm-up, and then measures repeated complete test
passes.
"""

from __future__ import annotations

import argparse
import gc
import json
import os
import platform
import random
import subprocess
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterable

import numpy as np
import pandas as pd
import torch
from torch.utils.data import DataLoader, WeightedRandomSampler

import paths
from classifier.data import DengueDataset
from classifier.utils_data import get_dataset
from classifier.workflow import config
from classifier.workflow.utils import (
    build_classifier_model,
    build_model_dir,
    get_latest_model_path,
    load_and_validate_folds,
    safe_name,
)


MODEL_CONFIGS = (
    {"name": "logreg", "display_name": "Logistic Regression", "model_type": "logreg", "pooling": "mean"},
    {"name": "ffnn", "display_name": "FFNN", "model_type": "ffnn", "pooling": "mean"},
    {"name": "denformer_first", "display_name": "DENFormer-first", "model_type": "denformer", "pooling": "first"},
    {"name": "denformer_mean", "display_name": "DENFormer-mean", "model_type": "denformer", "pooling": "mean"},
    {"name": "denformer_max", "display_name": "DENFormer-max", "model_type": "denformer", "pooling": "max"},
    {"name": "longformer", "display_name": "Longformer", "model_type": "longformer", "pooling": "mean"},
    {"name": "performer", "display_name": "Performer", "model_type": "performer", "pooling": "mean"},
)

DEFAULT_OUTPUT_ROOT = Path(paths.data_dir) / "benchmarks" / "efficiency"
DEFAULT_SPLIT_FILE = Path(paths.split_files["cdhit"])
DEFAULT_FOLD = "fold_1"
DEFAULT_CHECKPOINT_RUN_NAME = "cdhit"
DEFAULT_CHECKPOINT_EPOCHS = 100
DEFAULT_WARMUP_EPOCHS = 1
DEFAULT_MEASURED_EPOCHS = 3
DEFAULT_INFERENCE_WARMUP_BATCHES = 10
DEFAULT_INFERENCE_REPEATS = 3
DEFAULT_PAPER_XLSX_NAME = "efficiency_paper_table.xlsx"
BYTES_PER_GIB = 1024 ** 3

REQUIRED_EXISTING_RESULT_FILES = (
    "training_measurements.csv",
    "inference_measurements.csv",
    "efficiency_summary.csv",
)


@dataclass
class EpochMeasurement:
    model: str
    display_name: str
    model_type: str
    pooling: str
    fold: str
    phase: str
    repeat: int
    train_time_sec: float
    validation_time_sec: float
    total_time_sec: float
    train_loss: float
    train_accuracy: float
    validation_loss: float
    validation_accuracy: float
    peak_memory_allocated_gb: float | None
    peak_memory_reserved_gb: float | None


@dataclass
class InferenceMeasurement:
    model: str
    display_name: str
    model_type: str
    pooling: str
    fold: str
    repeat: int
    total_time_sec: float
    num_samples: int
    samples_per_sec: float
    latency_ms_per_sample: float
    accuracy: float
    peak_memory_allocated_gb: float | None
    peak_memory_reserved_gb: float | None
    checkpoint_path: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Benchmark training time, inference throughput, and peak CUDA memory "
            "for all classifier models under one controlled hardware allocation."
        )
    )
    parser.add_argument("--split-file", type=Path, default=DEFAULT_SPLIT_FILE)
    parser.add_argument("--fold", default=DEFAULT_FOLD)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--batch-size", type=int, default=config.BATCH_SIZE)
    parser.add_argument("--warmup-epochs", type=int, default=DEFAULT_WARMUP_EPOCHS)
    parser.add_argument("--measured-epochs", type=int, default=DEFAULT_MEASURED_EPOCHS)
    parser.add_argument(
        "--inference-warmup-batches",
        type=int,
        default=DEFAULT_INFERENCE_WARMUP_BATCHES,
    )
    parser.add_argument("--inference-repeats", type=int, default=DEFAULT_INFERENCE_REPEATS)
    parser.add_argument("--checkpoint-run-name", default=DEFAULT_CHECKPOINT_RUN_NAME)
    parser.add_argument("--checkpoint-epochs", type=int, default=DEFAULT_CHECKPOINT_EPOCHS)
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--log-every", type=int, default=100)
    parser.add_argument(
        "--source-run",
        type=Path,
        default=None,
        help=(
            "Existing completed benchmark run to reuse. If omitted, the latest "
            "completed run under --output-root is detected automatically."
        ),
    )
    parser.add_argument(
        "--table-only",
        action="store_true",
        help=(
            "Require an existing completed run and export the paper table without "
            "running training or inference."
        ),
    )
    parser.add_argument(
        "--force-benchmark",
        action="store_true",
        help="Run the benchmark even when reusable completed results already exist.",
    )
    parser.add_argument(
        "--paper-xlsx-name",
        default=DEFAULT_PAPER_XLSX_NAME,
        help="Filename for the formatted Excel paper table.",
    )
    parser.add_argument(
        "--models",
        nargs="+",
        choices=[model["name"] for model in MODEL_CONFIGS],
        default=[model["name"] for model in MODEL_CONFIGS],
        help="Subset of model configurations to benchmark. Default: all models.",
    )
    args = parser.parse_args()

    if args.batch_size <= 0:
        parser.error("--batch-size must be positive")
    if args.warmup_epochs < 0:
        parser.error("--warmup-epochs cannot be negative")
    if args.measured_epochs <= 0:
        parser.error("--measured-epochs must be positive")
    if args.inference_warmup_batches < 0:
        parser.error("--inference-warmup-batches cannot be negative")
    if args.inference_repeats <= 0:
        parser.error("--inference-repeats must be positive")
    if args.checkpoint_epochs <= 0:
        parser.error("--checkpoint-epochs must be positive")
    if args.log_every < 0:
        parser.error("--log-every cannot be negative")
    if args.table_only and args.force_benchmark:
        parser.error("--table-only and --force-benchmark cannot be used together")
    if Path(args.paper_xlsx_name).name != args.paper_xlsx_name:
        parser.error("--paper-xlsx-name must be a filename without directory components")
    if not args.paper_xlsx_name.lower().endswith(".xlsx"):
        parser.error("--paper-xlsx-name must end with .xlsx")
    return args


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def make_run_dir(output_root: Path) -> Path:
    job_id = os.environ.get("SLURM_JOB_ID", "local")
    run_dir = output_root / f"run_{utc_timestamp()}_job_{job_id}"
    run_dir.mkdir(parents=True, exist_ok=False)
    output_root.mkdir(parents=True, exist_ok=True)
    return run_dir


def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def synchronize(device: torch.device) -> None:
    if device.type == "cuda":
        torch.cuda.synchronize(device)


def reset_peak_memory(device: torch.device) -> None:
    if device.type == "cuda":
        synchronize(device)
        torch.cuda.reset_peak_memory_stats(device)


def peak_memory_gb(device: torch.device) -> tuple[float | None, float | None]:
    if device.type != "cuda":
        return None, None
    synchronize(device)
    allocated = torch.cuda.max_memory_allocated(device) / BYTES_PER_GIB
    reserved = torch.cuda.max_memory_reserved(device) / BYTES_PER_GIB
    return float(allocated), float(reserved)


def cleanup_cuda() -> None:
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()


def get_nvidia_smi() -> str | None:
    command = [
        "nvidia-smi",
        "--query-gpu=index,name,uuid,driver_version,memory.total,compute_cap",
        "--format=csv,noheader",
    ]
    try:
        completed = subprocess.run(command, check=True, capture_output=True, text=True)
    except (FileNotFoundError, subprocess.CalledProcessError):
        return None
    return completed.stdout.strip() or None


def get_device_info(device: torch.device) -> dict[str, Any]:
    info: dict[str, Any] = {
        "device": str(device),
        "cuda_available": bool(torch.cuda.is_available()),
        "torch_version": torch.__version__,
        "cuda_version": torch.version.cuda,
        "cudnn_version": torch.backends.cudnn.version(),
        "python_version": platform.python_version(),
        "platform": platform.platform(),
        "hostname": platform.node(),
        "nvidia_smi": get_nvidia_smi(),
        "gpu_allocation": os.environ.get("DENFORMER_BENCHMARK_GPU_ALLOCATION"),
    }
    if device.type == "cuda":
        idx = torch.cuda.current_device()
        props = torch.cuda.get_device_properties(idx)
        info.update(
            {
                "gpu_index": int(idx),
                "gpu_name": torch.cuda.get_device_name(idx),
                "gpu_total_memory_gb": float(props.total_memory / BYTES_PER_GIB),
                "gpu_compute_capability": f"{props.major}.{props.minor}",
                "gpu_multiprocessor_count": int(props.multi_processor_count),
            }
        )
    return info


def tracked_environment() -> dict[str, str | None]:
    names = (
        "SLURM_JOB_ID",
        "SLURM_JOB_NAME",
        "SLURM_JOB_PARTITION",
        "SLURM_CPUS_PER_TASK",
        "SLURM_MEM_PER_NODE",
        "SLURM_JOB_GPUS",
        "SLURM_STEP_GPUS",
        "CUDA_VISIBLE_DEVICES",
        "OMP_NUM_THREADS",
        "MKL_NUM_THREADS",
        "DENFORMER_BENCHMARK_GPU_ALLOCATION",
    )
    return {name: os.environ.get(name) for name in names}


def model_args(model_info: dict[str, str]) -> SimpleNamespace:
    return SimpleNamespace(
        model_type=model_info["model_type"],
        pooling=model_info["pooling"],
    )


def count_parameters(model: torch.nn.Module) -> tuple[int, int]:
    total = sum(parameter.numel() for parameter in model.parameters())
    trainable = sum(parameter.numel() for parameter in model.parameters() if parameter.requires_grad)
    return int(total), int(trainable)


def make_weighted_train_loader(
    dataset: DengueDataset,
    batch_size: int,
    seed: int,
) -> DataLoader:
    indices = np.asarray(dataset.indices, dtype=np.int64)
    labels = np.asarray(dataset.targets)[indices].astype(np.int64)
    classes, counts = np.unique(labels, return_counts=True)
    class_weights = {
        int(label): 1.0 - (float(count) / float(len(labels)))
        for label, count in zip(classes, counts)
    }
    sample_weights = torch.tensor(
        [class_weights[int(label)] for label in labels],
        dtype=torch.double,
    )
    generator = torch.Generator()
    generator.manual_seed(seed)
    sampler = WeightedRandomSampler(
        weights=sample_weights,
        num_samples=len(dataset),
        replacement=True,
        generator=generator,
    )
    return DataLoader(
        dataset,
        batch_size=batch_size,
        sampler=sampler,
        shuffle=False,
        num_workers=0,
        pin_memory=False,
    )


def make_eval_loader(dataset: DengueDataset, batch_size: int) -> DataLoader:
    return DataLoader(
        dataset,
        batch_size=batch_size,
        shuffle=False,
        num_workers=0,
        pin_memory=False,
    )


def run_train_pass(
    model: torch.nn.Module,
    loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    loss_fn: torch.nn.Module,
    device: torch.device,
    log_every: int,
) -> tuple[float, float]:
    model.train()
    loss_sum = torch.zeros((), device=device, dtype=torch.float64)
    correct = torch.zeros((), device=device, dtype=torch.long)
    sample_count = 0

    for batch_idx, (inputs, labels) in enumerate(loader, start=1):
        inputs = inputs.to(device)
        labels = labels.to(device)
        optimizer.zero_grad(set_to_none=True)
        outputs = model(inputs)
        loss = loss_fn(outputs, labels)
        loss.backward()
        optimizer.step()

        batch_size = int(labels.shape[0])
        loss_sum += loss.detach().to(torch.float64) * batch_size
        correct += (outputs.argmax(dim=1) == labels).sum()
        sample_count += batch_size

        if log_every and (batch_idx % log_every == 0 or batch_idx == len(loader)):
            print(f"    training batch {batch_idx}/{len(loader)}", flush=True)

    synchronize(device)
    return float(loss_sum.item() / sample_count), float(correct.item() / sample_count)


def run_eval_pass(
    model: torch.nn.Module,
    loader: DataLoader,
    loss_fn: torch.nn.Module,
    device: torch.device,
) -> tuple[float, float]:
    model.eval()
    loss_sum = torch.zeros((), device=device, dtype=torch.float64)
    correct = torch.zeros((), device=device, dtype=torch.long)
    sample_count = 0

    with torch.no_grad():
        for inputs, labels in loader:
            inputs = inputs.to(device)
            labels = labels.to(device)
            outputs = model(inputs)
            loss = loss_fn(outputs, labels)

            batch_size = int(labels.shape[0])
            loss_sum += loss.detach().to(torch.float64) * batch_size
            correct += (outputs.argmax(dim=1) == labels).sum()
            sample_count += batch_size

    synchronize(device)
    return float(loss_sum.item() / sample_count), float(correct.item() / sample_count)


def run_training_epoch(
    model_info: dict[str, str],
    model: torch.nn.Module,
    train_loader: DataLoader,
    validation_loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    loss_fn: torch.nn.Module,
    device: torch.device,
    fold: str,
    phase: str,
    repeat: int,
    log_every: int,
) -> EpochMeasurement:
    reset_peak_memory(device)
    synchronize(device)
    total_start = time.perf_counter()

    train_start = time.perf_counter()
    train_loss, train_accuracy = run_train_pass(
        model=model,
        loader=train_loader,
        optimizer=optimizer,
        loss_fn=loss_fn,
        device=device,
        log_every=log_every,
    )
    synchronize(device)
    train_time = time.perf_counter() - train_start

    validation_start = time.perf_counter()
    validation_loss, validation_accuracy = run_eval_pass(
        model=model,
        loader=validation_loader,
        loss_fn=loss_fn,
        device=device,
    )
    synchronize(device)
    validation_time = time.perf_counter() - validation_start
    total_time = time.perf_counter() - total_start
    peak_allocated, peak_reserved = peak_memory_gb(device)

    return EpochMeasurement(
        model=model_info["name"],
        display_name=model_info["display_name"],
        model_type=model_info["model_type"],
        pooling=model_info["pooling"],
        fold=fold,
        phase=phase,
        repeat=repeat,
        train_time_sec=float(train_time),
        validation_time_sec=float(validation_time),
        total_time_sec=float(total_time),
        train_loss=float(train_loss),
        train_accuracy=float(train_accuracy),
        validation_loss=float(validation_loss),
        validation_accuracy=float(validation_accuracy),
        peak_memory_allocated_gb=peak_allocated,
        peak_memory_reserved_gb=peak_reserved,
    )


def warmup_inference(
    model: torch.nn.Module,
    loader: DataLoader,
    device: torch.device,
    warmup_batches: int,
) -> int:
    if warmup_batches <= 0:
        return 0
    model.eval()
    completed = 0
    with torch.no_grad():
        for inputs, _ in loader:
            if completed >= warmup_batches:
                break
            inputs = inputs.to(device)
            outputs = model(inputs)
            probabilities = torch.softmax(outputs, dim=1)
            _ = probabilities.argmax(dim=1)
            completed += 1
    synchronize(device)
    return completed


def run_inference_pass(
    model_info: dict[str, str],
    model: torch.nn.Module,
    loader: DataLoader,
    device: torch.device,
    fold: str,
    repeat: int,
    checkpoint_path: Path,
) -> InferenceMeasurement:
    model.eval()
    reset_peak_memory(device)
    correct = torch.zeros((), device=device, dtype=torch.long)
    sample_count = 0

    synchronize(device)
    start = time.perf_counter()
    with torch.no_grad():
        for inputs, labels in loader:
            inputs = inputs.to(device)
            labels = labels.to(device)
            outputs = model(inputs)
            probabilities = torch.softmax(outputs, dim=1)
            predictions = probabilities.argmax(dim=1)
            correct += (predictions == labels).sum()
            sample_count += int(labels.shape[0])
    synchronize(device)
    elapsed = time.perf_counter() - start
    correct_count = int(correct.item())
    peak_allocated, peak_reserved = peak_memory_gb(device)

    return InferenceMeasurement(
        model=model_info["name"],
        display_name=model_info["display_name"],
        model_type=model_info["model_type"],
        pooling=model_info["pooling"],
        fold=fold,
        repeat=repeat,
        total_time_sec=float(elapsed),
        num_samples=int(sample_count),
        samples_per_sec=float(sample_count / elapsed),
        latency_ms_per_sample=float((elapsed / sample_count) * 1000.0),
        accuracy=float(correct_count / sample_count),
        peak_memory_allocated_gb=peak_allocated,
        peak_memory_reserved_gb=peak_reserved,
        checkpoint_path=str(checkpoint_path),
    )


def checkpoint_for_model(
    model_info: dict[str, str],
    fold: str,
    checkpoint_run_name: str,
    checkpoint_epochs: int,
) -> Path:
    run_suffix = f"_{checkpoint_run_name}" if checkpoint_run_name else ""
    model_dir = build_model_dir(
        output_dir=paths.logs_dir,
        model_type=model_info["model_type"],
        pooling=model_info["pooling"],
        k_type="ohe",
        run_suffix=run_suffix,
        epochs=checkpoint_epochs,
    )
    fold_model_dir = Path(model_dir) / f"split_{safe_name(fold)}"
    checkpoint = get_latest_model_path(str(fold_model_dir))
    if checkpoint is None:
        raise FileNotFoundError(
            f"No checkpoint found for {model_info['name']} in {fold_model_dir}. "
            "The inference benchmark requires the original trained checkpoint."
        )
    return Path(checkpoint)


def summarize_numeric(values: Iterable[float]) -> dict[str, float]:
    array = np.asarray(list(values), dtype=float)
    array = array[np.isfinite(array)]
    if array.size == 0:
        return {"mean": np.nan, "median": np.nan, "std": np.nan, "min": np.nan, "max": np.nan}
    return {
        "mean": float(np.mean(array)),
        "median": float(np.median(array)),
        "std": float(np.std(array)),
        "min": float(np.min(array)),
        "max": float(np.max(array)),
    }


def add_stats(row: dict[str, Any], prefix: str, values: Iterable[float]) -> None:
    for statistic, value in summarize_numeric(values).items():
        row[f"{prefix}_{statistic}"] = value


def build_summary(
    epoch_df: pd.DataFrame,
    inference_df: pd.DataFrame,
    parameter_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    params_df = pd.DataFrame(parameter_rows)

    for model_info in MODEL_CONFIGS:
        model_name = model_info["name"]
        measured_epochs = epoch_df[
            (epoch_df["model"] == model_name) & (epoch_df["phase"] == "measured")
        ]
        inference = inference_df[inference_df["model"] == model_name]
        if measured_epochs.empty or inference.empty:
            continue

        row: dict[str, Any] = {
            "model": model_name,
            "display_name": model_info["display_name"],
            "model_type": model_info["model_type"],
            "pooling": model_info["pooling"],
            "fold": measured_epochs["fold"].iloc[0],
            "measured_training_epochs": int(len(measured_epochs)),
            "inference_repeats": int(len(inference)),
        }

        parameter_match = params_df[params_df["model"] == model_name]
        if not parameter_match.empty:
            row.update(parameter_match.iloc[0].to_dict())

        add_stats(row, "train_time_sec", measured_epochs["train_time_sec"])
        add_stats(row, "validation_time_sec", measured_epochs["validation_time_sec"])
        add_stats(row, "training_epoch_total_time_sec", measured_epochs["total_time_sec"])
        add_stats(
            row,
            "training_peak_memory_allocated_gb",
            measured_epochs["peak_memory_allocated_gb"],
        )
        add_stats(
            row,
            "training_peak_memory_reserved_gb",
            measured_epochs["peak_memory_reserved_gb"],
        )
        add_stats(row, "inference_time_sec", inference["total_time_sec"])
        add_stats(row, "inference_samples_per_sec", inference["samples_per_sec"])
        add_stats(row, "inference_latency_ms_per_sample", inference["latency_ms_per_sample"])
        add_stats(
            row,
            "inference_peak_memory_allocated_gb",
            inference["peak_memory_allocated_gb"],
        )
        add_stats(
            row,
            "inference_peak_memory_reserved_gb",
            inference["peak_memory_reserved_gb"],
        )
        add_stats(row, "inference_accuracy", inference["accuracy"])
        rows.append(row)

    return pd.DataFrame(rows)


def build_paper_table(summary_df: pd.DataFrame) -> pd.DataFrame:
    columns = {
        "display_name": "Model",
        "trainable_params": "Trainable parameters",
        "train_time_sec_median": "Training time per epoch (s)",
        "training_peak_memory_allocated_gb_max": "Peak training GPU memory (GB)",
        "inference_time_sec_median": "Inference time (s)",
        "inference_samples_per_sec_median": "Inference throughput (genomes/s)",
        "inference_peak_memory_allocated_gb_max": "Peak inference GPU memory (GB)",
    }
    missing = [column for column in columns if column not in summary_df.columns]
    if missing:
        raise ValueError(
            "The efficiency summary is missing columns required for the paper table: "
            + ", ".join(missing)
        )
    return summary_df[list(columns)].rename(columns=columns).copy()


def write_excel_table(table: pd.DataFrame, destination: Path) -> None:
    """Write a publication-ready Excel table that can be copied into the paper."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Efficiency", index=False)
        worksheet = writer.book["Efficiency"]
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.print_title_rows = "1:1"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_border = Border(bottom=Side(style="medium", color="1F1F1F"))
        body_border = Border(bottom=Side(style="thin", color="D9E2F3"))

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border
        worksheet.row_dimensions[1].height = 42

        for row in worksheet.iter_rows(min_row=2):
            for column_index, cell in enumerate(row, start=1):
                cell.alignment = Alignment(
                    horizontal="left" if column_index == 1 else "right",
                    vertical="center",
                )
                cell.border = body_border

        for cell in worksheet["B"][1:]:
            cell.number_format = "#,##0"
        for column_letter in ("C", "D", "E", "F", "G"):
            for cell in worksheet[column_letter][1:]:
                cell.number_format = "0.00"

        widths = {
            1: 24,
            2: 22,
            3: 28,
            4: 31,
            5: 20,
            6: 34,
            7: 33,
        }
        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _completed_status(run_dir: Path) -> bool:
    status_path = run_dir / "status.json"
    if not status_path.exists():
        return True
    try:
        status = json.loads(status_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    return status.get("status") == "completed"


def reusable_run(run_dir: Path) -> bool:
    return (
        run_dir.is_dir()
        and _completed_status(run_dir)
        and all((run_dir / filename).is_file() for filename in REQUIRED_EXISTING_RESULT_FILES)
    )


def resolve_existing_run(output_root: Path, source_run: Path | None = None) -> Path | None:
    output_root = output_root.expanduser()
    if source_run is not None:
        candidate = source_run.expanduser()
        if not reusable_run(candidate):
            raise FileNotFoundError(
                f"Existing benchmark run is incomplete or unavailable: {candidate}"
            )
        return candidate

    candidates: list[Path] = []
    latest_pointer = output_root / "latest_run.txt"
    if latest_pointer.is_file():
        latest_text = latest_pointer.read_text(encoding="utf-8").strip()
        if latest_text:
            pointed = Path(latest_text).expanduser()
            candidates.append(pointed)
            candidates.append(output_root / pointed.name)

    if output_root.is_dir():
        candidates.extend(sorted(output_root.glob("run_*"), reverse=True))

    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        if reusable_run(candidate):
            return candidate
    return None


def export_paper_table_from_run(run_dir: Path, paper_xlsx_name: str) -> Path:
    summary_path = run_dir / "efficiency_summary.csv"
    summary_df = pd.read_csv(summary_path)
    paper_table = build_paper_table(summary_df)

    csv_path = run_dir / "efficiency_paper_table.csv"
    markdown_path = run_dir / "efficiency_paper_table.md"
    xlsx_path = run_dir / paper_xlsx_name

    paper_table.to_csv(csv_path, index=False)
    write_markdown_table(paper_table, markdown_path)
    write_excel_table(paper_table, xlsx_path)

    print(f"Reused completed benchmark run: {run_dir}", flush=True)
    print(f"Paper table CSV: {csv_path}", flush=True)
    print(f"Paper table Excel: {xlsx_path}", flush=True)
    return xlsx_path

def write_markdown_table(table: pd.DataFrame, destination: Path) -> None:
    try:
        text = table.to_markdown(index=False, floatfmt=".4f")
    except ImportError:
        text = table.to_csv(index=False)
    destination.write_text(text + "\n", encoding="utf-8")


def save_json(data: Any, destination: Path) -> None:
    destination.write_text(json.dumps(data, indent=2, allow_nan=False), encoding="utf-8")


def json_safe_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    clean = frame.replace({np.nan: None})
    return clean.to_dict(orient="records")


def main() -> None:
    args = parse_args()

    existing_run = resolve_existing_run(args.output_root, args.source_run)
    if existing_run is not None and not args.force_benchmark:
        export_paper_table_from_run(existing_run, args.paper_xlsx_name)
        return
    if args.table_only:
        raise FileNotFoundError(
            "No completed efficiency benchmark run was found. "
            "Provide --source-run or verify --output-root/latest_run.txt."
        )

    device = torch.device(config.DEVICE)
    if device.type != "cuda" or not torch.cuda.is_available():
        raise RuntimeError("The efficiency benchmark requires a CUDA GPU.")

    selected_models = [model for model in MODEL_CONFIGS if model["name"] in set(args.models)]
    run_dir = make_run_dir(args.output_root)
    print(f"Benchmark output: {run_dir}", flush=True)
    print(f"CUDA allocation: {os.environ.get('DENFORMER_BENCHMARK_GPU_ALLOCATION')}", flush=True)

    set_seed(args.seed)
    samples, targets = get_dataset(paths.embeddings_dir, "ohe")
    folds = load_and_validate_folds(args.split_file, len(samples), args.fold)
    if len(folds) != 1:
        raise RuntimeError(f"Expected one selected fold, found {len(folds)}")
    fold = folds[0]

    checkpoints = {
        model_info["name"]: checkpoint_for_model(
            model_info=model_info,
            fold=args.fold,
            checkpoint_run_name=args.checkpoint_run_name,
            checkpoint_epochs=args.checkpoint_epochs,
        )
        for model_info in selected_models
    }
    print("Checkpoint preflight completed for all selected models.", flush=True)

    train_dataset = DengueDataset(samples, targets, indices=fold["train_idx"])
    validation_dataset = DengueDataset(samples, targets, indices=fold["val_idx"])
    test_dataset = DengueDataset(samples, targets, indices=fold["test_idx"])
    validation_loader = make_eval_loader(validation_dataset, args.batch_size)
    test_loader = make_eval_loader(test_dataset, args.batch_size)

    metadata = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "protocol": {
            "models": [model["name"] for model in selected_models],
            "split_file": str(args.split_file),
            "fold": args.fold,
            "batch_size": args.batch_size,
            "warmup_epochs": args.warmup_epochs,
            "measured_epochs": args.measured_epochs,
            "inference_warmup_batches": args.inference_warmup_batches,
            "inference_repeats": args.inference_repeats,
            "checkpoint_run_name": args.checkpoint_run_name,
            "checkpoint_epochs": args.checkpoint_epochs,
            "seed": args.seed,
            "training_scope": "full training epoch followed by full validation pass",
            "inference_scope": (
                "complete test loader including host-to-device transfer, forward pass, "
                "softmax, argmax, and accuracy accumulation"
            ),
            "precision": "float32",
        },
        "dataset": {
            "num_samples": int(len(samples)),
            "sample_shape": list(samples.shape),
            "train_size": int(len(train_dataset)),
            "validation_size": int(len(validation_dataset)),
            "test_size": int(len(test_dataset)),
        },
        "device": get_device_info(device),
        "environment": tracked_environment(),
    }
    save_json(metadata, run_dir / "environment.json")
    save_json(
        {"status": "running", "started_at_utc": metadata["created_at_utc"]},
        run_dir / "status.json",
    )

    epoch_measurements: list[EpochMeasurement] = []
    inference_measurements: list[InferenceMeasurement] = []
    parameter_rows: list[dict[str, Any]] = []

    for model_position, model_info in enumerate(selected_models, start=1):
        print("\n" + "=" * 80, flush=True)
        print(
            f"MODEL {model_position}/{len(selected_models)}: {model_info['display_name']}",
            flush=True,
        )
        print("=" * 80, flush=True)

        cleanup_cuda()
        set_seed(args.seed)
        train_loader = make_weighted_train_loader(
            dataset=train_dataset,
            batch_size=args.batch_size,
            seed=args.seed,
        )
        model = build_classifier_model(
            args=model_args(model_info),
            emb_dim=config.EMB_DIM_OHE,
            config=config,
            device=device,
            attn=False,
        )
        total_params, trainable_params = count_parameters(model)
        parameter_rows.append(
            {
                "model": model_info["name"],
                "total_params": total_params,
                "trainable_params": trainable_params,
            }
        )
        loss_fn = torch.nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(model.parameters(), lr=config.LR)

        for warmup_idx in range(1, args.warmup_epochs + 1):
            print(f"  warm-up epoch {warmup_idx}/{args.warmup_epochs}", flush=True)
            measurement = run_training_epoch(
                model_info=model_info,
                model=model,
                train_loader=train_loader,
                validation_loader=validation_loader,
                optimizer=optimizer,
                loss_fn=loss_fn,
                device=device,
                fold=args.fold,
                phase="warmup",
                repeat=warmup_idx,
                log_every=args.log_every,
            )
            epoch_measurements.append(measurement)
            print(
                f"    excluded warm-up total: {measurement.total_time_sec:.3f} s",
                flush=True,
            )

        for measured_idx in range(1, args.measured_epochs + 1):
            print(f"  measured epoch {measured_idx}/{args.measured_epochs}", flush=True)
            measurement = run_training_epoch(
                model_info=model_info,
                model=model,
                train_loader=train_loader,
                validation_loader=validation_loader,
                optimizer=optimizer,
                loss_fn=loss_fn,
                device=device,
                fold=args.fold,
                phase="measured",
                repeat=measured_idx,
                log_every=args.log_every,
            )
            epoch_measurements.append(measurement)
            print(
                "    total={:.3f} s | peak allocated={:.3f} GB".format(
                    measurement.total_time_sec,
                    measurement.peak_memory_allocated_gb or float("nan"),
                ),
                flush=True,
            )

        del optimizer, loss_fn, model, train_loader
        cleanup_cuda()

        checkpoint_path = checkpoints[model_info["name"]]
        print(f"  inference checkpoint: {checkpoint_path}", flush=True)
        inference_model = build_classifier_model(
            args=model_args(model_info),
            emb_dim=config.EMB_DIM_OHE,
            config=config,
            device=device,
            attn=False,
        )
        state_dict = torch.load(checkpoint_path, map_location=device)
        inference_model.load_state_dict(state_dict)
        inference_model.eval()

        completed_warmup = warmup_inference(
            model=inference_model,
            loader=test_loader,
            device=device,
            warmup_batches=args.inference_warmup_batches,
        )
        print(f"  inference warm-up batches completed: {completed_warmup}", flush=True)

        for inference_idx in range(1, args.inference_repeats + 1):
            measurement = run_inference_pass(
                model_info=model_info,
                model=inference_model,
                loader=test_loader,
                device=device,
                fold=args.fold,
                repeat=inference_idx,
                checkpoint_path=checkpoint_path,
            )
            inference_measurements.append(measurement)
            print(
                "  inference {}/{}: {:.3f} s | {:.3f} genomes/s | peak {:.3f} GB".format(
                    inference_idx,
                    args.inference_repeats,
                    measurement.total_time_sec,
                    measurement.samples_per_sec,
                    measurement.peak_memory_allocated_gb or float("nan"),
                ),
                flush=True,
            )

        del state_dict, inference_model
        cleanup_cuda()

        epoch_df_partial = pd.DataFrame(asdict(item) for item in epoch_measurements)
        inference_df_partial = pd.DataFrame(asdict(item) for item in inference_measurements)
        epoch_df_partial.to_csv(run_dir / "training_measurements.csv", index=False)
        inference_df_partial.to_csv(run_dir / "inference_measurements.csv", index=False)

    epoch_df = pd.DataFrame(asdict(item) for item in epoch_measurements)
    inference_df = pd.DataFrame(asdict(item) for item in inference_measurements)
    summary_df = build_summary(epoch_df, inference_df, parameter_rows)
    paper_table = build_paper_table(summary_df)

    epoch_df.to_csv(run_dir / "training_measurements.csv", index=False)
    inference_df.to_csv(run_dir / "inference_measurements.csv", index=False)
    summary_df.to_csv(run_dir / "efficiency_summary.csv", index=False)
    paper_table.to_csv(run_dir / "efficiency_paper_table.csv", index=False)
    write_markdown_table(paper_table, run_dir / "efficiency_paper_table.md")
    write_excel_table(paper_table, run_dir / args.paper_xlsx_name)

    save_json(
        {
            "training_measurements": json_safe_records(epoch_df),
            "inference_measurements": json_safe_records(inference_df),
            "summary": json_safe_records(summary_df),
        },
        run_dir / "efficiency_results.json",
    )

    completed_at = datetime.now(timezone.utc).isoformat()
    save_json(
        {
            "status": "completed",
            "started_at_utc": metadata["created_at_utc"],
            "completed_at_utc": completed_at,
        },
        run_dir / "status.json",
    )
    (args.output_root / "latest_run.txt").write_text(str(run_dir) + "\n", encoding="utf-8")

    print("\nBenchmark completed.", flush=True)
    print(f"Summary: {run_dir / 'efficiency_summary.csv'}", flush=True)
    print(f"Paper table CSV: {run_dir / 'efficiency_paper_table.csv'}", flush=True)
    print(f"Paper table Excel: {run_dir / args.paper_xlsx_name}", flush=True)


if __name__ == "__main__":
    main()

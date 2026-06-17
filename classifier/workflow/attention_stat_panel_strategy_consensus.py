import argparse
import os
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import gridspec
from matplotlib.colors import BoundaryNorm, LinearSegmentedColormap, TwoSlopeNorm
from matplotlib.patches import Patch
from matplotlib import patheffects as pe

import paths
from classifier.workflow import config
from classifier.workflow.utils import (
    build_xai_output_dir,
)
from classifier.utils_attn_box import robust_shared_ylim_from_values
from classifier.workflow.attention_stat_panel import (
    SEROTYPES_ORDER,
    PAIRWISE_ORDER,
    PAIRWISE_SHORT,
    KW_EPS_THRESHOLDS,
    PAIRWISE_RB_THRESHOLDS,
    _safe_neglog10,
    _stars,
    _effect_stars,
)


DEFAULT_STRATEGY_IDS = ['continent', 'temporal', 'cdhit']
DEFAULT_STRATEGY_LABELS = ['Continent', 'Temporal', 'CD-HIT']
DEFAULT_STRATEGY_SPLIT_KEYS = ['continent', 'timebin', 'cdhit']
DEFAULT_STRATEGY_RUN_NAMES = ['continent', 'temporal', 'cdhit']
STRATEGY_COLORS = ['#cc79a7', '#f58518', '#8cd17d']
# Same strategy colors are used for boxes, legend and split-strategy stars.
# The former continent-blue was replaced by a softer pink to improve contrast
# on the p-value heatmaps; CD-HIT uses a lighter green for the same reason.
STRATEGY_STAR_COLORS = STRATEGY_COLORS
HEADER_FONTSIZE = 15
PAIRWISE_HEADER_FONTSIZE = 15
TICK_FONTSIZE = 14
REGION_FONTSIZE = 15
STAR_FONTSIZE = 14
STRATEGY_STAR_FONTSIZE = 12.2
LEGEND_FONTSIZE = HEADER_FONTSIZE
COLORBAR_LABEL_FONTSIZE = 14
COLORBAR_TICK_FONTSIZE = 12


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            'Build an aggregate attention stat panel across split strategies. '
            'The statistical annotations are conservative consensus calls: a region/comparison '
            'is marked significant only if it is significant in every requested split strategy.'
        )
    )
    parser.add_argument('--k', type=int, default=3, help='k-mer length used in the run folders.')
    parser.add_argument('--one-hot', action='store_true', help='Use ohe folder names instead of k-mer folder names.')
    parser.add_argument('--epochs', type=int, default=config.EPOCHS, help='Epoch suffix used in the run folders.')
    parser.add_argument('--pooling', type=str, default='mean', choices=['first', 'mean', 'max'])
    parser.add_argument('--alpha', type=float, default=0.05)
    parser.add_argument('--value-col', type=str, default='attention')

    parser.add_argument('--strategy-ids', nargs=3, default=DEFAULT_STRATEGY_IDS)
    parser.add_argument('--strategy-labels', nargs=3, default=DEFAULT_STRATEGY_LABELS)
    parser.add_argument('--strategy-split-keys', nargs=3, default=DEFAULT_STRATEGY_SPLIT_KEYS)
    parser.add_argument(
        '--strategy-run-names',
        nargs=3,
        default=DEFAULT_STRATEGY_RUN_NAMES,
        help=(
            'Run names used in the attention_box folders. Defaults assume the runs were named '
            'continent, temporal, cdhit. If a folder is not found, the script also tries the '
            'split-key name and the split-file stem.'
        ),
    )
    parser.add_argument(
        '--input-kind',
        default='attention_box',
        help='Aggregate-XAI input kind containing output-box-dataset folders.',
    )
    parser.add_argument(
        '--output-kind',
        default='attention_stat_panel',
        help='Aggregate-XAI output kind under which strategy_consensus will be created.',
    )
    parser.add_argument(
        '--out-name',
        default='continent_temporal_cdhit_consensus',
        help='Final output folder name after the k/pooling/epoch prefix.',
    )
    parser.add_argument(
        '--no-require-kw-for-pairwise',
        action='store_true',
        help='Do not gate pairwise consensus calls by the region-level KW consensus pass.',
    )
    parser.add_argument(
        '--no-require-effect-sign',
        action='store_true',
        help='For the effect-size consensus panel, do not require same signed pairwise effect direction across strategies.',
    )
    return parser.parse_args()


def _p_col(frame):
    if 'p_fdr' in frame.columns and pd.to_numeric(frame['p_fdr'], errors='coerce').notna().any():
        return 'p_fdr'
    if 'p' in frame.columns and pd.to_numeric(frame['p'], errors='coerce').notna().any():
        return 'p'
    raise ValueError('Expected a non-empty p_fdr or p column.')


def _pairwise_effect_col(frame):
    if 'rank_biserial' in frame.columns:
        return 'rank_biserial'
    if 'cliffs_delta' in frame.columns:
        return 'cliffs_delta'
    return None


def _dedupe(values):
    out = []
    seen = set()
    for value in values:
        if value is None:
            continue
        value = str(value).strip()
        if not value or value in seen:
            continue
        out.append(value)
        seen.add(value)
    return out


def _candidate_run_suffixes(run_name, split_key, split_file):
    # Primary candidate: the explicit run name used in the training/XAI commands.
    names = [run_name, split_key]
    # Common alias in the project: the temporal split file is keyed as timebin,
    # while plots are usually named temporal.
    if str(run_name).lower() == 'temporal' or str(split_key).lower() == 'timebin':
        names.extend(['timebin', 'temporal'])
    if str(run_name).lower() == 'cd-hit':
        names.append('cdhit')

    suffixes = [f'_{name}' for name in _dedupe(names)]
    # Fallback used by get_run_suffix when --run_name was not passed.
    suffixes.append(f'_{Path(split_file).stem}')
    return _dedupe(suffixes)


def _find_box_dataset_dir(args, k_type, strategy):
    split_key = strategy['split_key']
    if split_key not in paths.split_files:
        raise KeyError(f'Unknown split key {split_key!r}. Available: {sorted(paths.split_files)}')
    split_file = str(paths.split_files[split_key])

    candidates = []
    for run_suffix in _candidate_run_suffixes(strategy['run_name'], split_key, split_file):
        root = build_xai_output_dir(
            paths.logs_dir,
            args.input_kind,
            'denformer',
            args.pooling,
            k_type,
            run_suffix,
            args.epochs,
        )
        candidates.append(os.path.join(root, 'output-box-dataset'))

    for candidate in candidates:
        if os.path.isdir(candidate):
            return candidate, candidates
    return None, candidates


def _load_strategy_inputs(box_dataset_dir, strategy):
    value_col = str(getattr(strategy, 'value_col', strategy.get('value_col', 'attention')) if not isinstance(strategy, dict) else strategy.get('value_col', 'attention'))
    csv_name = 'attention_by_region_long.csv' if value_col == 'attention' else f'{value_col}_by_region_long.csv'
    df_path = os.path.join(box_dataset_dir, csv_name)
    kw_path = os.path.join(box_dataset_dir, 'kruskal_by_region.csv')
    pair_path = os.path.join(box_dataset_dir, 'pairwise_serotype_stats_by_region.csv')
    missing = [p for p in [df_path, kw_path, pair_path] if not os.path.exists(p)]
    if missing:
        raise FileNotFoundError(f'Missing required {value_col} box outputs:\n' + '\n'.join(missing))

    df = pd.read_csv(df_path)
    kw = pd.read_csv(kw_path)
    pairwise = pd.read_csv(pair_path)

    if 'region' not in kw.columns:
        first = kw.columns[0]
        kw = kw.rename(columns={first: 'region'})
    if 'region' not in pairwise.columns:
        raise ValueError(f'Missing region column in {pair_path}')
    if 'comparison' not in pairwise.columns:
        raise ValueError(f'Missing comparison column in {pair_path}')

    for frame in [df, kw, pairwise]:
        frame.insert(0, 'strategy_label', strategy['label'])
        frame.insert(0, 'strategy', strategy['id'])

    regions_order = list(pd.unique(df['region']))
    return df, kw, pairwise, regions_order


def _merge_region_orders(region_orders):
    out = []
    seen = set()
    for order in region_orders:
        for region in order:
            if region not in seen:
                out.append(region)
                seen.add(region)
    return out


def _star_from_worst_p(p_worst, pass_all):
    if not pass_all or not np.isfinite(p_worst):
        return ''
    return _stars(float(p_worst))


def _effect_magnitude_label(value, thresholds, signed=False):
    if pd.isna(value) or not np.isfinite(float(value)):
        return ''
    mag = abs(float(value)) if signed else float(value)
    if mag >= thresholds[2]:
        return 'large'
    if mag >= thresholds[1]:
        return 'medium'
    if mag >= thresholds[0]:
        return 'small'
    return 'negligible'


def build_kw_consensus(kw_all, regions_order, strategy_ids, alpha=0.05):
    work = kw_all.copy()
    work['_p_use'] = np.nan
    for strategy_id in strategy_ids:
        mask = work['strategy'] == strategy_id
        if mask.any():
            p_col = _p_col(work.loc[mask])
            work.loc[mask, '_p_use'] = pd.to_numeric(work.loc[mask, p_col], errors='coerce')

    p_wide = work.pivot_table(index='region', columns='strategy', values='_p_use', aggfunc='first')
    p_wide = p_wide.reindex(regions_order).reindex(columns=strategy_ids)

    eps_wide = None
    if 'epsilon_squared_R' in work.columns:
        eps_wide = work.pivot_table(index='region', columns='strategy', values='epsilon_squared_R', aggfunc='first')
        eps_wide = eps_wide.reindex(regions_order).reindex(columns=strategy_ids)

    rows = []
    for region in regions_order:
        pvals = p_wide.loc[region].astype(float).values if region in p_wide.index else np.full(len(strategy_ids), np.nan)
        finite = np.isfinite(pvals)
        all_available = bool(finite.all())
        all_significant = bool(all_available and np.all(pvals < alpha))
        p_worst = float(np.max(pvals)) if all_available else np.nan
        row = {
            'region': region,
            'kw_all_available': all_available,
            'kw_all_significant': all_significant,
            'kw_p_worst': p_worst,
            'kw_neg_log10_worst_p': _safe_neglog10([p_worst])[0] if np.isfinite(p_worst) else np.nan,
            'kw_consensus_star': _star_from_worst_p(p_worst, all_significant),
        }
        for strategy_id, pval in zip(strategy_ids, pvals):
            row[f'kw_p_{strategy_id}'] = pval
            row[f'kw_star_{strategy_id}'] = _stars(pval) if np.isfinite(pval) else ''
            row[f'kw_pass_{strategy_id}'] = bool(np.isfinite(pval) and pval < alpha)

        if eps_wide is not None and region in eps_wide.index:
            eps_vals = eps_wide.loc[region].astype(float).values
            row['kw_epsilon_min'] = float(np.nanmin(eps_vals)) if np.isfinite(eps_vals).any() else np.nan
            row['kw_epsilon_median'] = float(np.nanmedian(eps_vals)) if np.isfinite(eps_vals).any() else np.nan
            row['kw_epsilon_consensus_stars'] = _effect_stars(row['kw_epsilon_min'], KW_EPS_THRESHOLDS, signed=False)
            for strategy_id, value in zip(strategy_ids, eps_vals):
                row[f'kw_epsilon_{strategy_id}'] = value
        rows.append(row)
    return pd.DataFrame(rows)


def _sign_concordant(values):
    arr = np.asarray(values, dtype=float)
    if not np.isfinite(arr).all():
        return False, np.nan
    signs = np.sign(arr)
    if np.any(signs == 0):
        return False, np.nan
    first = signs[0]
    return bool(np.all(signs == first)), float(first)


def build_pairwise_consensus(
    pairwise_all,
    kw_consensus,
    regions_order,
    strategy_ids,
    alpha=0.05,
    require_kw=True,
    require_effect_sign=True,
):
    work = pairwise_all.copy()
    work['_p_use'] = np.nan
    for strategy_id in strategy_ids:
        mask = work['strategy'] == strategy_id
        if mask.any():
            p_col = _p_col(work.loc[mask])
            work.loc[mask, '_p_use'] = pd.to_numeric(work.loc[mask, p_col], errors='coerce')

    effect_col = _pairwise_effect_col(work)
    if effect_col is not None:
        work['_effect_use'] = pd.to_numeric(work[effect_col], errors='coerce')
    else:
        work['_effect_use'] = np.nan

    p_wide = work.pivot_table(index=['region', 'comparison'], columns='strategy', values='_p_use', aggfunc='first')
    p_wide = p_wide.reindex(columns=strategy_ids)
    effect_wide = work.pivot_table(index=['region', 'comparison'], columns='strategy', values='_effect_use', aggfunc='first')
    effect_wide = effect_wide.reindex(columns=strategy_ids)

    kw_pass = dict(zip(kw_consensus['region'], kw_consensus['kw_all_significant']))
    rows = []
    for region in regions_order:
        for comparison in PAIRWISE_ORDER:
            key = (region, comparison)
            if key in p_wide.index:
                pvals = p_wide.loc[key].astype(float).values
            else:
                pvals = np.full(len(strategy_ids), np.nan)
            finite = np.isfinite(pvals)
            all_available = bool(finite.all())
            pair_all_significant = bool(all_available and np.all(pvals < alpha))
            p_worst = float(np.max(pvals)) if all_available else np.nan

            if key in effect_wide.index:
                effects = effect_wide.loc[key].astype(float).values
            else:
                effects = np.full(len(strategy_ids), np.nan)

            if effect_col is not None:
                sign_ok, common_sign = _sign_concordant(effects)
            else:
                sign_ok, common_sign = True, np.nan

            kw_ok = bool(kw_pass.get(region, False)) if require_kw else True
            effect_sign_ok = bool(sign_ok) if require_effect_sign else True

            # Keep p-value consensus and effect-direction consensus separate.
            # The p-value panels should answer exactly this question:
            #   among KW-consensus regions, are all requested strategies significant,
            #   and what is the worst/max p-value star?
            # Effect sign concordance is a different diagnostic and is used only for
            # the effect-size panel and its columns.
            pvalue_consensus_pass = bool(pair_all_significant and kw_ok)
            effect_consensus_pass = bool(pvalue_consensus_pass and effect_sign_ok)

            effect_min_abs = float(np.nanmin(np.abs(effects))) if np.isfinite(effects).any() else np.nan
            effect_median = float(np.nanmedian(effects)) if np.isfinite(effects).any() else np.nan
            effect_worst_signed = float(common_sign * effect_min_abs) if sign_ok and np.isfinite(effect_min_abs) else np.nan

            row = {
                'region': region,
                'comparison': comparison,
                'pair_all_available': all_available,
                'pair_all_significant': pair_all_significant,
                'kw_all_significant': bool(kw_pass.get(region, False)),
                'sign_concordant': bool(sign_ok),
                'pair_pvalue_consensus_pass': pvalue_consensus_pass,
                'pair_pvalue_consensus_star': _star_from_worst_p(p_worst, pvalue_consensus_pass),
                'pair_effect_consensus_pass': effect_consensus_pass,

                # Backward-compatible names used by the p-value plot/export.
                'pair_consensus_pass': pvalue_consensus_pass,
                'pair_p_worst': p_worst,
                'pair_neg_log10_worst_p': _safe_neglog10([p_worst])[0] if np.isfinite(p_worst) else np.nan,
                'pair_consensus_star': _star_from_worst_p(p_worst, pvalue_consensus_pass),
                'pair_effect_col': effect_col if effect_col is not None else '',
                'pair_effect_min_abs': effect_min_abs,
                'pair_effect_median': effect_median,
                'pair_effect_worst_signed': effect_worst_signed,
                'pair_effect_consensus_stars': _effect_stars(effect_min_abs, PAIRWISE_RB_THRESHOLDS, signed=False)
                if effect_consensus_pass and np.isfinite(effect_min_abs) else '',
                'pair_effect_magnitude': _effect_magnitude_label(effect_min_abs, PAIRWISE_RB_THRESHOLDS, signed=False)
                if effect_consensus_pass and np.isfinite(effect_min_abs) else '',
            }
            for strategy_id, pval, eff in zip(strategy_ids, pvals, effects):
                row[f'pair_p_{strategy_id}'] = pval
                row[f'pair_star_{strategy_id}'] = _stars(pval) if np.isfinite(pval) else ''
                row[f'pair_pass_{strategy_id}'] = bool(np.isfinite(pval) and pval < alpha)
                row[f'pair_effect_{strategy_id}'] = eff
            rows.append(row)
    return pd.DataFrame(rows)


def _draw_strategy_box_panel(fig, outer_spec, df, regions_order, strategy_ids, strategy_labels, strategy_colors, value_col='attention'):
    nrows = len(regions_order)
    sub = gridspec.GridSpecFromSubplotSpec(nrows, 1, subplot_spec=outer_spec, hspace=0.0)
    axes = []
    n_strategies = len(strategy_ids)
    centers = np.arange(1, len(SEROTYPES_ORDER) + 1, dtype=float)
    offsets = np.linspace(-0.33, 0.33, n_strategies) if n_strategies > 1 else np.array([0.0])
    width = min(0.245, 0.78 / max(n_strategies, 1))
    # Use row-specific robust y-limits. A global 1st-99th percentile range
    # can still flatten low-dynamic regions when a few regions dominate the
    # upper tail, especially for sparse attribution signals such as GxI.
    shared_ylim = None

    for i, region in enumerate(regions_order):
        ax = fig.add_subplot(sub[i, 0])
        axes.append(ax)

        data = []
        positions = []
        colors = []
        for center, serotype in zip(centers, SEROTYPES_ORDER):
            for strategy_id, offset, color in zip(strategy_ids, offsets, strategy_colors):
                vals = df[
                    (df['region'] == region)
                    & (df['serotype'] == serotype)
                    & (df['strategy'] == strategy_id)
                ][value_col].dropna().values
                data.append(vals)
                positions.append(center + offset)
                colors.append(color)

        bp = ax.boxplot(
            data,
            positions=positions,
            widths=width,
            patch_artist=True,
            showfliers=False,
            whis=(5, 95),
            medianprops=dict(color='brown', linewidth=1.2),
            boxprops=dict(linewidth=0.9),
            whiskerprops=dict(linewidth=0.9),
            capprops=dict(linewidth=0.9),
        )
        for box, color in zip(bp['boxes'], colors):
            box.set_facecolor(color)
            box.set_alpha(0.88)

        ax.set_ylabel(region, fontsize=REGION_FONTSIZE, rotation=0, labelpad=30, va='center')
        # Use exact half-cell limits so the four DENV blocks have identical
        # horizontal width and the three strategy boxes stay visually centered
        # within each serotype block.
        ax.set_xlim(0.5, len(SEROTYPES_ORDER) + 0.5)
        ax.set_yticks([])
        ax.grid(True, axis='y', linestyle='--', linewidth=0.5, alpha=0.35)
        for sep in np.arange(1.5, len(SEROTYPES_ORDER), 1.0):
            ax.axvline(sep, color='lightgray', linewidth=0.6, alpha=0.45)

        if shared_ylim is not None:
            ax.set_ylim(*shared_ylim)
        else:
            all_vals = np.concatenate([v for v in data if len(v) > 0]) if any(len(v) for v in data) else np.array([])
            if len(all_vals) > 0:
                fallback_ylim = robust_shared_ylim_from_values(
                    all_vals,
                    percentiles=(1, 99),
                    pad_frac=0.15,
                    nonnegative=False,
                )
                if fallback_ylim is not None:
                    ax.set_ylim(*fallback_ylim)

        if i == 0:
            ax.set_xticks(centers)
            ax.set_xticklabels(SEROTYPES_ORDER, fontsize=HEADER_FONTSIZE)
            ax.xaxis.tick_top()
            ax.tick_params(axis='x', pad=3)
        else:
            ax.set_xticks([])
    return axes



def _truncate_colormap(cmap_name, minval=0.00, maxval=0.82, n=256):
    base = plt.get_cmap(cmap_name)
    return LinearSegmentedColormap.from_list(
        f'{cmap_name}_trunc_{minval:.2f}_{maxval:.2f}',
        base(np.linspace(minval, maxval, n)),
    )


PVALUE_CMAP = _truncate_colormap('Blues', 0.13, 0.66)


def _text_color_for_heat_value(value, vmax=300.0, threshold=0.52):
    if not np.isfinite(value) or vmax <= 0:
        return 'black'
    return 'white' if (float(value) / float(vmax)) >= threshold else 'black'


def _draw_strategy_star_triplet(ax, x, y, row, prefix, strategy_ids, strategy_colors, fontsize=STRATEGY_STAR_FONTSIZE):
    """Draw one colored star string per strategy inside a heatmap cell."""
    offsets = np.linspace(-0.35, 0.35, len(strategy_ids)) if len(strategy_ids) > 1 else [0.0]
    for dx, strategy_id, color in zip(offsets, strategy_ids, strategy_colors):
        txt = row.get(f'{prefix}_star_{strategy_id}', '')
        if pd.isna(txt) or str(txt) == '':
            continue
        ax.text(
            x + dx,
            y,
            str(txt),
            ha='center',
            va='center',
            fontsize=fontsize,
            color=color,
            fontweight='bold',
            clip_on=True,
            path_effects=[pe.withStroke(linewidth=0.45, foreground='black')],
        )


def _draw_kw_consensus_pvalue_panel(
    fig,
    spec,
    kw_consensus,
    regions_order,
    strategy_ids=None,
    strategy_colors=None,
    star_mode='consensus',
):
    ax = fig.add_subplot(spec)
    work = kw_consensus.set_index('region').reindex(regions_order)
    pvals = work['kw_p_worst'].astype(float).values.reshape(-1, 1)
    pass_mask = work['kw_all_significant'].astype(bool).values.reshape(-1, 1)
    values = _safe_neglog10(pvals)
    values[~pass_mask] = 0.0
    im = ax.imshow(values, aspect='auto', cmap=PVALUE_CMAP, vmin=0, vmax=300)
    ax.set_title('Kruskal-Wallis', fontsize=HEADER_FONTSIZE, pad=8)
    ax.set_xticks([0])
    ax.set_xticklabels([''], fontsize=TICK_FONTSIZE)
    ax.xaxis.tick_top()
    ax.set_yticks(np.arange(len(regions_order)))
    ax.set_yticklabels([])
    ax.tick_params(length=0)
    for r, region in enumerate(regions_order):
        if star_mode == 'strategy':
            _draw_strategy_star_triplet(
                ax,
                0,
                r,
                work.loc[region],
                'kw',
                strategy_ids or [],
                strategy_colors or STRATEGY_COLORS,
                fontsize=STRATEGY_STAR_FONTSIZE,
            )
        else:
            txt = str(work.loc[region, 'kw_consensus_star'])
            ax.text(
                0,
                r,
                txt,
                ha='center',
                va='center',
                fontsize=STAR_FONTSIZE,
                color=_text_color_for_heat_value(values[r, 0]),
                fontweight='bold' if txt else 'normal',
            )
    ax.set_xticks(np.arange(-.5, 1, 1), minor=True)
    ax.set_yticks(np.arange(-.5, len(regions_order), 1), minor=True)
    ax.grid(which='minor', color='lightgray', linestyle='-', linewidth=0.4, alpha=0.45)
    return ax, im

def _pivot_pairwise_consensus(pairwise_consensus, value_col, regions_order):
    matrix = pairwise_consensus.pivot(index='region', columns='comparison', values=value_col)
    present = [c for c in PAIRWISE_ORDER if c in matrix.columns]
    if not present:
        present = list(matrix.columns)
    return matrix.reindex(regions_order)[present]


def _draw_pairwise_consensus_pvalue_panel(
    fig,
    spec,
    pairwise_consensus,
    regions_order,
    strategy_ids=None,
    strategy_colors=None,
    star_mode='consensus',
):
    ax = fig.add_subplot(spec)
    pmat = _pivot_pairwise_consensus(pairwise_consensus, 'pair_p_worst', regions_order)
    pass_mat = _pivot_pairwise_consensus(pairwise_consensus, 'pair_consensus_pass', regions_order).astype(bool)
    star_mat = _pivot_pairwise_consensus(pairwise_consensus, 'pair_consensus_star', regions_order)
    values = _safe_neglog10(pmat.values)
    values[~pass_mat.values] = 0.0
    im = ax.imshow(values, aspect='auto', cmap=PVALUE_CMAP, vmin=0, vmax=300)
    ax.set_title('', fontsize=HEADER_FONTSIZE, pad=8)
    labels = [PAIRWISE_SHORT[PAIRWISE_ORDER.index(c)] if c in PAIRWISE_ORDER else str(c) for c in pmat.columns]
    ax.set_xticks(list(range(len(pmat.columns))))
    ax.set_xticklabels(labels, fontsize=PAIRWISE_HEADER_FONTSIZE)
    ax.xaxis.tick_top()
    ax.set_yticks(np.arange(len(regions_order)))
    ax.set_yticklabels([])
    ax.tick_params(length=0)
    pair_lookup = pairwise_consensus.set_index(['region', 'comparison'])
    for r, region in enumerate(regions_order):
        for c, comparison in enumerate(pmat.columns):
            if star_mode == 'strategy':
                if (region, comparison) in pair_lookup.index:
                    row = pair_lookup.loc[(region, comparison)]
                    # Pairwise post-hoc annotations are only interpreted for regions
                    # that pass the KW consensus rule. This keeps the strategy-star
                    # view aligned with the worst-p consensus view.
                    if bool(row.get('kw_all_significant', True)):
                        _draw_strategy_star_triplet(
                            ax,
                            c,
                            r,
                            row,
                            'pair',
                            strategy_ids or [],
                            strategy_colors or STRATEGY_STAR_COLORS,
                            fontsize=STRATEGY_STAR_FONTSIZE,
                        )
            else:
                txt = str(star_mat.iloc[r, c]) if pd.notna(star_mat.iloc[r, c]) else ''
                ax.text(
                    c,
                    r,
                    txt,
                    ha='center',
                    va='center',
                    fontsize=STAR_FONTSIZE,
                    color=_text_color_for_heat_value(values[r, c]),
                    fontweight='bold' if txt else 'normal',
                )
    ax.set_xticks(np.arange(-.5, values.shape[1], 1), minor=True)
    ax.set_yticks(np.arange(-.5, len(regions_order), 1), minor=True)
    ax.grid(which='minor', color='lightgray', linestyle='-', linewidth=0.4, alpha=0.45)
    return ax, im

def _draw_kw_consensus_effect_panel(fig, spec, kw_consensus, regions_order):
    ax = fig.add_subplot(spec)
    work = kw_consensus.set_index('region').reindex(regions_order)
    if 'kw_epsilon_min' not in work.columns:
        raise ValueError('Missing kw_epsilon_min in KW consensus table.')
    values = work['kw_epsilon_min'].astype(float).values.reshape(-1, 1)
    pass_mask = work['kw_all_significant'].astype(bool).values.reshape(-1, 1)
    values_plot = np.where(pass_mask, values, 0.0)
    boundaries = [0.0, KW_EPS_THRESHOLDS[0], KW_EPS_THRESHOLDS[1], KW_EPS_THRESHOLDS[2], 1.0]
    norm = BoundaryNorm(boundaries, ncolors=256, clip=True)
    im = ax.imshow(values_plot, aspect='auto', cmap='Blues', norm=norm)
    ax.set_title('Kruskal-Wallis', fontsize=HEADER_FONTSIZE, pad=8)
    ax.set_xticks([0])
    ax.set_xticklabels([''], fontsize=TICK_FONTSIZE)
    ax.xaxis.tick_top()
    ax.set_yticks(np.arange(len(regions_order)))
    ax.set_yticklabels([])
    ax.tick_params(length=0)
    for r, region in enumerate(regions_order):
        txt = work.loc[region, 'kw_epsilon_consensus_stars'] if bool(work.loc[region, 'kw_all_significant']) else ''
        ax.text(0, r, txt, ha='center', va='center', fontsize=STAR_FONTSIZE, color='black')
    ax.set_xticks(np.arange(-.5, 1, 1), minor=True)
    ax.set_yticks(np.arange(-.5, len(regions_order), 1), minor=True)
    ax.grid(which='minor', color='lightgray', linestyle='-', linewidth=0.4, alpha=0.45)
    return ax, im


def _draw_pairwise_consensus_effect_panel(fig, spec, pairwise_consensus, regions_order):
    ax = fig.add_subplot(spec)
    mat = _pivot_pairwise_consensus(pairwise_consensus, 'pair_effect_worst_signed', regions_order)
    pass_col = 'pair_effect_consensus_pass' if 'pair_effect_consensus_pass' in pairwise_consensus.columns else 'pair_consensus_pass'
    pass_mat = _pivot_pairwise_consensus(pairwise_consensus, pass_col, regions_order).astype(bool)
    star_mat = _pivot_pairwise_consensus(pairwise_consensus, 'pair_effect_consensus_stars', regions_order)
    values = mat.values.astype(float)
    values_plot = np.where(pass_mat.values, values, 0.0)
    norm = TwoSlopeNorm(vmin=-1.0, vcenter=0.0, vmax=1.0)
    im = ax.imshow(values_plot, aspect='auto', cmap='coolwarm', norm=norm)
    ax.set_title('', fontsize=HEADER_FONTSIZE, pad=8)
    labels = [PAIRWISE_SHORT[PAIRWISE_ORDER.index(c)] if c in PAIRWISE_ORDER else str(c) for c in mat.columns]
    ax.set_xticks(list(range(len(mat.columns))))
    ax.set_xticklabels(labels, fontsize=PAIRWISE_HEADER_FONTSIZE)
    ax.xaxis.tick_top()
    ax.set_yticks(np.arange(len(regions_order)))
    ax.set_yticklabels([])
    ax.tick_params(length=0)
    for r in range(values.shape[0]):
        for c in range(values.shape[1]):
            ax.text(c, r, str(star_mat.iloc[r, c]) if pd.notna(star_mat.iloc[r, c]) else '', ha='center', va='center', fontsize=STAR_FONTSIZE, color='black')
    ax.set_xticks(np.arange(-.5, values.shape[1], 1), minor=True)
    ax.set_yticks(np.arange(-.5, len(regions_order), 1), minor=True)
    ax.grid(which='minor', color='lightgray', linestyle='-', linewidth=0.4, alpha=0.45)
    return ax, im


def plot_consensus_pvalue_panel(
    df,
    kw_consensus,
    pairwise_consensus,
    regions_order,
    strategy_ids,
    strategy_labels,
    title=None,
    star_mode='consensus',
    value_col='attention',
):
    fig = plt.figure(figsize=(20.8, max(8.2, 0.74 * len(regions_order))))
    # Make the KW column the same physical width as each pairwise column.
    # Pairwise has six columns, so KW width ~= pairwise_width / 6.
    gs = gridspec.GridSpec(1, 4, width_ratios=[7.15, 1.51, 9.05, 0.28], wspace=0.115)
    _draw_strategy_box_panel(fig, gs[0], df, regions_order, strategy_ids, strategy_labels, STRATEGY_COLORS, value_col=value_col)
    _, _ = _draw_kw_consensus_pvalue_panel(
        fig,
        gs[1],
        kw_consensus,
        regions_order,
        strategy_ids=strategy_ids,
        strategy_colors=STRATEGY_STAR_COLORS if star_mode == 'strategy' else STRATEGY_COLORS,
        star_mode=star_mode,
    )
    _, im_pair = _draw_pairwise_consensus_pvalue_panel(
        fig,
        gs[2],
        pairwise_consensus,
        regions_order,
        strategy_ids=strategy_ids,
        strategy_colors=STRATEGY_STAR_COLORS if star_mode == 'strategy' else STRATEGY_COLORS,
        star_mode=star_mode,
    )
    cax = fig.add_subplot(gs[3])
    cb = fig.colorbar(im_pair, cax=cax)
    cb.set_label(r'$-\log_{10}(\max p_{FDR})$ across strategies', fontsize=COLORBAR_LABEL_FONTSIZE, labelpad=10)
    cb.ax.tick_params(labelsize=COLORBAR_TICK_FONTSIZE)

    handles = [Patch(facecolor=color, edgecolor='black', label=label, alpha=0.88) for color, label in zip(STRATEGY_COLORS, strategy_labels)]
    fig.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, 0.985), ncol=len(handles), frameon=False, fontsize=LEGEND_FONTSIZE)
    fig.subplots_adjust(left=0.050, right=0.978, top=0.885, bottom=0.055)
    return fig


def plot_consensus_effect_panel(df, kw_consensus, pairwise_consensus, regions_order, strategy_ids, strategy_labels, title=None, value_col='attention'):
    fig = plt.figure(figsize=(20.8, max(8.2, 0.74 * len(regions_order))))
    gs = gridspec.GridSpec(1, 3, width_ratios=[7.15, 1.51, 9.05], wspace=0.125)
    _draw_strategy_box_panel(fig, gs[0], df, regions_order, strategy_ids, strategy_labels, STRATEGY_COLORS, value_col=value_col)
    _, im_kw = _draw_kw_consensus_effect_panel(fig, gs[1], kw_consensus, regions_order)
    _, im_pair = _draw_pairwise_consensus_effect_panel(fig, gs[2], pairwise_consensus, regions_order)

    handles = [Patch(facecolor=color, edgecolor='black', label=label, alpha=0.88) for color, label in zip(STRATEGY_COLORS, strategy_labels)]
    fig.legend(handles=handles, loc='upper center', bbox_to_anchor=(0.5, 0.985), ncol=len(handles), frameon=False, fontsize=LEGEND_FONTSIZE)

    fig.subplots_adjust(left=0.050, right=0.86, top=0.885, bottom=0.09)

    cax_kw = fig.add_axes([0.875, 0.58, 0.020, 0.30])
    cb_kw = fig.colorbar(
        im_kw,
        cax=cax_kw,
        boundaries=[0.0, KW_EPS_THRESHOLDS[0], KW_EPS_THRESHOLDS[1], KW_EPS_THRESHOLDS[2], 1.0],
        ticks=[
            KW_EPS_THRESHOLDS[0] / 2.0,
            (KW_EPS_THRESHOLDS[0] + KW_EPS_THRESHOLDS[1]) / 2.0,
            (KW_EPS_THRESHOLDS[1] + KW_EPS_THRESHOLDS[2]) / 2.0,
            (KW_EPS_THRESHOLDS[2] + 1.0) / 2.0,
        ],
    )
    cb_kw.ax.set_yticklabels(['<0.01', '* >=0.01', '** >=0.06', '*** >=0.14'])
    cb_kw.ax.tick_params(labelsize=COLORBAR_TICK_FONTSIZE)
    cb_kw.set_label('min KW epsilon-squared', fontsize=COLORBAR_LABEL_FONTSIZE, labelpad=8)

    cax_pair = fig.add_axes([0.940, 0.18, 0.020, 0.64])
    cb_pair = fig.colorbar(im_pair, cax=cax_pair)
    cb_pair.set_ticks([-1.0, -0.43, -0.28, -0.11, 0.0, 0.11, 0.28, 0.43, 1.0])
    cb_pair.ax.tick_params(labelsize=COLORBAR_TICK_FONTSIZE)
    cb_pair.set_label('signed min |rank-biserial|', fontsize=COLORBAR_LABEL_FONTSIZE, labelpad=8)

    fig.text(
        0.45,
        0.018,
        'Consensus cells require significance in all strategies. Effect-size stars use the weakest magnitude across strategies. '
        'KW thresholds: 0.01/0.06/0.14; pairwise thresholds: |rank-biserial| 0.11/0.28/0.43.',
        ha='center',
        va='bottom',
        fontsize=10,
    )
    return fig




def _wide_first_value(frame, index_cols, value_col, strategy_ids, prefix=None):
    """Return a strategy-wide table for one value column.

    The input frame is expected to contain one row per strategy and index key.
    Missing columns are represented as NaN, so the export remains robust if an
    older box output does not contain every effect-size field.
    """
    prefix = value_col if prefix is None else prefix
    base = frame[index_cols].drop_duplicates().copy()
    if value_col not in frame.columns:
        for strategy_id in strategy_ids:
            base[f'{prefix}_{strategy_id}'] = np.nan
        return base
    wide = frame.pivot_table(
        index=index_cols,
        columns='strategy',
        values=value_col,
        aggfunc='first',
    ).reindex(columns=strategy_ids)
    wide.columns = [f'{prefix}_{strategy_id}' for strategy_id in wide.columns]
    return wide.reset_index()


def _merge_wide_tables(tables, index_cols):
    out = None
    for table in tables:
        if out is None:
            out = table.copy()
        else:
            out = out.merge(table, on=index_cols, how='outer')
    return out if out is not None else pd.DataFrame(columns=index_cols)


def build_kw_pvalue_export(kw_all, kw_consensus, regions_order, strategy_ids):
    work = kw_all.copy()
    work['_p_use'] = np.nan
    work['_p_source'] = ''
    for strategy_id in strategy_ids:
        mask = work['strategy'] == strategy_id
        if mask.any():
            p_col = _p_col(work.loc[mask])
            work.loc[mask, '_p_use'] = pd.to_numeric(work.loc[mask, p_col], errors='coerce')
            work.loc[mask, '_p_source'] = p_col
    pvals = _wide_first_value(work, ['region'], '_p_use', strategy_ids, prefix='p')
    stars = _wide_first_value(work, ['region'], '_p_use', strategy_ids, prefix='star')
    for strategy_id in strategy_ids:
        col = f'star_{strategy_id}'
        stars[col] = stars[col].map(lambda v: _stars(v) if pd.notna(v) and np.isfinite(float(v)) else '')
    sources = _wide_first_value(work, ['region'], '_p_source', strategy_ids, prefix='p_source')
    out = _merge_wide_tables([pvals, stars, sources], ['region'])
    keep = ['region', 'kw_p_worst', 'kw_neg_log10_worst_p', 'kw_all_significant', 'kw_consensus_star']
    keep = [c for c in keep if c in kw_consensus.columns]
    out = out.merge(kw_consensus[keep], on='region', how='left')
    return out.set_index('region').reindex(regions_order).reset_index()


def build_pairwise_pvalue_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids):
    work = pairwise_all.copy()
    work['_p_use'] = np.nan
    work['_p_source'] = ''
    for strategy_id in strategy_ids:
        mask = work['strategy'] == strategy_id
        if mask.any():
            p_col = _p_col(work.loc[mask])
            work.loc[mask, '_p_use'] = pd.to_numeric(work.loc[mask, p_col], errors='coerce')
            work.loc[mask, '_p_source'] = p_col
    idx = ['region', 'comparison']
    pvals = _wide_first_value(work, idx, '_p_use', strategy_ids, prefix='p')
    stars = _wide_first_value(work, idx, '_p_use', strategy_ids, prefix='star')
    for strategy_id in strategy_ids:
        col = f'star_{strategy_id}'
        stars[col] = stars[col].map(lambda v: _stars(v) if pd.notna(v) and np.isfinite(float(v)) else '')
    sources = _wide_first_value(work, idx, '_p_source', strategy_ids, prefix='p_source')
    out = _merge_wide_tables([pvals, stars, sources], idx)
    keep = [
        'region', 'comparison', 'pair_p_worst', 'pair_neg_log10_worst_p',
        'kw_all_significant', 'pair_all_significant', 'pair_pvalue_consensus_pass',
        'pair_pvalue_consensus_star', 'pair_consensus_star',
    ]
    keep = [c for c in keep if c in pairwise_consensus.columns]
    out = out.merge(pairwise_consensus[keep], on=idx, how='left')
    out['region'] = pd.Categorical(out['region'], categories=regions_order, ordered=True)
    out['comparison'] = pd.Categorical(out['comparison'], categories=PAIRWISE_ORDER, ordered=True)
    return out.sort_values(['region', 'comparison']).reset_index(drop=True)


def build_kw_effect_export(kw_all, kw_consensus, regions_order, strategy_ids):
    idx = ['region']
    tables = []
    for col in ['eta_squared_H', 'epsilon_squared_R', 'effect_size_magnitude']:
        tables.append(_wide_first_value(kw_all, idx, col, strategy_ids, prefix=col))
    out = _merge_wide_tables(tables, idx)
    keep = [
        'region', 'kw_epsilon_min', 'kw_epsilon_median',
        'kw_epsilon_consensus_stars', 'kw_all_significant',
    ]
    keep = [c for c in keep if c in kw_consensus.columns]
    out = out.merge(kw_consensus[keep], on='region', how='left')
    return out.set_index('region').reindex(regions_order).reset_index()


def build_pairwise_effect_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids):
    idx = ['region', 'comparison']
    tables = []
    for col in [
        'VDA', 'VDA_folded', 'rank_biserial', 'cliffs_delta',
        'effect_size_magnitude', 'dominant_serotype',
    ]:
        tables.append(_wide_first_value(pairwise_all, idx, col, strategy_ids, prefix=col))
    out = _merge_wide_tables(tables, idx)
    keep = [
        'region', 'comparison', 'pair_effect_col', 'pair_effect_min_abs',
        'pair_effect_median', 'pair_effect_worst_signed',
        'pair_effect_consensus_stars', 'pair_effect_magnitude',
        'sign_concordant', 'pair_effect_consensus_pass', 'pair_pvalue_consensus_pass',
    ]
    keep = [c for c in keep if c in pairwise_consensus.columns]
    out = out.merge(pairwise_consensus[keep], on=idx, how='left')
    out['region'] = pd.Categorical(out['region'], categories=regions_order, ordered=True)
    out['comparison'] = pd.Categorical(out['comparison'], categories=PAIRWISE_ORDER, ordered=True)
    return out.sort_values(['region', 'comparison']).reset_index(drop=True)


def _autosize_excel_columns(writer, sheet_name, table):
    worksheet = writer.sheets[sheet_name]
    for idx, column in enumerate(table.columns, start=1):
        max_len = max([len(str(column))] + [len(str(v)) for v in table[column].head(200).values])
        width = min(max(max_len + 2, 10), 34)
        col_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[col_letter].width = width
    worksheet.freeze_panes = 'A2'


def _validate_pvalue_star_consistency(kw_consensus, pairwise_consensus, strategy_ids, require_kw=True):
    """Fail fast if the two p-value panels can disagree logically.

    In the standard p-value panel, stars are based on the worst/max p-value.
    In the split-strategy-stars panel, the colored stars are the individual
    strategy p-values. Therefore, when all individual strategy p-values are
    significant in an eligible cell, the standard panel must contain the star
    computed from the worst p-value.
    """
    problems = []

    for _, row in kw_consensus.iterrows():
        passes = [bool(row.get(f'kw_pass_{sid}', False)) for sid in strategy_ids]
        expected_pass = all(passes)
        expected_star = _star_from_worst_p(row.get('kw_p_worst', np.nan), expected_pass)
        actual_star = '' if pd.isna(row.get('kw_consensus_star', '')) else str(row.get('kw_consensus_star', ''))
        if actual_star != expected_star:
            problems.append(
                f"KW {row.get('region')}: expected {expected_star!r} from max p, found {actual_star!r}"
            )

    for _, row in pairwise_consensus.iterrows():
        passes = [bool(row.get(f'pair_pass_{sid}', False)) for sid in strategy_ids]
        kw_ok = bool(row.get('kw_all_significant', False)) if require_kw else True
        expected_pass = bool(all(passes) and kw_ok)
        expected_star = _star_from_worst_p(row.get('pair_p_worst', np.nan), expected_pass)
        actual_star = '' if pd.isna(row.get('pair_consensus_star', '')) else str(row.get('pair_consensus_star', ''))
        if actual_star != expected_star:
            problems.append(
                f"Pairwise {row.get('region')} {row.get('comparison')}: "
                f"expected {expected_star!r} from max p, found {actual_star!r}"
            )

    if problems:
        preview = '\n'.join(problems[:20])
        extra = '' if len(problems) <= 20 else f"\n... and {len(problems) - 20} more"
        raise AssertionError(
            'Inconsistent p-value consensus/star annotations.\n' + preview + extra
        )



def _scientific_bool(value):
    if pd.isna(value):
        return ''
    return 'yes' if bool(value) else 'no'


def _strategy_display_names(strategy_ids, strategy_labels=None):
    labels = list(strategy_labels) if strategy_labels is not None else list(strategy_ids)
    if len(labels) != len(strategy_ids):
        labels = list(strategy_ids)
    return {sid: label for sid, label in zip(strategy_ids, labels)}


def _make_readme_table(value_col, strategy_ids, strategy_labels, alpha):
    signal_name = 'Transformer attention' if value_col == 'attention' else 'Gradient x Input (GxI)'
    strategies = ', '.join(_strategy_display_names(strategy_ids, strategy_labels).values())
    rows = [
        ('Signal', signal_name),
        ('Split strategies', strategies),
        ('Kruskal-Wallis tests', 'Performed independently for each genomic region across the four DENV serotypes.'),
        ('Pairwise tests', 'Two-sided Mann-Whitney U tests for all six DENV serotype pairs within each genomic region.'),
        ('Multiple testing correction', 'Benjamini-Hochberg false discovery rate correction; 13 tests for Kruskal-Wallis and 78 tests for pairwise comparisons.'),
        ('Consensus p-value rule', f'A result is marked as consensus-significant only when all requested split strategies have corrected p-values below alpha={alpha:g}.'),
        ('Pairwise gating rule', 'Pairwise consensus is evaluated for regions passing the Kruskal-Wallis consensus rule.'),
        ('Kruskal-Wallis effect sizes', 'Eta-squared H and epsilon-squared R are reported; epsilon-squared is used as the primary regional effect-size summary.'),
        ('Pairwise effect sizes', 'Vargha-Delaney A, folded Vargha-Delaney A, rank-biserial correlation, and Cliff\'s delta are reported.'),
        ('Primary signed pairwise effect size', 'Rank-biserial correlation; positive values indicate higher signal values for the first serotype in the comparison, negative values for the second.'),
        ('Interpretation', 'Statistics quantify differences in model-derived XAI signals and should not be interpreted as causal biological evidence.'),
    ]
    return pd.DataFrame(rows, columns=['Item', 'Description'])


def build_supplementary_kw_pvalues(kw_all, kw_consensus, regions_order, strategy_ids, strategy_labels):
    base = build_kw_pvalue_export(kw_all, kw_consensus, regions_order, strategy_ids)
    label_map = _strategy_display_names(strategy_ids, strategy_labels)
    out = pd.DataFrame({'Region': base['region']})
    for sid in strategy_ids:
        label = label_map[sid]
        if f'p_{sid}' in base.columns:
            out[f'Corrected p-value ({label})'] = base[f'p_{sid}']
        if f'star_{sid}' in base.columns:
            out[f'Significance ({label})'] = base[f'star_{sid}']
    if 'kw_p_worst' in base.columns:
        out['Worst corrected p-value across strategies'] = base['kw_p_worst']
    if 'kw_all_significant' in base.columns:
        out['Consensus significant'] = base['kw_all_significant'].map(_scientific_bool)
    if 'kw_consensus_star' in base.columns:
        out['Consensus significance'] = base['kw_consensus_star']
    return out


def build_supplementary_pairwise_pvalues(pairwise_all, pairwise_consensus, regions_order, strategy_ids, strategy_labels):
    base = build_pairwise_pvalue_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids)
    label_map = _strategy_display_names(strategy_ids, strategy_labels)
    out = pd.DataFrame({'Region': base['region'].astype(str), 'Comparison': base['comparison'].astype(str)})
    for sid in strategy_ids:
        label = label_map[sid]
        if f'p_{sid}' in base.columns:
            out[f'Corrected p-value ({label})'] = base[f'p_{sid}']
        if f'star_{sid}' in base.columns:
            out[f'Significance ({label})'] = base[f'star_{sid}']
    if 'pair_p_worst' in base.columns:
        out['Worst corrected p-value across strategies'] = base['pair_p_worst']
    if 'kw_all_significant' in base.columns:
        out['KW consensus significant'] = base['kw_all_significant'].map(_scientific_bool)
    if 'pair_pvalue_consensus_pass' in base.columns:
        out['Pairwise consensus significant'] = base['pair_pvalue_consensus_pass'].map(_scientific_bool)
    if 'pair_pvalue_consensus_star' in base.columns:
        out['Pairwise consensus significance'] = base['pair_pvalue_consensus_star']
    return out


def build_supplementary_kw_effect_sizes(kw_all, kw_consensus, regions_order, strategy_ids, strategy_labels):
    base = build_kw_effect_export(kw_all, kw_consensus, regions_order, strategy_ids)
    label_map = _strategy_display_names(strategy_ids, strategy_labels)
    out = pd.DataFrame({'Region': base['region']})
    for sid in strategy_ids:
        label = label_map[sid]
        if f'epsilon_squared_R_{sid}' in base.columns:
            out[f'Epsilon-squared ({label})'] = base[f'epsilon_squared_R_{sid}']
        if f'eta_squared_H_{sid}' in base.columns:
            out[f'Eta-squared ({label})'] = base[f'eta_squared_H_{sid}']
        if f'effect_size_magnitude_{sid}' in base.columns:
            out[f'Effect-size magnitude ({label})'] = base[f'effect_size_magnitude_{sid}']
    if 'kw_epsilon_min' in base.columns:
        out['Minimum epsilon-squared across strategies'] = base['kw_epsilon_min']
    if 'kw_epsilon_median' in base.columns:
        out['Median epsilon-squared across strategies'] = base['kw_epsilon_median']
    if 'kw_epsilon_consensus_stars' in base.columns:
        out['Consensus effect-size stars'] = base['kw_epsilon_consensus_stars']
    if 'kw_all_significant' in base.columns:
        out['KW consensus significant'] = base['kw_all_significant'].map(_scientific_bool)
    return out


def build_supplementary_pairwise_effect_sizes(pairwise_all, pairwise_consensus, regions_order, strategy_ids, strategy_labels):
    base = build_pairwise_effect_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids)
    label_map = _strategy_display_names(strategy_ids, strategy_labels)
    out = pd.DataFrame({'Region': base['region'].astype(str), 'Comparison': base['comparison'].astype(str)})
    for sid in strategy_ids:
        label = label_map[sid]
        for raw_col, nice in [
            ('rank_biserial', 'Rank-biserial correlation'),
            ('cliffs_delta', "Cliff's delta"),
            ('VDA', 'Vargha-Delaney A'),
            ('VDA_folded', 'Folded Vargha-Delaney A'),
            ('effect_size_magnitude', 'Effect-size magnitude'),
            ('dominant_serotype', 'Dominant serotype'),
        ]:
            col = f'{raw_col}_{sid}'
            if col in base.columns:
                out[f'{nice} ({label})'] = base[col]
    for col, nice in [
        ('pair_effect_min_abs', 'Weakest absolute pairwise effect across strategies'),
        ('pair_effect_median', 'Median signed pairwise effect across strategies'),
        ('pair_effect_worst_signed', 'Weakest signed pairwise effect across strategies'),
    ]:
        if col in base.columns:
            out[nice] = base[col]
    if 'sign_concordant' in base.columns:
        out['Effect direction concordant across strategies'] = base['sign_concordant'].map(_scientific_bool)
    if 'pair_effect_consensus_pass' in base.columns:
        out['Effect-size consensus pass'] = base['pair_effect_consensus_pass'].map(_scientific_bool)
    if 'pair_effect_consensus_stars' in base.columns:
        out['Consensus effect-size stars'] = base['pair_effect_consensus_stars']
    if 'pair_effect_magnitude' in base.columns:
        out['Consensus effect-size magnitude'] = base['pair_effect_magnitude']
    if 'pair_pvalue_consensus_pass' in base.columns:
        out['Pairwise p-value consensus significant'] = base['pair_pvalue_consensus_pass'].map(_scientific_bool)
    return out


def build_supplementary_counts(df_all, strategy_ids, strategy_labels):
    label_map = _strategy_display_names(strategy_ids, strategy_labels)
    counts = df_all.groupby(['strategy', 'region', 'serotype']).size().reset_index(name='Number of samples')
    counts['Split strategy'] = counts['strategy'].map(label_map).fillna(counts['strategy'])
    counts = counts.rename(columns={'region': 'Region', 'serotype': 'Serotype'})
    return counts[['Split strategy', 'Region', 'Serotype', 'Number of samples']]


def _format_excel_workbook(writer):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    header_font = Font(bold=True)
    border = Border(bottom=Side(style='thin', color='BFBFBF'))
    for ws in writer.book.worksheets:
        ws.freeze_panes = 'A2'
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '0.000E+00' if (abs(cell.value) < 0.001 and cell.value != 0) else '0.0000'
                cell.alignment = Alignment(wrap_text=False, vertical='center')


def export_supplementary_excel(
    df_all,
    kw_all,
    pairwise_all,
    kw_consensus,
    pairwise_consensus,
    regions_order,
    strategy_ids,
    strategy_labels,
    out_dir,
    value_col='attention',
    alpha=0.05,
):
    """Export a paper-supplement-ready workbook without internal paths or diagnostics."""
    xlsx_path = os.path.join(out_dir, f'supplementary_{value_col}_strategy_consensus.xlsx')
    sheets = {
        'README': _make_readme_table(value_col, strategy_ids, strategy_labels, alpha),
        'KW p-values': build_supplementary_kw_pvalues(kw_all, kw_consensus, regions_order, strategy_ids, strategy_labels),
        'Pairwise p-values': build_supplementary_pairwise_pvalues(pairwise_all, pairwise_consensus, regions_order, strategy_ids, strategy_labels),
        'KW effect sizes': build_supplementary_kw_effect_sizes(kw_all, kw_consensus, regions_order, strategy_ids, strategy_labels),
        'Pairwise effect sizes': build_supplementary_pairwise_effect_sizes(pairwise_all, pairwise_consensus, regions_order, strategy_ids, strategy_labels),
        'Sample counts': build_supplementary_counts(df_all, strategy_ids, strategy_labels),
    }
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for sheet_name, table in sheets.items():
            table.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            _autosize_excel_columns(writer, sheet_name[:31], table)
        _format_excel_workbook(writer)
    print(f'Saved supplementary workbook -> {xlsx_path}')
    return xlsx_path

def export_consensus_excel(
    df_all,
    kw_all,
    pairwise_all,
    kw_consensus,
    pairwise_consensus,
    inputs_summary,
    regions_order,
    strategy_ids,
    out_dir,
    value_col='attention',
):
    """Export one workbook collecting p-values and effect sizes for all strategies.

    The workbook is intentionally organized into separate tabs for p-values and
    effect sizes, so it can be used directly when reporting both the three
    strategy-specific statistics and the conservative strategy-consensus summary.
    """
    xlsx_path = os.path.join(out_dir, f'{value_col}_strategy_consensus_raw.xlsx')
    sheets = {
        'kw_pvalues': build_kw_pvalue_export(kw_all, kw_consensus, regions_order, strategy_ids),
        'pairwise_pvalues': build_pairwise_pvalue_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids),
        'kw_effect_sizes': build_kw_effect_export(kw_all, kw_consensus, regions_order, strategy_ids),
        'pairwise_effect_sizes': build_pairwise_effect_export(pairwise_all, pairwise_consensus, regions_order, strategy_ids),
        'kw_consensus_raw': kw_consensus,
        'pairwise_consensus_raw': pairwise_consensus,
        'inputs': inputs_summary,
        'counts_by_strategy_region': df_all.groupby(['strategy', 'strategy_label', 'region', 'serotype']).size().reset_index(name='n'),
    }
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        for sheet_name, table in sheets.items():
            # Excel sheet names are capped at 31 characters.
            safe_name = sheet_name[:31]
            table.to_excel(writer, sheet_name=safe_name, index=False)
            _autosize_excel_columns(writer, safe_name, table)
    print(f'Saved -> {xlsx_path}')
    return xlsx_path


def main():
    args = parse_args()
    print('\nArguments:')
    for key, value in vars(args).items():
        print(f'  {key}: {value}')

    k_type = 'ohe' if args.one_hot else f'k{args.k}'
    strategy_ids = list(args.strategy_ids)
    strategy_labels = list(args.strategy_labels)
    strategies = []
    for strategy_id, label, split_key, run_name in zip(
        strategy_ids,
        args.strategy_labels,
        args.strategy_split_keys,
        args.strategy_run_names,
    ):
        strategies.append({
            'id': strategy_id,
            'label': label,
            'split_key': split_key,
            'run_name': run_name,
            'value_col': args.value_col,
        })

    df_parts = []
    kw_parts = []
    pairwise_parts = []
    region_orders = []
    input_rows = []

    for strategy in strategies:
        box_dataset_dir, candidates = _find_box_dataset_dir(args, k_type, strategy)
        if box_dataset_dir is None:
            raise FileNotFoundError(
                f'Could not find output-box-dataset for strategy {strategy["id"]!r}. Tried:\n'
                + '\n'.join(candidates)
            )
        print(f'Input [{strategy["id"]}] -> {box_dataset_dir}')
        df, kw, pairwise, regions_order = _load_strategy_inputs(box_dataset_dir, strategy)
        df_parts.append(df)
        kw_parts.append(kw)
        pairwise_parts.append(pairwise)
        region_orders.append(regions_order)
        input_rows.append({
            'strategy': strategy['id'],
            'strategy_label': strategy['label'],
            'split_key': strategy['split_key'],
            'run_name': strategy['run_name'],
            'box_dataset_dir': box_dataset_dir,
            f'n_{args.value_col}_rows': len(df),
            'n_regions': len(regions_order),
        })

    regions_order = _merge_region_orders(region_orders)
    df_all = pd.concat(df_parts, axis=0, ignore_index=True)
    kw_all = pd.concat(kw_parts, axis=0, ignore_index=True)
    pairwise_all = pd.concat(pairwise_parts, axis=0, ignore_index=True)

    kw_consensus = build_kw_consensus(kw_all, regions_order, strategy_ids, alpha=args.alpha)
    pairwise_consensus = build_pairwise_consensus(
        pairwise_all,
        kw_consensus,
        regions_order,
        strategy_ids,
        alpha=args.alpha,
        require_kw=not args.no_require_kw_for_pairwise,
        require_effect_sign=not args.no_require_effect_sign,
    )
    _validate_pvalue_star_consistency(
        kw_consensus,
        pairwise_consensus,
        strategy_ids,
        require_kw=not args.no_require_kw_for_pairwise,
    )

    # Keep the consensus outputs directly under the model-level stat-panel folder:
    #   aggregate_xai/attention_stat_panel/denformer_<pooling>/strategy_consensus/
    # without an additional run-name subfolder.
    model_stat_panel_dir = os.path.dirname(build_xai_output_dir(
        paths.logs_dir,
        args.output_kind,
        'denformer',
        args.pooling,
        k_type,
        f'_{args.out_name}',
        args.epochs,
    ))
    out_dir = os.path.join(model_stat_panel_dir, 'strategy_consensus')
    os.makedirs(out_dir, exist_ok=True)
    print(f'Output strategy consensus dir: {out_dir}')

    df_all_path = os.path.join(out_dir, f'{args.value_col}_by_region_long_all_strategies.csv')
    kw_all_path = os.path.join(out_dir, 'kruskal_by_region_all_strategies.csv')
    pair_all_path = os.path.join(out_dir, 'pairwise_serotype_stats_by_region_all_strategies.csv')
    kw_consensus_path = os.path.join(out_dir, 'kruskal_by_region_strategy_consensus.csv')
    pair_consensus_path = os.path.join(out_dir, 'pairwise_serotype_strategy_consensus.csv')

    df_all.to_csv(df_all_path, index=False)
    kw_all.to_csv(kw_all_path, index=False)
    pairwise_all.to_csv(pair_all_path, index=False)
    kw_consensus.to_csv(kw_consensus_path, index=False)
    pairwise_consensus.to_csv(pair_consensus_path, index=False)
    print(f'Saved -> {df_all_path}')
    print(f'Saved -> {kw_all_path}')
    print(f'Saved -> {pair_all_path}')
    print(f'Saved -> {kw_consensus_path}')
    print(f'Saved -> {pair_consensus_path}')

    score_label = 'attention' if args.value_col == 'attention' else args.value_col.upper()
    title_prefix = 'split-strategy consensus | denformer ' + args.pooling + f' | raw {score_label} regional statistics'

    fig = plot_consensus_pvalue_panel(
        df_all,
        kw_consensus,
        pairwise_consensus,
        regions_order,
        strategy_ids,
        strategy_labels,
        title=title_prefix + ' | p-values',
        value_col=args.value_col,
    )
    for ext in ['png', 'pdf']:
        path = os.path.join(out_dir, f'{args.value_col}_region_strategy_consensus_pvalues_raw.{ext}')
        fig.savefig(path, dpi=300, bbox_inches='tight')
        print(f'Saved -> {path}')
    plt.close(fig)

    fig = plot_consensus_pvalue_panel(
        df_all,
        kw_consensus,
        pairwise_consensus,
        regions_order,
        strategy_ids,
        strategy_labels,
        title=title_prefix + ' | p-values | split-strategy stars',
        star_mode='strategy',
        value_col=args.value_col,
    )
    for ext in ['png', 'pdf']:
        path = os.path.join(out_dir, f'{args.value_col}_region_strategy_consensus_pvalues_raw_strategy_stars.{ext}')
        fig.savefig(path, dpi=300, bbox_inches='tight')
        print(f'Saved -> {path}')
    plt.close(fig)

    if 'kw_epsilon_min' in kw_consensus.columns and pairwise_consensus['pair_effect_col'].astype(str).str.len().max() > 0:
        fig = plot_consensus_effect_panel(
            df_all,
            kw_consensus,
            pairwise_consensus,
            regions_order,
            strategy_ids,
            strategy_labels,
            title=title_prefix + ' | effect sizes',
            value_col=args.value_col,
        )
        for ext in ['png', 'pdf']:
            path = os.path.join(out_dir, f'{args.value_col}_region_strategy_consensus_effect_size_raw.{ext}')
            fig.savefig(path, dpi=300, bbox_inches='tight')
            print(f'Saved -> {path}')
        plt.close(fig)

    inputs_summary = pd.DataFrame(input_rows)
    excel_path = export_consensus_excel(
        df_all,
        kw_all,
        pairwise_all,
        kw_consensus,
        pairwise_consensus,
        inputs_summary,
        regions_order,
        strategy_ids,
        out_dir,
        value_col=args.value_col,
    )
    supplementary_excel_path = export_supplementary_excel(
        df_all,
        kw_all,
        pairwise_all,
        kw_consensus,
        pairwise_consensus,
        regions_order,
        strategy_ids,
        strategy_labels,
        out_dir,
        value_col=args.value_col,
        alpha=args.alpha,
    )

    summary_path = os.path.join(out_dir, f'{args.value_col}_strategy_consensus_summary.csv')
    pd.DataFrame([{
        'output_dir': out_dir,
        'excel_path': excel_path,
        'supplementary_excel_path': supplementary_excel_path,
        f'n_{args.value_col}_rows': len(df_all),
        'n_regions': len(regions_order),
        'strategies': ';'.join(strategy_ids),
        'alpha': args.alpha,
        'pvalue_rule': 'KW consensus: all strategy KW p-values < alpha. Pairwise p-value consensus: KW consensus region and all strategy pairwise p-values < alpha. Stars use max/worst p-value.',
        'effect_size_rule': 'Effect-size panel uses weakest absolute effect across strategies and additionally requires concordant pairwise effect direction unless disabled.',
        'pairwise_pvalue_requires_kw_consensus': not args.no_require_kw_for_pairwise,
        'pairwise_effect_requires_sign_concordance': not args.no_require_effect_sign,
    }]).to_csv(summary_path, index=False)
    print(f'Saved -> {summary_path}')


if __name__ == '__main__':
    main()

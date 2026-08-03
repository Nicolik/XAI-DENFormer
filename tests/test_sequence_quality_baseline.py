from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from classifier.workflow.reports.make_sequence_quality_baseline import (
    ModelSpec,
    aggregate_predictions,
    align_features_to_canonical_order,
    build_supplementary_table,
    write_supplementary_table_workbook,
    write_outputs,
    evaluate_protocol,
    extract_feature_table,
    folds_from_main_loader,
    sequence_features,
    split_qc,
)


def test_sequence_features_counts_categories():
    values = sequence_features("ACGTNNRY-.")
    assert values["sequence_length"] == 10
    assert values["acgt_count"] == 4
    assert values["n_count"] == 2
    assert values["iupac_count"] == 2
    assert values["gap_count"] == 2
    assert values["non_acgt_count"] == 6
    assert np.isclose(values["gc_fraction"], 0.5)


def test_extract_feature_table_preserves_global_fasta_order(tmp_path: Path):
    fasta_map = {
        "DENV1_merged_meta.fasta": "DENV1",
        "DENV2_merged_meta.fasta": "DENV2",
        "DENV3_merged_meta.fasta": "DENV3",
        "DENV4_merged_meta.fasta": "DENV4",
    }
    for idx, (filename, serotype) in enumerate(fasta_map.items()):
        (tmp_path / filename).write_text(
            f">{serotype}_a\n{'A' * (10 + idx)}\n>{serotype}_b\n{'C' * (11 + idx)}\n",
            encoding="utf-8",
        )

    features = extract_feature_table(tmp_path)
    assert features["source_index"].tolist() == list(range(8))
    assert features["fasta_serotype"].tolist()[:2] == ["DENV1", "DENV1"]
    assert features["padding_length"].min() == 0


def test_evaluate_protocol_uses_main_fold_structure():
    rows = []
    for index in range(16):
        label = index % 4
        rows.append(
            {
                "index": index,
                "sequence_id": f"seq_{index}",
                "serotype": f"DENV{label + 1}",
                "label": label,
                "sequence_length": 10000 + label * 10,
                "padding_fraction": 0.01 * label,
                "n_fraction": 0.001 * label,
                "iupac_fraction": 0.0,
                "gap_fraction": 0.0,
                "other_fraction": 0.0,
                "non_acgt_fraction": 0.001 * label,
                "gc_fraction": 0.4 + 0.02 * label,
            }
        )
    features = pd.DataFrame(rows)
    folds = [
        {
            "fold_id": "fold_1",
            "train_idx": np.arange(12, dtype=np.int64),
            "val_idx": np.array([], dtype=np.int64),
            "test_idx": np.arange(12, 16, dtype=np.int64),
        }
    ]
    specs = [
        ModelSpec("majority", tuple(), "dummy"),
        ModelSpec("length_logreg", ("sequence_length",), "logreg"),
    ]

    fold_metrics, predictions, coefficients = evaluate_protocol(
        "synthetic", folds, features, specs
    )
    assert len(fold_metrics) == 2
    assert len(predictions) == 8
    assert not coefficients.empty

    aggregate, per_class, confusion = aggregate_predictions(predictions)
    assert len(aggregate) == 2
    assert len(per_class) == 8
    assert len(confusion) == 32


def test_align_features_to_saved_ohe_order_and_targets(tmp_path: Path):
    features = pd.DataFrame(
        [
            {"source_index": 0, "sequence_id": "b", "fasta_serotype": "DENV2", "serotype": "DENV2", "label": 1},
            {"source_index": 1, "sequence_id": "a", "fasta_serotype": "DENV1", "serotype": "DENV1", "label": 0},
            {"source_index": 2, "sequence_id": "dup", "fasta_serotype": "DENV2", "serotype": "DENV2", "label": 1},
            {"source_index": 3, "sequence_id": "dup", "fasta_serotype": "DENV1", "serotype": "DENV1", "label": 0},
        ]
    )
    ids = tmp_path / "seq_ids_ohe.txt"
    labels = tmp_path / "label_matrix.txt"
    ids.write_text("a\ndup\nb\ndup\n", encoding="utf-8")
    labels.write_text("0\n0\n1\n1\n", encoding="utf-8")

    aligned = align_features_to_canonical_order(features, ids, labels)
    assert aligned["index"].tolist() == [0, 1, 2, 3]
    assert aligned["sequence_id"].tolist() == ["a", "dup", "b", "dup"]
    assert aligned["label"].tolist() == [0, 0, 1, 1]
    assert aligned["serotype"].tolist() == ["DENV1", "DENV1", "DENV2", "DENV2"]


def test_align_features_requires_canonical_files(tmp_path: Path):
    features = pd.DataFrame(
        [{"source_index": 0, "sequence_id": "a", "serotype": "DENV1", "label": 0}]
    )
    with pytest.raises(FileNotFoundError, match="sequence-ID order"):
        align_features_to_canonical_order(
            features, tmp_path / "missing_seq_ids.txt", tmp_path / "missing_labels.txt"
        )


def test_folds_from_main_loader_ignores_descriptive_columns(tmp_path: Path):
    split_path = tmp_path / "split.csv"
    pd.DataFrame(
        [
            {"fold": "A", "index": 0, "split": "train", "serotype": "WRONG", "sequence_id": "wrong"},
            {"fold": "A", "index": 1, "split": "train", "serotype": "WRONG", "sequence_id": "wrong"},
            {"fold": "A", "index": 2, "split": "val", "serotype": "WRONG", "sequence_id": "wrong"},
            {"fold": "A", "index": 3, "split": "test", "serotype": "WRONG", "sequence_id": "wrong"},
        ]
    ).to_csv(split_path, index=False)

    folds = folds_from_main_loader(split_path, 4)
    assert len(folds) == 1
    assert folds[0]["train_idx"].tolist() == [0, 1]
    assert folds[0]["val_idx"].tolist() == [2]
    assert folds[0]["test_idx"].tolist() == [3]

    qc = split_qc("synthetic", folds, 4)
    assert qc["target_label_source"] == "label_matrix.txt"
    assert qc["split_loader"] == "classifier.workflow.utils.load_and_validate_folds"


def _synthetic_aggregate_metrics() -> pd.DataFrame:
    models = [
        "majority",
        "length_logreg",
        "n_logreg",
        "gc_logreg",
        "length_gc_logreg",
        "length_n_gc_logreg",
        "quality_logreg",
        "quality_histgb",
    ]
    protocols = ["continent", "timebin", "cdhit"]
    rows = []
    for model_index, model in enumerate(models):
        for protocol_index, protocol in enumerate(protocols):
            value = 0.10 + 0.01 * model_index + 0.001 * protocol_index
            rows.append(
                {
                    "protocol": protocol,
                    "model": model,
                    "accuracy": value,
                    "balanced_accuracy": value + 0.10,
                    "macro_f1_all4": value + 0.20,
                }
            )
    return pd.DataFrame(rows)


def test_supplementary_table_has_atomic_metric_values():
    table = build_supplementary_table(_synthetic_aggregate_metrics())
    assert table.columns.tolist() == [
        "Baseline",
        "Metric",
        "Geographical",
        "Temporal",
        "CD-HIT",
    ]
    assert len(table) == 24
    assert table.iloc[:3]["Baseline"].tolist() == ["Majority class"] * 3
    assert table.iloc[:3]["Metric"].tolist() == [
        "Accuracy",
        "Balanced accuracy",
        "Macro-F1",
    ]
    assert table.iloc[0]["Geographical"] == 10.0
    assert table.iloc[1]["Geographical"] == 20.0
    assert table.iloc[2]["Geographical"] == 30.0
    assert isinstance(table.iloc[0]["Geographical"], float)
    assert table.iloc[-1]["Baseline"] == "All quality features (HGB)"


def test_supplementary_workbook_contains_only_table(tmp_path: Path):
    from openpyxl import load_workbook

    output = tmp_path / "table.xlsx"
    write_supplementary_table_workbook(output, _synthetic_aggregate_metrics())
    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Table_S1"]

    assert sheet.max_row == 25
    assert sheet.max_column == 5
    assert [sheet.cell(1, col).value for col in range(1, 6)] == [
        "Baseline",
        "Metric",
        "Geographical",
        "Temporal",
        "CD-HIT",
    ]
    assert sheet["C2"].value == 10.0
    assert sheet["C3"].value == 20.0
    assert sheet["C4"].value == 30.0
    assert "A2:A4" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    all_values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
    assert not any("Supplementary Table" in str(value) for value in all_values)
    assert not any("Values are" in str(value) for value in all_values)
    assert not any(" / " in str(value) for value in all_values)


def _assert_helvetica_8(workbook) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                assert cell.font.name == "Helvetica"
                assert cell.font.sz == 8


def test_all_generated_workbooks_use_helvetica_8(tmp_path: Path):
    from openpyxl import load_workbook

    features = pd.DataFrame(
        [
            {
                "serotype": "DENV1",
                "sequence_length": 10000,
                "padding_fraction": 0.01,
                "n_fraction": 0.001,
                "iupac_fraction": 0.0,
                "gap_fraction": 0.0,
                "other_fraction": 0.0,
                "non_acgt_fraction": 0.001,
                "gc_fraction": 0.45,
            },
            {
                "serotype": "DENV2",
                "sequence_length": 10010,
                "padding_fraction": 0.0,
                "n_fraction": 0.002,
                "iupac_fraction": 0.0,
                "gap_fraction": 0.0,
                "other_fraction": 0.0,
                "non_acgt_fraction": 0.002,
                "gc_fraction": 0.47,
            },
        ]
    )
    simple = pd.DataFrame([{"protocol": "continent", "value": 1.0}])
    predictions = pd.DataFrame([{"index": 0, "prediction": 0}])
    coefficients = pd.DataFrame([{"protocol": "continent", "coefficient": 0.1}])
    aggregate = _synthetic_aggregate_metrics()
    per_class = pd.DataFrame([{"protocol": "continent", "class": "DENV1", "f1": 1.0}])
    confusion = pd.DataFrame([{"protocol": "continent", "true": 0, "pred": 0, "n": 1}])
    qc = pd.DataFrame([{"protocol": "continent", "status": "PASS"}])

    write_outputs(
        tmp_path,
        features,
        simple,
        predictions,
        coefficients,
        aggregate,
        per_class,
        confusion,
        qc,
    )

    _assert_helvetica_8(load_workbook(tmp_path / "Sequence_Quality_Baseline.xlsx"))
    _assert_helvetica_8(
        load_workbook(tmp_path / "Supplementary_Table_Sequence_Quality_Baselines.xlsx")
    )

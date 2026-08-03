import tempfile
import unittest
from unittest.mock import patch
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from classifier.workflow.reports.export_performance_statistics_workbook import export_workbook
from classifier.workflow.reports.make_performance_statistics import (
    _write_csv_outputs,
    _write_workbook,
)

from classifier.workflow.reports.performance_statistics import (
    AnalysisConfig,
    ABLATION_PAIRS,
    BENCHMARK_PAIRS,
    DEFAULT_MODELS,
    DEFAULT_PROTOCOLS,
    exact_mcnemar,
    holm_adjust,
    run_analysis,
)


class PerformanceStatisticsTest(unittest.TestCase):
    def test_holm_is_monotone_and_bounded(self):
        adjusted = holm_adjust([0.01, 0.04, 0.03])
        self.assertTrue(np.all(adjusted >= np.array([0.01, 0.04, 0.03])))
        self.assertTrue(np.all(adjusted <= 1.0))

    def test_exact_mcnemar(self):
        a = np.array([True, True, True, False, False])
        b = np.array([False, False, True, True, False])
        result = exact_mcnemar(a, b)
        self.assertEqual(result["a_correct_b_wrong"], 2)
        self.assertEqual(result["a_wrong_b_correct"], 1)
        self.assertEqual(result["discordant_total"], 3)

    def test_end_to_end_synthetic(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            logs = root / "logs"
            split_path = root / "cdhit_splits.csv"

            n = 48
            indices = np.arange(n)
            labels = np.repeat(np.arange(4), n // 4)
            cluster_ids = np.arange(n) // 2
            split_rows = []
            for fold in range(2):
                for idx in indices:
                    split_rows.append(
                        {
                            "fold": f"fold_{fold + 1}",
                            "index": int(idx),
                            "split": "test" if idx % 2 == fold else "train",
                            "serotype": int(labels[idx] + 1),
                            "cluster_id": int(cluster_ids[idx]),
                        }
                    )
            pd.DataFrame(split_rows).to_csv(split_path, index=False)

            rng = np.random.default_rng(7)
            model_error_rates = {
                "denformer_mean": 0.03,
                "denformer_first": 0.08,
                "denformer_max": 0.06,
                "longformer": 0.05,
                "performer": 0.06,
                "ffnn": 0.10,
                "logreg": 0.12,
            }
            for protocol in DEFAULT_PROTOCOLS:
                for model in DEFAULT_MODELS:
                    preds = labels.copy()
                    error_mask = rng.random(n) < model_error_rates[model]
                    preds[error_mask] = (preds[error_mask] + 1) % 4
                    for fold in range(2):
                        test_idx = indices[indices % 2 == fold]
                        out_dir = logs / model / protocol / "metrics" / f"split_fold_{fold + 1}"
                        out_dir.mkdir(parents=True, exist_ok=True)
                        # Legacy prediction files did not store global indices.
                        np.savez(
                            out_dir / "predictions_test.npz",
                            labels=labels[test_idx],
                            preds=preds[test_idx],
                            probs=np.eye(4)[preds[test_idx]],
                        )

            outputs = run_analysis(
                logs_dir=logs,
                cdhit_split_file=split_path,
                config=AnalysisConfig(
                    bootstrap_reps=50,
                    permutation_reps=100,
                    seed=3,
                    bootstrap_batch_size=16,
                    permutation_batch_size=32,
                ),
                protocol_split_files={protocol: split_path for protocol in DEFAULT_PROTOCOLS},
            )
            expected_pairs = len(DEFAULT_PROTOCOLS) * (len(BENCHMARK_PAIRS) + len(ABLATION_PAIRS))
            self.assertEqual(len(outputs.primary_accuracy), expected_pairs)
            self.assertEqual(len(outputs.macro_f1), expected_pairs)
            self.assertEqual(len(outputs.per_class_f1), expected_pairs * 4)
            self.assertTrue((outputs.qc_alignment["alignment_status"] == "PASS").all())

            workbook_path = root / "Supplementary_File_S3_Model_Performance_Statistics.xlsx"
            _write_workbook(outputs, workbook_path)
            workbook = load_workbook(workbook_path, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "README",
                    "Benchmark_primary",
                    "Ablation_primary",
                    "Accuracy_full",
                    "Macro_F1",
                    "Per_class_F1",
                    "Model_summary",
                    "QC_alignment",
                    "Methods_parameters",
                ],
            )
            self.assertNotIn("Paper_table", workbook.sheetnames)
            self.assertNotIn("Table_S3", workbook.sheetnames)

            cached_dir = root / "cached_statistics"
            _write_csv_outputs(outputs, cached_dir)
            parameters_csv = cached_dir / "csv" / "methods_parameters.csv"
            parameters_csv.unlink()

            workbook_from_cache = cached_dir / "reexported.xlsx"
            with patch.object(pd, "read_excel", side_effect=AssertionError("must not be called")):
                export_workbook(
                    output_dir=cached_dir,
                    workbook_path=workbook_from_cache,
                    source_workbook=workbook_path,
                )
            cached_workbook = load_workbook(workbook_from_cache, read_only=True)
            self.assertEqual(cached_workbook.sheetnames, workbook.sheetnames)
            self.assertTrue(parameters_csv.is_file())

            second_workbook = cached_dir / "reexported_from_csv_only.xlsx"
            export_workbook(
                output_dir=cached_dir,
                workbook_path=second_workbook,
                source_workbook=root / "missing.xlsx",
            )
            self.assertTrue(second_workbook.is_file())


if __name__ == "__main__":
    unittest.main()
